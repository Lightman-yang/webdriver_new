# _*_ coding: UTF-8 _*_
# 开发人员  :light
# 开发时间  :2022/9/3 16:31
import openpyxl


# 加载本地数据
def excelboot01(nn):
    zzzzz1 = []
    # data =openpyxl.load_workbook(wu)
    data = openpyxl.load_workbook(r"D:\webdriver_new\lw\game_name.xlsx")
    # 获取工作表 有三种方法
    zz1 = data.active  # 不知道表名称的 用这种
    # zz1=data['Sheet'] #知道表面的用第二种
    # zz1=data.get_sheet_by_name('Sheet1') #第三种不知道和第二种有什么区别
    # 获取表A1数据
    # z_max=zz1.max_row #获取最大行数
    # z_min=zz1.min_row #获取最小行数
    # c_min=zz1.max_column #获取最大数列
    # c_min = zz1.min_column  # 获取最小数列
    # cell.coordinate 定位数据
    # for i in zz1['A2':'H2' ]:
    #     print(len(i))
    # 获取行数
    max_row_num = zz1.max_row
    row_list2 = []
    for i in range(nn + 1, max_row_num + 1):

        for row in zz1[i]:
            row_list2.append(row.value)

        # print(i[1].value,i[2].value,i[3].value,i[4].value,i[5].value,i[6].value)
        # for cell in i:
        #
        #      zzzzz1.append(cell.value)
        # print(cell.value,'1')
        # print('cell',cell.value,cell.coordinate)
        # print(zz1.min_row+1,zz1.max_row+1)
        # zz = zz1['A{}'.format(i)]
        #     zzz = cell.value
        #     print(zzz)
        #     zz_list = list(zzz.split(",")) # 用','分割并转换list数据
        #     zzzzz=zz_list[0:8]
        # # #print(type(zz.value),type(zz_list),zz_list[0:6])
        # # print(zz_list[0:6],'第一行%s' %i)
        #     zzzzz1.append(zzzzz)#添加到zzzzz1变量中
        # print(zz_list,"最大行为'%s',第一行数据为'%s'" %  (zz1.min_row,zz1.max_row))
    # print(row_list2)
    # zz.close
    # print(zzzzz)

    print(row_list2[0:7], type(row_list2[7]))
    return row_list2[2:8]


if __name__ == '__main__':
    aa = excelboot01(1)
    print(aa)

    #
    # #c=[1,2,3,4,5,'6']
    # #print(aaa_arr)
    # #print(c)
    # print(aaa_arr[1][0:]) #获取元素
