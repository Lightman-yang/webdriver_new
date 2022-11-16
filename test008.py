#_*_ coding: UTF-8 _*_
#开发人员  :light
#开发时间  :2022/9/4 9:59
import openpyxl

import test005
ZZ=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33]
a=[9, 10, 12, 18, 29, 32]
# zz=test005.JiaFa(a)
# jj=test005.JianFa(a)
# print('减',"=",zz,"加","=",jj)

# zzzzz1 = []
# data =openpyxl.load_workbook(wu)
data1 = openpyxl.load_workbook(r"D:\TEST001\test.xlsx")
data=data1['Sheet2']
# 获取活动表
#zz1 = data.active
data.cell(row=3,column=2,value="%s" % a)#row=1,(行1) colum=2（2列）
data1.save("test.xlsx")
data1.close()