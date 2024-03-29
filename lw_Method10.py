import os
import random
import smtplib
import time
import traceback
# import pandas as pd
from email.header import Header
from email.mime.text import MIMEText
from random import uniform
from threading import Thread, RLock  # 导入线程函数
from time import sleep  # 导入时间休眠函数

import numpy as np
import openpyxl
import pyautogui
import pydirectinput  as dt
from PIL import Image
from comtypes.client import CreateObject
from win32gui import FindWindow

from python_findpicture import Caozuolei1


# 继承Caozuolei1函数。
class Caozuolei(Caozuolei1):
    time.sleep(0.5)
    mutex1 = RLock()
    mutex2 = RLock()

    # # 绑定窗口句柄
    # # 如果函数运行期间想要停止，请把鼠标移动到屏幕得左上角（0，0）位置，
    # # 这触发pyautogui.FaailSafeException异常，从而终止程序运行。
    # dt.FAILSAFE = True  # 默认True则鼠标(0,0)可触发异常；False不触发
    #
    # hwnd1 = win32gui.WindowFromPoint((308, 444))  # 请填写 x 和 y 坐标
    # print(hwnd1)
    # time.sleep(1)
    #
    # '''步骤8 通过窗口句柄 获取当前窗口的【左、上、右、下】四个方向的坐标位置'''
    # # 通过句柄值获取当前窗口的【左、上、右、下】四个方向的坐标位置
    # # left, top, right, bottom = win32gui.GetWindowRect('句柄值')
    # left, top, right, bottom = win32gui.GetWindowRect(hwnd1)
    # print(left, top, right, bottom)
    #
    # # 通过句柄将窗口放到最前
    # # win32gui.SetForegroundWindow("句柄值")
    # # dpyautogui.press('alt')
    # win32gui.SetForegroundWindow(hwnd1)
    # time.sleep(2)

    def __init__(self):
        super().__init__()
        # 当自身有构造方法时候，继承构造方法用super
        # 参数 窗口类名="地下城与勇士"
        self.hwnd = FindWindow("地下城与勇士", None)
        self.path = os.getcwd()
        try:
            self.lw = CreateObject('lw.lwsoft3')
        except OSError:
            print(self.path)
            os.system(r'regsvr32 /s %s\lw.dll' % self.path)
            self.lw = CreateObject('lw.lwsoft3')
        # 绑定窗口
        self.BindWindow()

    def BindWindow(self):
        self.lw.SetWindowState(self.hwnd, 12)  # 激活窗口
        self.lw.BindWindow(self.hwnd, 0, 0, 0, 0, 0)
        # self.lw.EnableRealMouse(2,20,30)
        self.lw.SetDict(0, "find.txt")
        # print("find")

    # 发送邮件
    def send_mails_QQsmtp(self, from_mail, to_mail, note_content, apartment_name, subject,
                          user_name, mail_host, mail_port, mymail_user, mymail_pwd):
        print(f'正在向{to_mail}发生邮件......')
        # mail msg 设置

        msg = MIMEText(note_content, 'plain', 'utf-8')
        msg['From'] = apartment_name
        msg['To'] = user_name
        msg['Subject'] = Header(subject, 'utf-8')
        with smtplib.SMTP_SSL(mail_host, mail_port) as smtObj:
            # 登陆发件人邮箱（公寓财务处邮箱）
            smtObj.login(mymail_user, mymail_pwd)
            # 发送
            smtObj.sendmail(from_mail, to_mail, msg.as_string())
            print('邮件成功发送！')

    def youjian(self, pvp=0):
        pmp = [[0, '0'],
               [1, "阳-pc1-v4"],
               [2, "qqlight-pc2-v1"],
               [3, "撒旦-pc1-v2"],
               [4, "心动依恋-pc2-v2"],
               [5, "阿斯顿-pc1-v3"],
               [6, "梦想-pc2-v32"],
               [7, "风化雪月-pc1-v1"],
               [8, "柳杨-pc2-v4"],
               [9, "超越起跑线-pc3-v1"],
               [10, "-pv4-1"], ]
        user_info = ['286224275@qq.com']

        apartment_dict = {
            'from_mail': '1074952101@qq.com',
            # 密码pwd
            'pwd': 'uewdydfgwvakghhf',
            'sub': '通知提醒',
            'apartment_name': '1074952101@qq.com'
        }
        """查询用户信息发送邮件"""
        # 使用邮件服务商提供的SMTP服务，需要设置服务器/端口号/用户名/口令（即授权码）等
        # QQmail_host, QQmail_port = 'smtp.qq.com', 465
        QQmail_host, QQmail_port = 'smtp.qq.com', 465
        QQmail_user = apartment_dict.get('from_mail')
        QQmail_pwd = apartment_dict.get('pwd')
        apartment_name = apartment_dict.get('apartment_name')
        sub = apartment_dict.get('sub')
        to_mail = user_info[0]  # 收件邮箱
        user_name = user_info[0]  # 亲爱的xxx用户xx
        # room_no = user_info[2]  # 房间号

        note_content = f'亲爱的{(pmp[pvp][1])}，你好：系统显示，游戏已经崩溃了！\
                                 \n{apartment_name}\
                                 \n崩溃时间为{time.strftime("%Y-%m-%d %H:%M-%S秒", time.localtime())}！'
        self.send_mails_QQsmtp(QQmail_user, to_mail, note_content, apartment_name, sub,
                               user_name, QQmail_host, QQmail_port, QQmail_user, QQmail_pwd)

        # print("执行过了")

    # 获取窗口客户区域的宽度和高度,会设置X,Y。
    def GetClientSize(self):
        self.lw.SetWindowState(self.hwnd, 12)  # 激活窗口
        ret = self.lw.GetClientSize(self.hwnd)
        self.width = ret[1]
        self.height = ret[2]

    # 解除绑定
    def UnBind(self):
        self.lw.UnBindWindow()
        print("解除绑定")
        # 按键key值是多少

    def KeyPress1(self, key):
        self.lw.KeyPress(key)

    # 按键操作，32代表space  1代表一次
    def KeyPress(self, num):
        for i in range(0, num):
            time.sleep(uniform(2.1, 4.9))
            s = self.lw.KeyPress(32, 1)
            print("跳", i + 1, "次")

    def MoveTo(self, x, y):
        self.lw.MoveTo(x, y)

    # 鼠标点击左键操作，理论是先进行鼠标移动，然后在进行鼠标点击左键操作
    def LeftClick(self, x, y, delay=uniform(0.01, 0.15)):
        self.lw.SetWindowState(self.hwnd, 12)  # 激活窗口
        # def LeftClick(self, x1, y1, xr, yr, delay=uniform(0.3, 0.5)):
        #     self.lw.SetWindowState(self.hwnd, 12)  # 激活窗口
        #     x = x1 + random.randint(0, xr)
        #     y = y1 + random.randint(0, yr)
        self.lw.MoveTo(x, y)
        self.lw.LeftClick()
        print("点击了鼠标左键")
        time.sleep(delay)

    def FindStrFastEx(self, x1, y1, x2, y2, string, color_format, sim, isbackcolor, n=0):
        ret = self.lw.FindStrFastEx(x1, y1, x2, y2, string, color_format, sim, isbackcolor)

        if len(ret) > 0:
            ss = ret.split("|")
            print(ss)
            num = len(ss)
            s = ss[n].split(",")
            self.x = int(s[1])
            self.y = int(s[2])
            print(len(ret), self.x, self.y, "测试")
            return 1
        else:
            return 0

    def Find_Pic(self, x1, y1, x2, y2, pic_name, delta_color, sim, dir, timeout, ischick, chickdex, chickdey,
                 chickdely):
        ret = self.lw.findpic(x1, y1, x2, y2, pic_name, delta_color, sim, dir, timeout, ischick, chickdex, chickdey,
                              chickdely)

        if ret == 1:

            print("找到图片, 横向坐标为：" + self.lw.x + "纵向坐标为：" + self.lw.y + "序号为：" + self.lw.idx)
        else:
            try:
                print("没有！")
            except OSError as de:
                print(de, 'Find_Pic')

                traceback.print_exc()

            except Exception as e:
                print(e, 'Find_Pic')

                traceback.print_exc()

    def movingFigur_Down(self, time_sleep):

        time.sleep(0.15)  # 按下两秒
        dt.keyDown('down')  # 向下移动
        time.sleep(time_sleep)  # 按下两秒
        dt.keyUp('down')  # ：模拟按键松开

    # 方法为像右移动
    def movingFigur_right(self, time_sleep):
        time.sleep(0.15)  # 像右移动
        dt.keyDown('right')  # 像右移动
        time.sleep(time_sleep)  # 像右移动
        dt.keyUp('right')  # ：模拟按键松开右

    # 方法为像上移动
    def movingFigur_up(self, time_sleep):
        time.sleep(0.15)  # 像右移动
        dt.keyDown('up')  # 像右移动
        time.sleep(time_sleep)  # 像右移动
        dt.keyUp('up')  # ：模拟按键松开右

    # 方法为像左移动
    def movingFigur_left(self, time_sleep):
        time.sleep(0.15)  # 像左移动
        dt.keyDown('left')  # 像左移动
        time.sleep(time_sleep)  # 像左移动
        dt.keyUp('left')  # ：模拟按键松开左

    def sleep_1000(self):
        while dt.press('+'):
            time.sleep(1000)
        else:
            dt.press('num9')

    def Coordinate_point(self, x, y):  # 坐标点
        # x11, y11 = left + 378, top + 452,
        x11, y11 = Caozuolei().left + x, Caozuolei().top + y,
        return x11, y11

    # 080037F840018002BFE37F8E06180C30192030406080C11F82000$行$0.0.66$15.14.15.14

    def movingfigur_Down(self, time_sleep):

        time.sleep(0.15)  # 按下两秒
        dt.keyDown('down')  # 向下移动
        time.sleep(time_sleep)  # 按下两秒
        dt.keyUp('down')  # ：模拟按键松开

    # 方法为像右移动
    def movingfigur_right(self, time_sleep):
        time.sleep(0.15)  # 像右移动
        dt.keyDown('right')  # 像右移动
        time.sleep(time_sleep)  # 像右移动
        dt.keyUp('right')  # ：模拟按键松开右

    # 方法为像上移动
    def movingfigur_up(self, time_sleep):
        time.sleep(0.15)  # 像右移动
        dt.keyDown('up')  # 像右移动
        time.sleep(time_sleep)  # 像右移动
        dt.keyUp('up')  # ：模拟按键松开右

    # 方法为像左移动
    def movingfigur_left(self, time_sleep):
        time.sleep(0.15)  # 像左移动
        dt.keyDown('left')  # 像左移动
        time.sleep(time_sleep)  # 像左移动
        dt.keyUp('left')  # ：模拟按键松开左

    # 字库
    def Set_Dict(self, index, file):

        ret = self.lw.SetDict(index, file)
        if ret == 1:
            print("自库设置成功！")

        else:
            print("自库设置失败！")

    # 切换字库
    def Use_Dict(self, index):

        ret = self.lw.UseDict(index)
        if ret == 1:
            print("切换字库{}成功！".format(index))
            return 1

        else:
            print("切换字库{}失败！".format(index))
            return 0

    def FuBen_INFO1(self):  # 副本地图信息数据实时获取
        print('self====', self)
        while True:
            # self.Use_Dict(1)
            # c.Set_Dict(0, "测试2.txt")
            aa100 = self.Find_Ocr(
                x1=0,
                y1=0,
                x2=800,
                y2=600,
                color_format="#360",
                sim=0.85,
                linesign=" ",
                isbackcolor=0)
            print("打印aa100", aa100)
            if str(aa100) is None or aa100 == 0 or str(aa100) in '崩溃':
                print('时间', time.time())
                continue
            elif "靓仔" in str(aa100):
                print(str(aa100))
                # b = time.strftime("%y-%m-%d_%H$%M$%S", time.localtime())
                #
                # img = pyautogui.screenshot(region=[48, 84, 848, 684])
                # img = Image.fromarray(np.uint8(img))
                # #img.save('D:\webdriver_new\lw\Tpshot{}s.png'.format(b))

                dt.press('9')
                time.sleep(1)  # 按下两秒
                #return
                continue
            elif "靓仔" not in str(aa100):
                print('没有靓仔')
                return
            else:

                try:
                    print()
                    return
                except OSError as de:
                    print(de, 'FuBen_INFO1')
                    return
                    traceback.print_exc()
                except Exception as e:
                    print(e, 'FuBen_INFO1')
                    return
                    traceback.print_exc()

    def FuBen_INFO2(self, oopp=0):  # 副本地图信息数据实时获取
        # self.Use_Dict(1)
        # c.Set_Dict(0, "测试2.txt")
        # a = 0
        while True:
            # time.sleep(1)
            aa = self.Find_Ocr(
                x1=0,
                y1=0,
                x2=1200,
                y2=1200,
                color_format="#360",
                sim=0.85,
                linesign=" ",
                isbackcolor=0)
            if "五陵" in str(aa):
                return 100
            elif "再次挑战" not in str(aa) and oopp == 0:
                dt.press('g')
                dt.press('d')
                # time.sleep(0.75)  # 按下19秒
                dt.press('f')
                dt.press('y')
            elif "再次挑战" not in str(aa) and oopp == 1:
                dt.press('d')
                dt.press('q')
                time.sleep(1)  # 按下两秒
                dt.press('f')
                dt.press('w')
                time.sleep(0.7)  # 按下两秒
                dt.press('w')
            elif "再次挑战" not in str(aa) and oopp == 2:
                return 1

            else:

                return

    def FuBen_INFO3(self, zzz=0):  # 副本地图信息如果自动存在，证明没活力点了，直接退出
        # self.Use_Dict(1)
        # c.Set_Dict(0, "测试2.txt") 靓丽=#70  靓仔=#360
        while True:
            # time.sleep(0.15)
            aabbcc = self.Find_Ocr(
                x1=0,
                y1=0,
                x2=800,
                y2=600,
                color_format="#360",
                sim=0.85,
                linesign=" ",
                isbackcolor=0)
            bbb = self.Find_Ocr(
                x1=0,
                y1=0,
                x2=800,
                y2=600,
                color_format="#70",
                sim=0.85,
                linesign=" ",
                isbackcolor=0)
            if ("靓仔" in str(aabbcc) and zzz == 1) or ("靓丽" in str(bbb) and zzz == 1):
                dt.press('9')
                time.sleep(1)
                ret_values.append(2)
                return 2
            elif "再次挑战" in str(aabbcc):
                ret_values.append(1)
                dt.press('f12')
                time.sleep(1)
                # dt.press('.')
                time.sleep(2)
                print("再次挑战", 1)

                return 1


            else:
                print("再次挑战", 0)
                ret_values.append(0)
                return 0

    def FuBen_INFO(self):  # 副本地图信息数据实时获取
        # self.Use_Dict(0)
        while True:
            # time.sleep(0.15)
            aam = self.Find_Ocr(
                x1=0,
                y1=0,
                x2=800,
                y2=600,
                color_format="#360",
                sim=0.99,
                linesign=" ",
                isbackcolor=0)
            print('FuBen_INFO', aam, '==aam')
            if str(aam) is None or str(aam) in '崩溃':
                print(str(aam))
                print('狗屎')
                continue
            elif "五陵" in str(aam):
                print("在副本外面")
                return 100
                # continue
            elif "第一关" in str(aam) and "开门" not in str(aam):
                print('第一关')
                return 1
                # continue
            elif "开门" in str(aam):
                print('第二关')
                return 2
                # continue
            elif "凹凸" in str(aam) and "星星" not in str(aam):
                print("在第三关凹凸")
                return 3
            elif "大鱼海棠" in str(aam) or "夜空下" in str(aam):
                print("在第四关")
                return 5
            elif "星星" in str(aam):
                print("在第5关")
                return 6
            elif "花花" in str(aam) and "空空" in str(aam):
                print("在6关")
                return 7
            elif "花花" in str(aam) and "好绿" not in str(aam):
                print("在6关")
                return 7
            elif "空空" in str(aam) and "花花" not in str(aam):
                print("在6关")
                return 77
            elif "好绿" in str(aam) or (("好绿" in str(aam)) and ("花花" in str(aam))):
                print("在7关")
                return 8
            elif "最后" in str(aam) and "再次挑战" not in str(aam):
                print("在boss房间")
                return 9
            elif "再次挑战" in str(aam):
                print("通关完成")
                return 10
            else:
                try:

                    print("0,没找到")
                    return 0
                except OSError as de:
                    print(de, 'FuBen_INFO')
                    traceback.print_exc()
                    return 0
                except Exception as e:
                    print(e, 'FuBen_INFO')

                    traceback.print_exc()
                    return 0

    def FuBen_INFO11(self):  # 进图校验是否选择“永恒之光研究所”

        # self.Use_Dict(0)
        while True:
            # time.sleep(0.15)
            aa = self.Find_Ocr(
                x1=31,
                y1=245,
                x2=260,
                y2=33,
                color_format="#331",
                sim=0.8,
                linesign=" ",
                isbackcolor=0)
            if "永恒之光" in str(aa):

                print("永恒之光")

                return 1
            else:
                dt.press('up')
                continue

    def FuBen_INFO12(self):  # 进图校验是否选择“永恒之光研究所”

        # self.Use_Dict(0)
        while True:
            # time.sleep(0.15)

            aa = self.Find_Ocr(
                x1=12,
                y1=140,
                x2=760,
                y2=600,
                color_format="#422",
                sim=0.8,
                linesign=" ",
                isbackcolor=0)
            print("aa:", str(aa))
            bb = self.Find_Ocr(
                x1=0,
                y1=0,
                x2=760,
                y2=600,
                color_format="#360",
                sim=0.79,
                linesign=" ",
                isbackcolor=0)
            print("bb:", str(bb))
            if ("无敌" in str(aa) and "快递员" not in str(aa)) and "数字" not in str(aa):
                print("无敌师很酷 是：超越起跑线  账号,9")
                return 9
            elif "冒险师" in str(aa) and "快递员" not in str(aa):
                print("力气师很大 是：曌飂飏  账号,8")
                return 8
            elif "水水" in str(aa) and "快递员" not in str(aa):
                print("色彩不足啊 是：风花雪月 账号,6")
                return 7
            elif "造就师" in str(bb) or '造就师' in str(aa):
                print("造就师,6")
                return 6
            elif "最叼" in str(bb) or "最" in str(bb):
                print("最,1")

                return 5
            elif "不足" in str(aa):
                if "无敌" not in str(bb):
                    print("不足,1")

                    return 2
                else:
                    return 3
            elif ("一二" in bb or "快递员" in bb) and "无敌" not in str(bb):

                print("一二,4")

                return 4
            elif "士拉" in str(aa):

                print("士拉,2")

                return 1
            elif "数字" in str(aa):

                print(str(aa), "0110,3")

                return 3
            else:
                print('没有找到人物pvp')
                continue

    def FuBen_INFO13(self):  # 校验是否选择“传送门”

        # self.Use_Dict(0)
        while True:
            # time.sleep(0.15)

            aa = self.Find_Ocr(
                x1=0,
                y1=0,
                x2=760,
                y2=450,
                color_format="#353",
                sim=0.95,
                linesign=" ",
                isbackcolor=0)
            print(aa)
            if "传送" in str(aa):

                print(str(aa), "有传送,1")

                return 1
            else:
                print('没有传送')
                return 0

    def FuBen_INFOxx(self, num):

        if num == 1:
            # self.forxunhuan(601,478)
            self.forxunhuan()
            return
        elif num == 2:
            # self.forxunhuan(202,500)
            self.forxunhuan()
            return
        elif num == 3:
            # self.forxunhuan(202, 520)
            self.forxunhuan()
            return
        elif num == 4:
            self.forxunhuan()
            # self.forxunhuan(530, 485)
            return
        elif num == 5:
            self.forxunhuan()
            # self.forxunhuan(415, 420)
            return
        elif num == 6:
            dt.press('up')

            time.sleep(0.013)  # 按下两秒
            dt.keyDown('up')  # ：模拟按键按下
            time.sleep(1.35)  # 按下19秒
            dt.keyUp('up')  # ：模拟按键松开按键
            # self.forxunhuan(493, 490)
            self.forxunhuan()
            return
        elif num == 7:
            # self.forxunhuan(131, 393)
            self.forxunhuan()
            dt.press('right')

            time.sleep(0.013)  # 按下两秒
            dt.keyDown('right')  # ：模拟按键按下
            time.sleep(1.56)  # 按下19秒
            dt.keyUp('right')  # ：模拟按键松开按键
        else:
            # self.youjian()  # 发送邮件

            print('FuBen_INFOxx什么也没有找到')
            time.sleep(10)

    def FuBen_INFO6(self, b=0, c=1):  # 副本地图信息数据实时获取
        # self.Set_Dict(0, 'test3.txt')
        # self.Use_Dict(0)
        # c.Set_Dict(0, "test3.txt")
        # time.sleep(2)
        cc = c

        while True:
            print(b, c, 'B和C')
            cc = cc + 1
            print('FuBen_INFO6 开始')
            aa6 = self.Find_Ocr(
                x1=392,
                y1=63,
                x2=853,
                y2=574,
                color_format="#380",
                sim=0.88,
                linesign=" ",
                isbackcolor=0)
            time.sleep(0.15)
            aa7 = self.Find_Ocr(
                x1=5,
                y1=5,
                x2=800,
                y2=600,
                color_format="#380",
                sim=0.88,
                linesign=" ",
                isbackcolor=0)
            time.sleep(0.15)
            if str(aa6) is None or str(aa7) is None:
                print(str(aa6), str(aa7), 'aa is None')
                continue

            elif "开洞" in str(aa6):
                print('开洞', cc)
                return
            elif cc == 6 and "开洞" not in str(aa6) and b == 1:
                dt.press('right')
                time.sleep(0.013)  # 按下两秒
                dt.keyDown('right')  # ：模拟按键按下
                time.sleep(0.2)  # 按下19秒
                dt.keyUp('right')  # ：模拟按键松开按键
                cc += 1
                continue
            elif cc == 7 and "开洞" not in str(aa6) and b == 1:
                dt.press('right')
                dt.press('alt')

                continue
            elif (cc == 25 or cc == 51) and "开洞" not in str(aa6):
                dt.press('alt')
                continue
            elif 50 > cc > 7 and "开洞" not in str(aa6):
                dt.press('right')
                continue
            elif cc <= 7 and "开洞" not in str(aa7):
                print("9等待9")
                dt.press('g')
                time.sleep(0.075)
                dt.press('y')
                time.sleep(1)

                continue

            elif "开洞" in str(aa7):
                print('开洞', cc, str(aa7))
                return
            else:

                try:
                    print('FuBen_INFO6 #000')  # 000
                    # self.youjian()
                    break
                except OSError as de:
                    print(de, 'FuBen_INFO6')

                    traceback.print_exc()
                    break
                except Exception as e:
                    print(e, 'FuBen_INFO6')

                    traceback.print_exc()
                    break

    def FuBen_INFO666(self, b=0, c=1):  # 副本地图信息数据实时获取
        # self.Set_Dict(0, 'test3.txt')
        # self.Use_Dict(0)
        # c.Set_Dict(0, "test3.txt")
        # time.sleep(2)
        cc = c
        while True:
            cc = cc + 1

            aa = self.Find_Ocr(
                x1=392,
                y1=63,
                x2=853,
                y2=574,
                color_format="#380",
                sim=0.88,
                linesign=" ",
                isbackcolor=0)
            aa1 = self.Find_Ocr(
                x1=0,
                y1=0,
                x2=800,
                y2=600,
                color_format="#380",
                sim=0.88,
                linesign=" ",
                isbackcolor=0)
            if "开洞" in aa:
                print('开洞', cc)
                return

            elif cc == 6 and "开洞" not in aa and b == 1:
                dt.press('right')
                time.sleep(0.013)  # 按下两秒
                dt.keyDown('right')  # ：模拟按键按下
                time.sleep(0.2)  # 按下19秒
                dt.keyUp('right')  # ：模拟按键松开按键
                continue
            elif (cc == 25 or cc == 51) and "开洞" not in aa:
                dt.press('alt')
                continue
            elif 50 > cc > 7 and "开洞" not in aa:
                dt.press('right')
                time.sleep(0.013)  # 按下两秒
                dt.keyDown('right')  # ：模拟按键按下
                time.sleep(0.19)  # 按下19秒
                dt.keyUp('right')  # ：模拟按键松开按键
                continue
            elif cc <= 7 and "开洞" not in aa:
                print("9等待9")
                dt.press('g')
                time.sleep(0.075)
                dt.press('y')
                time.sleep(1)

                continue

            elif "开洞" in aa1:
                print('开洞', cc, aa1)
                return
            else:
                print('FuBen_INFO6 #000')  # 000
                # self.youjian()
                break

    def FuBen_INFO99(self):  # 副本地图信息数据实时获取
        for i in range(1, 80):
            time.sleep(0.05)
            aa = self.Find_Ocr(
                x1=480,
                y1=22,
                x2=800,
                y2=600,
                color_format="#360",
                sim=0.8,
                linesign=" ",
                isbackcolor=0)
            if "返回" in str(aa) or "再次挑战" in str(aa) or "经验值" in str(aa) or "通关时间" in str(aa):
                print('总经验值,请选择奖励，返回')
                zjyz[0] = 1
                return 1
            elif '崩溃' in str(aa):
                print('从新调用self.FuBen_INFO99()')
                continue
            else:
                print('没找到总经验值，返回', aa)

                continue


    def FuBen_INFO66(self, ss=0):  # 副本地图信息数据实时获取
        # self.Set_Dict(0, 'test3.txt')
        # self.Use_Dict(0)
        # c.Set_Dict(0, "test3.txt")
        # cc = c
        sss2 = ss
        bs=0
        # print('门外面')
        while True:
            # cc = cc + 1
            # print('进来了')
            aa00 = self.Find_Ocr(
                x1=300,
                y1=63,
                x2=800,
                y2=574,
                color_format="#380",
                sim=1,
                linesign=" ",
                isbackcolor=0)

            aa11 = self.Find_Ocr(
                x1=0,
                y1=0,
                x2=800,
                y2=600,
                color_format="#380",
                sim=1,
                linesign=" ",
                isbackcolor=0)

            aa12 = self.Find_Ocr(
                x1=0,
                y1=0,
                x2=800,
                y2=600,
                color_format="#130",
                sim=1,
                linesign=" ",
                isbackcolor=0)

            aa13 = self.Find_Ocr(
                x1=0,
                y1=0,
                x2=800,
                y2=600,
                color_format="#422",
                sim=0.99,
                linesign=" ",
                isbackcolor=0)
            aa016 = self.Find_Ocr(
                x1=300,
                y1=63,
                x2=800,
                y2=574,
                color_format="#380",
                sim=75,
                linesign=" ",
                isbackcolor=0)
            aa017 = self.Find_Ocr(
                x1=200,
                y1=63,
                x2=800,
                y2=574,
                color_format="#380",
                sim=75,
                linesign=" ",
                isbackcolor=0)
            aamm = self.Find_Ocr(
                x1=350,
                y1=0,
                x2=800,
                y2=450,
                color_format="#360",
                sim=0.99,
                linesign=" ",
                isbackcolor=0)
            print(aa00, '=aa00')  # 开洞  =aa00
            print(aa11, '=aa11')  # 开洞  =aa11
            print(aa12, '=aa12')  # 洞府  =aa12
            print(aa13, '=aa13')
            print(aa016, '=aa06')
            print('')
            print(sss2, 'bs=================', bs)
            print(aamm, 'aamm=================', aamm)
            if '崩溃' in str(aa11) or '崩溃' in str(aa13):

                # if  aa00 in '崩溃' or aa11 in '崩溃' or aa13 in '崩溃' or aa12 in '崩溃':
                print('FuBen_INFO66崩溃')
                continue
            elif aa11 == 0 and aa13 == 0 and aa12 == 0 and aa00 == 0 and aamm == 0:
                # if  aa00 in '崩溃' or aa11 in '崩溃' or aa13 in '崩溃' or aa12 in '崩溃':
                print('FuBen_INFO66崩溃')
                self.Set_Dict(0, 'test3.txt')
                continue
            elif (sss2 == 8 or sss2 == 7) and bs <= 4:
                print('测试这个新任务，', time.time())
                bs += 1
                if "大鱼海棠" in str(aamm) or "开洞" in str(aa00) or "开洞" in str(aa016) or "开洞" in str(aa017):
                    print("在第四关")
                    return
                elif aa11 == 0 and aa13 == 0 and aa12 == 0 and aa00 == 0 and aamm == 0:
                    # if  aa00 in '崩溃' or aa11 in '崩溃' or aa13 in '崩溃' or aa12 in '崩溃':
                    print('FuBen_INFO66崩溃')
                    self.Set_Dict(0, 'test3.txt')
                    continue
                else:
                    try:

                        dt.press('right')
                        dt.keyDown('right')  # ：模拟按键按下 向下
                        time.sleep(0.35)
                        dt.keyUp('right')  # ：模拟按键松开按键
                        print('FuBen_INFO66 sss2==', sss2)
                        continue

                    except OSError as de:
                        print(de, 'FuBen_INFO66 ==]')
                        traceback.print_exc()
                        continue
                    except Exception as e:
                        print(e, 'FuBen_INFO66 ==')
                        traceback.print_exc()
                        continue


            elif ("开洞" not in str(aa00) or "开洞" not in str(aa016) or "开洞" not in str(aa017)) and bs <= 7:

                print('开洞，洞口没找到')
                bs += 1
                continue
            elif '开府' in str(aa12):
                if sss2 == 2:
                    dt.press('d')

                    print('FuBen_INFO66 步骤三')
                    continue

                else:
                    try:

                        dt.press('g')
                        print('FuBen_INFO66 步骤四')
                        continue

                    except OSError as de:
                        print(de, 'FuBen_INFO66 ==]')
                        traceback.print_exc()
                        continue
                    except Exception as e:
                        print(e, 'FuBen_INFO66 ==')
                        traceback.print_exc()
                        continue






            elif '开洞' not in str(aa00) and sss2 == 2:  # and aa00 !=None:
                print(aa00, aa11, aa12, aa13, 'aa is None or aa1 is None or aa3 is None or aa2 is None')

                dt.press('right')
                dt.keyDown('right')  # ：模拟按键按下 向下
                time.sleep(0.35)
                dt.keyUp('right')  # ：模拟按键松开按键
                continue
            elif "开洞" in str(aa11) and "开洞" not in str(aa00) and '开府' not in str(aa12) or '开洞' not in str(aa00):
                if sss2 == 7:
                    dt.press('right')
                    dt.keyDown('right')  # ：模拟按键按下 向下
                    time.sleep(0.35)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    print('FuBen_INFO66 步骤一')
                    continue
                elif "开洞" in str(aa11) or "开洞" in str(aa11) and '开洞' not in str(aa00):
                    print('??????????????????????????????????????????????')
                    dt.press('right')
                    continue
                elif aa11 == 0 and aa13 == 0 and aa12 == 0 and aa00 == 0:
                    # if  aa00 in '崩溃' or aa11 in '崩溃' or aa13 in '崩溃' or aa12 in '崩溃':
                    print('FuBen_INFO66崩溃')
                    self.Set_Dict(0, 'test3.txt')
                    continue

                if "开洞" in str(aa00) or "开洞" in str(aa016) or "开洞" in str(aa017):

                    print('开洞，洞口没找到')
                    return
                elif aa00 == 0 or aa017 == 0 or aa016 == 0:
                    continue
                else:
                    try:

                        dt.press('right')
                        dt.keyDown('right')  # ：模拟按键按下 向下
                        time.sleep(0.25)
                        dt.keyUp('right')  # ：模拟按键松开按键
                        print('FuBen_INFO66 步骤一')
                        continue
                    except OSError as de:
                        print(de, 'FuBen_INFO66 ==]')
                        traceback.print_exc()
                        continue
                    except Exception as e:
                        print(e, 'FuBen_INFO66 ==')
                        traceback.print_exc()
                        continue

            elif "开洞" not in str(aa11) and "开洞" not in str(aa00) and '开府' not in str(aa12):
                dt.press('right')
                dt.keyDown('right')  # ：模拟按键按下 向下
                time.sleep(0.25)
                dt.keyUp('right')  # ：模拟按键松开按
                continue
            elif aa00 is None or aa11 is None or aa13 is None or aa12 is None:
                if "开洞" in str(aa00):
                    dt.press('right')
                    print('开洞，洞口没找到')
                    return
                else:
                    try:

                        print(aa00, aa11, aa12, aa13, 'aa is None or aa1 is None or aa3 is None or aa2 is None')
                        dt.press('right')

                        print('崩溃向前走一下')
                        continue
                    except OSError as de:
                        print(de, 'FuBen_INFO66 ==]')
                        traceback.print_exc()
                        continue
                    except Exception as e:
                        print(e, 'FuBen_INFO66 ==')
                        traceback.print_exc()
                        continue

            elif "开洞" not in str(aa00) and "开洞" in str(aa11):
                dt.press('right')
                print('开洞，洞口没找到')
                continue
            elif "开洞" not in str(aa00) and "开洞" in str(aa11):
                dt.press('right')
                print('开洞，洞口没找到')
                continue
            elif "开洞" in str(aa00):
                print('开洞')
                return

            elif '开府' in str(aa11):
                if sss2 == 2:
                    dt.press('d')

                    print('FuBen_INFO66 步骤三')
                    continue
                else:

                    try:

                        dt.press('g')
                        print('FuBen_INFO66 步骤四')
                        continue

                    except OSError as de:
                        print(de, 'FuBen_INFO66 ==]')
                        traceback.print_exc()
                        continue
                    except Exception as e:
                        print(e, 'FuBen_INFO66 ==')
                        traceback.print_exc()
                        continue
            elif '德拉' in str(aa13) or '防护罩' in str(aa13):
                dt.press('y')

                dt.press('g')
                print('FuBen_INFO66 步骤五')
                continue

            elif '德拉' not in str(aa13) and '开洞' not in str(aa00) and '开洞' not in str(aa12) and '开洞' not in str(
                    aa11) and sss2 == 7:
                print('非')
                dt.press('right')
                dt.keyDown('right')  # ：模拟按键按下 向下
                time.sleep(0.3)
                dt.keyUp('right')  # ：模拟按键松开按键
                print('----------')
                # sss2 = +7
                print('FuBen_INFO66 步骤六')

                continue
            elif '非' in str(aa13) and '非' in str(aa00) and '非' in str(aa12) and '非' in str(aa11) and sss2 == 7:
                print('非')
                dt.press('right')
                dt.keyDown('right')  # ：模拟按键按下 向下
                time.sleep(0.5)
                dt.keyUp('right')  # ：模拟按键松开按键
                print('----------')
                sss2 = +1
                continue
            elif '非' in str(aa13) and '非' in str(aa00) and '非' in str(aa12) and '非' in str(aa11):
                print('非')
                print('FuBen_INFO66 步骤七')

                # self.Set_Dict(0, 'test3.txt')
                continue

            elif '德拉' not in str(aa13) and sss2 == 1:
                if "开洞" in str(aa00):
                    # dt.press('right')
                    print('开洞，洞口没找到')
                    return
                elif aa00 == 0:
                    continue

                elif '开府' in str(aa12):
                    if sss2 == 2:
                        dt.press('d')

                        print('FuBen_INFO66 步骤三')
                        continue
                    else:

                        dt.press('g')
                        print('FuBen_INFO66 步骤四')
                        continue
                elif '花花' in str(aa12):
                    dt.press('right')
                    continue
                else:

                    try:

                        dt.press('right')
                        dt.keyDown('right')  # ：模拟按键按下 向下
                        time.sleep(0.3)
                        dt.keyUp('right')  # ：模拟按键松开按键

                    except OSError as de:
                        print(de, 'FuBen_INFO66 ==]')
                        traceback.print_exc()
                        continue
                    except Exception as e:
                        print(e, 'FuBen_INFO66 ==')
                        traceback.print_exc()
                        continue

            # elif str(aa00) is None or str(aa11) is None or str(aa13) is None or str(aa12) is None or \
            #         str(aa00) in '崩溃' or str(aa11) in '崩溃' or str(aa12) in '崩溃' or str(aa13) in '崩溃':
            #     print(aa00, aa11, aa12, aa13, 'aa is None or aa1 is None or aa3 is None or aa2 is None')
            #     dt.press('right')
            #     dt.keyDown('right')  # ：模拟按键按下 向下
            #     time.sleep(0.3)
            #     dt.keyUp('right')  # ：模拟按键松开按键
            #     print('崩溃向前走一下')
            #     continue

            elif str(aa11) in '崩溃' or str(aa13) in '崩溃' or str(aa12) in '崩溃':
                # if  aa00 in '崩溃' or aa11 in '崩溃' or aa13 in '崩溃' or aa12 in '崩溃':
                print('FuBen_INFO66崩溃')
                continue
            elif aa11 == 0 and aa13 == 0 and aa12 == 0 and aa00 == 0:
                # if  aa00 in '崩溃' or aa11 in '崩溃' or aa13 in '崩溃' or aa12 in '崩溃':
                print('FuBen_INFO66崩溃')
                self.Set_Dict(0, 'test3.txt')
                continue
            elif "开洞" not in str(aa00) and "开洞" not in str(aa11):
                dt.press('right')
                print('开洞，洞口没找到pop')
                continue
            else:
                try:
                    dt.press('right')
                    print('FuBen_INFO66,123')
                    # break
                    continue
                except OSError as de:
                    print(de, '  continue 九')
                    traceback.print_exc()
                    continue
                except Exception as e:
                    print(e, '  continue 九')
                    traceback.print_exc()
                    continue

    def yuren(self, num_parameter, move_seepx, move_seepy, Restart_computer_parameter, sss, aa1, bb1, cc1, dd1,
              hh1=0.75, hh2=0.75):
        # 女气功 花花
        print(num_parameter, move_seepx, move_seepy, Restart_computer_parameter, sss, aa1, bb1, cc1, dd1, hh1, hh2)
        # 定位坐标[606,401,75,499, 502,481, 460,481, 360,454]
        # self.forxunhuan(460,481)
        for i in range(1, 60):
            time.sleep(0.5)
            num = num_parameter  # num不能是奇数 运行几次
            # move_seep = -0.52  # 57.7   气功4.2  40.8
            # move_seep1= -0.26
            mmmm = 0
            move_seep = -move_seepx - 0.1
            move_seep1 = -move_seepy
            # move_seep = -0.8  # 57.70
            # move_seep1 = -0.23
            m_button = 'h'
            print(move_seep)
            a_error = 0
            Restart_computer = Restart_computer_parameter  # Restart_computer为0或者1，0关闭电脑，1不关闭电脑
            for j in range(1, 9):
                # print('right开始按下{}次'.format(j))
                # a_error = 0
                if j == 1:
                    global ret_values
                    ret_values = []
                    time.sleep(0.65)  # 按下两秒
                    dt.press('a')
                    t13 = Thread(target=self.FuBen_INFO3,
                                 args=(1,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t13.start()  # 开始运行t1线程
                    time.sleep(0.7)  # 按下两秒
                    dt.press('e')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('w')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('r')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(2.6 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('x')  # ：模拟按键按下
                    time.sleep(0.15)  # 按下两秒
                    dt.press('y')
                    t13.join()
                    if ret_values[0] == 1:
                        j = 8
                        i = num
                        print('测试', t13)
                        break
                        # elif self.FuBen_INFO3() == 2:
                    elif ret_values[0] == 2:
                        j = 8
                        break
                    # if self.FuBen_INFO3() == 1:
                    #     j = 8
                    #     i = num
                    #     print('测试')
                    #     break
                    else:
                        try:
                            self.FuBen_INFO66(2)

                            self.FuBen_INFO6()
                            # dt.keyDown('down')  # ：模拟按键按下 向下
                            # time.sleep(0.5 + move_seep1)
                            # dt.keyUp('down')  # ：模拟按键松开按键
                            dt.press('9')
                            time.sleep(0.3)  # 按下两秒
                            if self.FuBen_INFO() == 1:
                                self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                                # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                                # time.sleep(1)  # 按下两秒
                                dt.keyDown('right')  # ：模拟按键按下 向下
                                time.sleep(1.5 + move_seep1)
                                dt.keyUp('right')  # ：模拟按键松开按键


                            else:
                                pass
                        except OSError as de:
                            print(de)
                            print(i)
                            traceback.print_exc()
                        except Exception as e:
                            print(e)
                            print(i)
                            traceback.print_exc()

                elif j == 2:
                    time.sleep(0.3)  # 按下两秒
                    dt.press('w')
                    time.sleep(0.6)  # 按下两秒
                    dt.press('e')
                    time.sleep(0.7)  # 按下两秒

                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.8 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键
                    # self.Find_srt("熟练者", "#422", "功师", "#422")
                    dt.press('q')
                    time.sleep(0.65)
                    dt.press('d')
                    time.sleep(0.1)  # 按下两秒
                    dt.press('q')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('right')
                    dt.keyDown('right')  # ：模拟按键按下 向下
                    time.sleep(1.3 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    self.FuBen_INFO66(2)

                    self.FuBen_INFO6()
                    dt.press('9')
                    time.sleep(0.7)
                    self.FuBen_INFO1()
                    # time.sleep(1)  # 按下两秒
                    print('二操作9成功')
                    # time.sleep(1.2)
                    # Caozuolei().FuBen_INFO1()

                    # time.sleep(1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.45 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键





                elif j == 3:
                    time.sleep(0.3)  # 按下两秒
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.66 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    # time.sleep(0.1)  # 按下两秒
                    dt.press('w')
                    time.sleep(0.5)  # 按下两
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('w')
                    dt.press('d')

                    time.sleep(0.8)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.55)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    self.FuBen_INFO66(2)
                    self.FuBen_INFO6()
                    time.sleep(0.5)
                    dt.press('9')
                    time.sleep(0.75)
                    self.FuBen_INFO1()
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)

                    dt.press('left')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(4.1 + move_seep)
                    dt.keyUp('left')  # ：模拟按键松开按键
                    dt.press('x')
                    time.sleep(0.5)
                    dt.press('y')
                    time.sleep(0.75)
                    dt.press('f')
                    dt.press('w')
                    time.sleep(0.1)  # 按下两秒
                    dt.press('w')
                    self.FuBen_INFO66(2)
                    self.FuBen_INFO6(1)
                    dt.press('9')
                    time.sleep(0.75)
                    self.FuBen_INFO1()
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.7 + move_seep)
                    t3 = Thread(target=self.forxunhuanY,
                                args=(sss, aa1, bb1, cc1, dd1, hh1, hh2,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t3.start()  # 开始运行t1线程
                    print('<---')
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(3.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                elif j == 4:
                    time.sleep(0.5)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.8 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('f')
                    time.sleep(0.01)  # 按下两秒
                    dt.press('w')
                    time.sleep(0.7)
                    dt.press('g')
                    time.sleep(0.65)
                    dt.press('h')
                    time.sleep(0.7)
                    dt.press('ctrl')
                    time.sleep(0.7)
                    dt.press('f')
                    time.sleep(0.01)  # 按下两秒
                    dt.press('w')
                    time.sleep(0.7)
                    dt.press('r')
                    # time.sleep(0.7)
                    # dt.press('f')
                    # time.sleep(0.01)  # 按下两秒
                    # dt.press('w')
                    self.FuBen_INFO66(1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.45 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    self.FuBen_INFO6(0, 6)

                    dt.press('9')
                    time.sleep(0.75)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.FuBen_INFO1()

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.7 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 5:
                    time.sleep(0.05)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.78 + move_seep1)
                    dt.press('d')
                    dt.keyUp('down')  # ：模拟按键松开按键

                    time.sleep(0.07)  # 按下两秒
                    dt.press('w')
                    time.sleep(0.87)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.47 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.35 + move_seep1)
                    dt.keyUp('up')  # ：模拟按键松开按
                    # dt.keyDown('left')  # ：模拟按键按下
                    # time.sleep(1 + move_seep1)
                    # dt.keyUp('left')  # ：模拟按键松开按键
                    dt.press('f')
                    time.sleep(0.1)  # 按下两秒
                    dt.press('w')
                    time.sleep(0.66)  # 按下两秒
                    dt.press('d')
                    time.sleep(0.1)  # 按下两秒
                    dt.press('w')
                    time.sleep(0.7)  # 按下两秒
                    # dt.press('d')
                    # time.sleep(0.1)  # 按下两秒
                    # dt.press('q')

                    self.FuBen_INFO66()
                    self.FuBen_INFO6()

                    if self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2) == 7:
                        dt.press('f')
                        time.sleep(0.1)  # 按下两秒
                        dt.press('w')
                        print('结束')
                    elif self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2) == 77:
                        dt.press('f')
                        time.sleep(0.1)  # 按下两秒
                        dt.press('w')
                        time.sleep(0.7)  # 按下两秒
                        dt.press('f')
                        time.sleep(0.1)  # 按下两秒
                        dt.press('w')

                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(0.5)
                        dt.keyUp('left')  # ：模拟按键松开按键
                        time.sleep(0.56)
                    else:
                        dt.press('9')
                        time.sleep(0.75)

                        self.FuBen_INFO1()
                        dt.press('up')
                        # time.sleep(1)
                        time.sleep(0.56)
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(0.9 + move_seep1)
                        dt.keyUp('right')  # ：模拟按键松开按键


                elif j == 6:
                    # time.sleep(0.65)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(0.6)
                    dt.keyDown('right')  # ：模拟按键按下 向下

                    time.sleep(0.58 + move_seep1)

                    dt.keyUp('up')  # ：模拟按键松开按键
                    time.sleep(0.5)
                    dt.press('f')
                    dt.keyUp('right')

                    dt.press('w')
                    time.sleep(0.68)
                    dt.press('f')

                    dt.press('w')
                    time.sleep(0.85)
                    # self.Find_srt(aa1, bb1, cc1, dd1)
                    dt.keyDown('right')  # ：模拟按键按下 向下

                    time.sleep(0.9 + move_seep1)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('down')
                    self.FuBen_INFO66(2)
                    self.FuBen_INFO6(0, 6)

                    dt.press('9')
                    time.sleep(0.75)
                    dt.press('e')
                    self.FuBen_INFO1()
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    if self.FuBen_INFO13() == 0:
                        mmmm = 0.45
                        dt.press('left')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(3.75 + move_seep)
                        dt.keyUp('left')  # ：模拟按键松开按键

                        time.sleep(0.5)
                        dt.press('w')
                        time.sleep(0.7)  # 按下两秒
                        dt.press('r')
                        time.sleep(0.75)
                        dt.press('d')
                        time.sleep(0.1)  # 按下两秒
                        dt.press('q')
                        self.FuBen_INFO66(2)
                        self.FuBen_INFO6(1)
                        dt.press('9')
                        time.sleep(0.7)
                        self.FuBen_INFO1()
                        # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                        self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(3.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                    else:
                        # Caozuolei().FuBen_INFO1()

                        time.sleep(0.75)  # 按下两秒
                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按
                        time.sleep(0.55 - mmmm)
                    # c=1



                elif j == 7:  # 奇数 反之偶数
                    time.sleep(0.55 - mmmm)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.15 + move_seep1 - mmmm)
                    dt.press('d')
                    dt.keyUp('up')  # ：模拟按键松开按键

                    time.sleep(0.01)  # 按下两秒
                    dt.press('w')
                    time.sleep(0.66)  # 按下19秒
                    dt.press('r')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('h')
                    time.sleep(0.75)  # 按下19秒
                    dt.press('w')
                    time.sleep(0.75)  # 按下19秒

                    dt.press('d')
                    time.sleep(0.66)  # 按下两秒
                    dt.press('q')
                    time.sleep(0.66)  # 按下两秒
                    dt.press('d')
                    time.sleep(0.1)  # 按下两秒
                    dt.press('q')
                    time.sleep(0.66)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.15 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键
                    self.FuBen_INFO66(2)

                    self.FuBen_INFO1()
                    # time.sleep(1)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2, -15)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    dt.press('9')
                    time.sleep(0.5)
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.6 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键


                elif j == 8:  # 奇数 反之偶数

                    time.sleep(0.3)
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.15 + move_seep)  # 按下19秒
                    dt.press('d')
                    dt.keyUp('right')  # ：模拟按键松开按键
                    time.sleep(0.01)  # 按下19秒
                    dt.press('w')
                    time.sleep(0.65)  # 按下19秒
                    dt.press('w')
                    time.sleep(0.65)  # 按下19秒
                    dt.press('s')
                    time.sleep(1.8)  # 按下两秒
                    for oalt in range(1, 6):
                        dt.press('d')
                        time.sleep(0.01)  # 按下两秒
                        dt.press('w')
                        print('什么 {} 什么'.format(oalt))
                        time.sleep(0.35)  # 按下两秒
                    dt.press('d')
                    time.sleep(0.01)  # 按下两秒
                    dt.press('r')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('d')
                    time.sleep(0.01)  # 按下两秒
                    dt.press('w')
                    self.FuBen_INFO2(1)
                    dt.press('0')

                    time.sleep(0.75)  # 按下两秒
                    for ii in range(1, 4):
                        num_num = num // 2
                        if i == num_num or i == 10 or i == num or i == 1:
                            time.sleep(5)  # 按下两秒
                            # from python_findpicture import Caozuolei1
                            self.sellGoods_xy()  # 点击一键出售按钮
                            print('等待3')
                            time.sleep(1.5)  # 按下两秒
                            dt.press('a')
                            time.sleep(1)  # 按下两秒
                            dt.press('space')
                            time.sleep(1)
                            dt.press('left')
                            time.sleep(1)
                            dt.press('space')
                            break
                        else:
                            break
                    time.sleep(5)  # 按下两秒
                    dt.press('esc')

                    # print('0')
                    # print(i)

                    time.sleep(2)

                    if i == num:
                        dt.press('.')
                        dt.press('f12')
                        time.sleep(2)
                    else:
                        dt.press('f10')
                    time.sleep(3)
                    # print('方法5运行{}'.format(j))

                time.sleep(0.5)
                # print('大循环{}'.format(j))
            sss = i * 6
            if i == num and Restart_computer == 0:
                # os.system("shutdown -s -t 30")
                b = time.strftime("%y-%m-%d_%H$%M$%S", time.localtime())

                img = pyautogui.screenshot(region=[48, 84, 848, 684])
                img = Image.fromarray(np.uint8(img))
                # img.save('D:\webdriver_new\lw\Tpshot{}s.png'.format(b))
                time.sleep(1)
                print(b)
                break
            elif i == num and Restart_computer == 1:
                os.system("shutdown -s -t 30")  # 30秒关闭电脑
                break

            print('费了{}疲劳'.format(sss))
        '''t
        dt.keyDown('a')#：模拟按键按下
        time.sleep(2) #按下两秒0
        dt.keyUp('a') #：模拟按键松开time.sleep(2)
        '''

    def naiMa(self, num_parameter, move_seepx, move_seepy, Restart_computer_parameter, sss, aa1, bb1, cc1, dd1,
              hh1=0.75, hh2=0.75):
        # 女气功 花花
        print(num_parameter, move_seepx, move_seepy, Restart_computer_parameter, sss, aa1, bb1, cc1, dd1, hh1, hh2)
        # 定位坐标[606,401,75,499, 502,481, 460,481, 360,454]
        # self.forxunhuan(460,481)
        for i in range(1, 60):
            time.sleep(0.5)
            num = num_parameter  # num不能是奇数 运行几次
            # move_seep = -0.52  # 57.7   气功4.2  40.8
            # move_seep1= -0.26
            mmmm = 0
            move_seep = -move_seepx - 0.1
            move_seep1 = -move_seepy
            # move_seep = -0.8  # 57.70
            # move_seep1 = -0.23
            m_button = 'h'
            print(move_seep)
            a_error = 0
            Restart_computer = Restart_computer_parameter  # Restart_computer为0或者1，0关闭电脑，1不关闭电脑
            t001 = Thread(target=self.timedaojishi, args=(pvp,))  # 定义线程t2，线程任务为5.30s倒计时，无参数
            t001.start()  # 开始运行t1线程
            print('开始')

            Restart_computer = Restart_computer_parameter  # Restart_computer为0或者1，0关闭电脑，1不关闭电脑
            for j in range(1, 9):
                # print('right开始按下{}次'.format(j))
                # a_error = 0
                if j == 1:
                    global ret_values
                    ret_values = []
                    time.sleep(0.65)  # 按下两秒
                    dt.press('h')
                    t13 = Thread(target=self.FuBen_INFO3,
                                 args=(1,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t13.start()  # 开始运行t1线程
                    time.sleep(0.65)  # 按下两秒
                    dt.press('q')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('d')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('g')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(2.6 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键fy
                    time.sleep(0.15)  # 按下两秒
                    dt.press('left')  # ：模拟按键松开按键

                    dt.press('y')
                    time.sleep(0.2)  # 按下两秒
                    t13.join()
                    if ret_values[0] == 1:
                        j = 8
                        i = num
                        print('测试', t13)
                        break
                    # elif self.FuBen_INFO3() == 2:
                    elif ret_values[0] == 2:
                        j = 8
                        break
                    # if self.FuBen_INFO3() == 1:
                    #     j = 8
                    #     i = num
                    #     print('测试')
                    #     break
                    else:
                        self.FuBen_INFO66()

                        self.FuBen_INFO6()
                        dt.keyDown('down')  # ：模拟按键按下 向下
                        time.sleep(0.5 + move_seep1)
                        dt.keyUp('down')  # ：模拟按键松开按键
                        dt.press('9')
                        if self.FuBen_INFO() == 1:
                            self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                            # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                            time.sleep(0.5)  # 按下两秒
                            dt.keyDown('right')  # ：模拟按键按下 向下
                            time.sleep(1.5 + move_seep1)
                            dt.keyUp('right')  # ：模拟按键松开按键


                        else:
                            pass

                elif j == 2:
                    dt.press('q')
                    time.sleep(0.6)  # 按下两秒
                    dt.press('d')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('g')
                    time.sleep(0.7)  # 按下两秒

                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(1 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键
                    # self.Find_srt("熟练者", "#422", "功师", "#422")
                    dt.press('f')
                    time.sleep(0.7)  # 按下两秒

                    dt.press('right')
                    dt.keyDown('right')  # ：模拟按键按下 向下
                    time.sleep(1.3 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('r')
                    time.sleep(0.1)  # 按下两秒

                    self.FuBen_INFO66(2)

                    self.FuBen_INFO6()
                    dt.press('9')
                    time.sleep(0.3)
                    self.FuBen_INFO1()
                    # time.sleep(1)  # 按下两秒
                    print('二操作9成功')
                    # time.sleep(1.2)
                    # Caozuolei().FuBen_INFO1()

                    # time.sleep(1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.35 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键





                elif j == 3:

                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.66 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    time.sleep(0.1)  # 按下两秒
                    dt.press('q')
                    time.sleep(0.5)  # 按下两
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.35 + move_seep)
                    dt.press('s')
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('q')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('d')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('g')
                    time.sleep(0.8)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.5)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    self.FuBen_INFO6()
                    # self.FuBen_INFO6()
                    # time.sleep(0.5)
                    dt.press('9')
                    time.sleep(0.5)
                    self.FuBen_INFO1()
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)

                    dt.press('left')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(3.9 + move_seep)
                    dt.keyUp('left')  # ：模拟按键松开按键
                    dt.press('q')
                    time.sleep(0.65)
                    dt.press('f')
                    time.sleep(0.75)
                    dt.press('q')

                    time.sleep(0.7)  # 按下两秒
                    dt.press('d')
                    self.FuBen_INFO66(2)
                    self.FuBen_INFO6(1)
                    dt.press('9')
                    time.sleep(0.7)
                    self.FuBen_INFO1()
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)
                    t3 = Thread(target=self.forxunhuanYnama,
                                args=(sss, aa1, bb1, cc1, dd1, hh1, hh2,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t3.start()  # 开始运行t1线程
                    # self.forxunhuanY(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    print('<---')
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(3.4 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                elif j == 4:
                    time.sleep(0.5)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.86 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('s')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('f')
                    time.sleep(0.7)
                    dt.press('e')
                    time.sleep(1.5)
                    dt.press('h')
                    time.sleep(0.7)

                    time.sleep(0.7)
                    dt.press('g')

                    # self.FuBen_INFO66(1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.4 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    self.FuBen_INFO66(1)
                    # time.sleep(1)
                    dt.press('9')
                    time.sleep(0.6)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.FuBen_INFO1()

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.6 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 5:
                    time.sleep(0.05)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.77 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    dt.press('t')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('q')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('d')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('g')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.2 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.43 + move_seep1)
                    dt.keyUp('up')  # ：模拟按键松开按
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(0.8 + move_seep1)
                    dt.keyUp('left')  # ：模拟按键松开按键
                    dt.press('r')
                    time.sleep(0.63)  # 按下两秒
                    dt.press('q')
                    time.sleep(0.53)  # 按下两秒

                    self.FuBen_INFO66()
                    self.FuBen_INFO6()

                    if self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2) == 7:
                        dt.press('f')
                        time.sleep(0.1)  # 按下两秒
                        dt.press('w')
                        print('结束')
                    elif self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2) == 77:
                        dt.press('f')
                        time.sleep(2)  # 按下两秒
                        dt.press('w')
                        time.sleep(0.7)  # 按下两秒
                        dt.press('f')
                        time.sleep(0.1)  # 按下两秒
                        dt.press('w')

                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(0.5)
                        dt.keyUp('left')  # ：模拟按键松开按键
                        time.sleep(0.56)
                    else:
                        dt.press('9')
                        time.sleep(0.75)
                        self.FuBen_INFO1()
                        dt.press('up')
                        # time.sleep(1)
                        time.sleep(0.56)
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(0.95 + move_seep1)
                        dt.keyUp('right')  # ：模拟按键松开按键


                elif j == 6:
                    time.sleep(0.65)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(0.85)
                    dt.keyDown('right')  # ：模拟按键按下 向下

                    time.sleep(0.58 + move_seep1)

                    dt.keyUp('up')  # ：模拟按键松开按键
                    time.sleep(0.63)
                    dt.keyUp('right')
                    dt.press('f')

                    time.sleep(0.68)
                    dt.press('q')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('r')

                    time.sleep(0.85)
                    # self.Find_srt(aa1, bb1, cc1, dd1)
                    dt.keyDown('right')  # ：模拟按键按下 向下
                    time.sleep(0.7 + move_seep1)
                    t31 = Thread(target=self.FuBen_INFO66,
                                 args=(1,))  # 定义线程t2，
                    t31.start()  # 开始运行t1线程
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('down')
                    dt.press('down')

                    # self.FuBen_INFO6(0, 6)

                    time.sleep(0.5)
                    dt.press('9')
                    t31.join()
                    time.sleep(0.75)
                    self.FuBen_INFO1()
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    if self.FuBen_INFO13() == 0:
                        mmmm = 0.45
                        dt.press('left')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(0.02)
                        dt.press('up')
                        time.sleep(3.4 + move_seep)
                        dt.keyUp('left')  # ：模拟按键松开按键

                        time.sleep(0.1)
                        dt.press('q')
                        time.sleep(0.7)  # 按下两秒
                        dt.press('d')
                        time.sleep(0.75)
                        dt.press('g')
                        time.sleep(0.7)  # 按下两秒
                        dt.press('f')
                        self.FuBen_INFO66(2)
                        self.FuBen_INFO6(1)
                        dt.press('9')
                        time.sleep(0.8)
                        self.FuBen_INFO1()
                        self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(3.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                    else:
                        # Caozuolei().FuBen_INFO1()

                        time.sleep(0.75)  # 按下两秒
                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按

                    # c=1



                elif j == 7:  # 奇数 反之偶数
                    time.sleep(0.55 - mmmm)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.55 + move_seep1 - mmmm)
                    dt.keyUp('up')  # ：模拟按键松开按键

                    dt.press('w')
                    time.sleep(0.05)  # 按下两秒
                    dt.press('w')
                    time.sleep(0.65)  # 按下19秒

                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.25 + move_seep)  # 按下19秒
                    dt.press('s')
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.95 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键
                    time.sleep(0.3)  # 按下19秒

                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, -15)
                    # self.FuBen_INFO666(0, 7)
                    self.FuBen_INFO66(7)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2, -15)
                    self.FuBen_INFO1()
                    # time.sleep(1)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    dt.press('9')
                    time.sleep(0.5)  # 按下19秒

                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.3 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键


                elif j == 8:  # 奇数 反之偶数

                    time.sleep(0.75)
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.35 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('y')
                    time.sleep(0.65)  # 按下19秒
                    dt.press('q')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('t')
                    time.sleep(0.73)  # 按下19秒
                    dt.press('s')
                    time.sleep(0.75)  # 按下19秒
                    dt.press('f')

                    time.sleep(0.65)  # 按下两秒
                    for oalt in range(1, 13):
                        dt.press('ctrl')
                        dt.press('y')
                        dt.press('f')
                        dt.press('d')
                        print('什么 {} 什么'.format(oalt))
                        time.sleep(0.35)  # 按下两秒

                    time.sleep(0.1)  # 按下两秒

                    self.FuBen_INFO2(1)
                    dt.press('0')

                    time.sleep(1)  # 按下两秒
                    for ii in range(1, 4):
                        num_num = num // 2
                        if i == num_num or i == 10 or i == num or i == 1:
                            time.sleep(5)  # 按下两秒
                            # from python_findpicture import Caozuolei1
                            self.sellGoods_xy()  # 点击一键出售按钮
                            print('等待3')
                            time.sleep(1.5)  # 按下两秒
                            dt.press('a')
                            time.sleep(1)  # 按下两秒
                            dt.press('space')
                            time.sleep(1)
                            dt.press('left')
                            time.sleep(1)
                            dt.press('space')
                            break
                        else:
                            break
                    time.sleep(5)  # 按下两秒
                    dt.press('esc')

                    # print('0')
                    # print(i)

                    time.sleep(2)

                    if i == num:
                        dt.press('.')
                        dt.press('f12')
                        time.sleep(2)
                    else:
                        dt.press('f10')
                    time.sleep(3)
                    # print('方法5运行{}'.format(j))

                time.sleep(0.5)
                # print('大循环{}'.format(j))
            sss = i * 6
            if i == num and Restart_computer == 0:
                # os.system("shutdown -s -t 30")
                b = time.strftime("%y-%m-%d_%H$%M$%S", time.localtime())

                img = pyautogui.screenshot(region=[48, 84, 848, 684])
                img = Image.fromarray(np.uint8(img))
                # img.save('D:\webdriver_new\lw\Tpshot{}s.png'.format(b))
                time.sleep(1)
                print(b)
                break
            elif i == num and Restart_computer == 1:
                os.system("shutdown -s -t 30")  # 30秒关闭电脑
                break

            print('费了{}疲劳'.format(sss))
        '''t
        dt.keyDown('a')#：模拟按键按下
        time.sleep(2) #按下两秒0
        dt.keyUp('a') #：模拟按键松开time.sleep(2)
        '''

    def gongjianshou(self, num_parameter, move_seepx, move_seepy, Restart_computer_parameter, sss, aa1, bb1, cc1, dd1,
                     hh1=0.75, hh2=0.75):
        # 弓箭手 旅人
        print(num_parameter, move_seepx, move_seepy, Restart_computer_parameter, sss, aa1, bb1, cc1, dd1, hh1, hh2)
        # 定位坐标[606,401,75,499, 502,481, 460,481, 360,454]
        # self.forxunhuan(460,481)
        for i in range(1, 60):
            time.sleep(0.5)
            num = num_parameter  # num不能是奇数 运行几次
            # move_seep = -0.52  # 57.7   气功4.2  40.8
            # move_seep1= -0.26

            move_seep = -move_seepx
            move_seep1 = -move_seepy
            # move_seep = -0.8  # 57.70
            # move_seep1 = -0.23
            m_button = 'h'
            print(move_seep)
            a_error = 0
            Restart_computer = Restart_computer_parameter  # Restart_computer为0或者1，0关闭电脑，1不关闭电脑
            for j in range(1, 9):
                # print('right开始按下{}次'.format(j))
                # a_error = 0
                if j == 1:
                    time.sleep(0.75)  # 按下两秒
                    dt.press('y')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('d')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('f')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('x')
                    time.sleep(0.65)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(2.6 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('left')  # ：模拟按键按下
                    time.sleep(0.15)  # 按下两秒
                    dt.press('x')

                    if self.FuBen_INFO3() == 1:
                        j = 8
                        i = num
                        print('测试')
                        break
                    else:
                        self.FuBen_INFO66()

                        self.FuBen_INFO6()
                        dt.keyDown('down')  # ：模拟按键按下 向下
                        time.sleep(0.5 + move_seep1)
                        dt.keyUp('down')  # ：模拟按键松开按键
                        dt.press('9')
                        if self.FuBen_INFO() == 1:

                            self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                            time.sleep(1)  # 按下两秒
                            dt.keyDown('right')  # ：模拟按键按下 向下
                            time.sleep(1.5 + move_seep1)
                            dt.keyUp('right')  # ：模拟按键松开按键


                        else:
                            pass

                elif j == 2:
                    dt.press('e')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('x')
                    time.sleep(0.1)
                    dt.press('x')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('g')
                    time.sleep(0.1)
                    dt.press('x')
                    time.sleep(0.1)
                    dt.press('x')
                    time.sleep(0.5)

                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(1 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键
                    # self.Find_srt("熟练者", "#422", "功师", "#422")

                    time.sleep(0.2)  # 按下两秒
                    dt.press('right')
                    dt.keyDown('right')  # ：模拟按键按下 向下
                    time.sleep(1 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    self.FuBen_INFO66()

                    self.FuBen_INFO6()
                    dt.press('9')
                    time.sleep(1)
                    self.FuBen_INFO1()
                    # time.sleep(1)  # 按下两秒
                    print('二操作9成功')
                    # time.sleep(1.2)
                    # Caozuolei().FuBen_INFO1()

                    # time.sleep(1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.35 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键





                elif j == 3:

                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.66 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    time.sleep(0.5)  # 按下两秒
                    dt.press('x')
                    time.sleep(0.3)  # 按下两
                    dt.press('x')
                    time.sleep(0.3)  # 按下两
                    dt.press('x')
                    time.sleep(0.3)  # 按下两
                    dt.press('f')
                    time.sleep(0.7)  # 按下两

                    dt.press('y')
                    time.sleep(0.7)  # 按下两

                    dt.press('g')
                    time.sleep(0.7)  # 按下两
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.keyUp('left')  # ：模拟按键松开按键

                    dt.press('x')
                    time.sleep(0.3)  # 按下两
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.55)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    self.FuBen_INFO66()
                    self.FuBen_INFO6()
                    time.sleep(0.5)
                    dt.press('9')
                    time.sleep(1)
                    self.FuBen_INFO1()
                    self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)

                    dt.press('left')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(4.1 + move_seep)
                    dt.keyUp('left')  # ：模拟按键松开按键
                    dt.press('x')
                    time.sleep(0.5)
                    dt.press('x')
                    time.sleep(0.75)
                    dt.press('y')
                    time.sleep(0.75)
                    dt.press('d')
                    time.sleep(0.75)
                    dt.press('x')
                    self.FuBen_INFO66()
                    self.FuBen_INFO6(1)
                    dt.press('9')
                    time.sleep(1)
                    self.FuBen_INFO1()
                    self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(3.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                elif j == 4:
                    time.sleep(0.5)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.8 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('f')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('h')
                    time.sleep(0.7)
                    dt.press('x')
                    time.sleep(0.5)
                    dt.press('x')
                    time.sleep(0.7)
                    dt.press('f')
                    time.sleep(0.5)  # 按下两秒

                    self.FuBen_INFO66()
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.45 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    self.FuBen_INFO6(0, 6)
                    time.sleep(1)
                    dt.press('9')
                    time.sleep(1.2)

                    self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.FuBen_INFO1()

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.7 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 5:
                    time.sleep(0.05)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.75 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    dt.press('e')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('d')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('y')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.28 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.55 + move_seep1)
                    dt.keyUp('up')  # ：模拟按键松开按
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(1 + move_seep1)
                    dt.keyUp('left')  # ：模拟按键松开按键

                    time.sleep(0.8)  # 按下两秒

                    self.FuBen_INFO66()
                    self.FuBen_INFO6()

                    if self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2) == 7:
                        dt.press('f')
                        time.sleep(0.1)  # 按下两秒
                        dt.press('w')
                        print('结束')
                    elif self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2) == 77:
                        dt.press('f')
                        time.sleep(0.1)  # 按下两秒
                        dt.press('w')

                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(0.5)
                        dt.keyUp('left')  # ：模拟按键松开按键
                        time.sleep(0.56)
                    else:
                        dt.press('9')
                        time.sleep(0.75)
                        self.FuBen_INFO1()
                        # time.sleep(1)
                        time.sleep(0.56)
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(0.95 + move_seep1)
                        dt.keyUp('right')  # ：模拟按键松开按键


                elif j == 6:
                    time.sleep(0.75)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(0.65)
                    dt.keyDown('right')  # ：模拟按键按下 向下

                    time.sleep(0.45 + move_seep1)

                    dt.keyUp('up')  # ：模拟按键松开按键
                    time.sleep(0.6)
                    dt.keyUp('right')
                    dt.press('e')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('r')
                    time.sleep(0.7)

                    # self.Find_srt(aa1, bb1, cc1, dd1)
                    dt.keyDown('right')  # ：模拟按键按下 向下
                    time.sleep(0.9 + move_seep1)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    self.FuBen_INFO66()
                    self.FuBen_INFO6(c=6)
                    time.sleep(0.1)
                    dt.press('9')
                    time.sleep(0.75)
                    self.FuBen_INFO1()
                    self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    if self.FuBen_INFO13() == 0:
                        dt.press('left')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(3.75 + move_seep)
                        dt.keyUp('left')  # ：模拟按键松开按键

                        time.sleep(0.2)
                        dt.press('f')
                        time.sleep(0.7)  # 按下两秒
                        dt.press('g')
                        time.sleep(0.75)
                        dt.press('f')
                        time.sleep(0.7)  # 按下两秒
                        dt.press('x')
                        time.sleep(0.5)  # 按下两秒
                        dt.press('x')
                        time.sleep(0.5)  # 按下两秒
                        self.FuBen_INFO66()
                        self.FuBen_INFO6(1)
                        dt.press('9')
                        time.sleep(1)
                        self.FuBen_INFO1()
                        self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(3.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                    else:
                        # Caozuolei().FuBen_INFO1()

                        time.sleep(0.75)  # 按下两秒
                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按

                    # c=1



                elif j == 7:  # 奇数 反之偶数
                    time.sleep(0.75)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.8 + move_seep1)
                    dt.keyUp('up')  # ：模拟按键松开按键
                    dt.press('w')
                    time.sleep(0.75)  # 按下19秒
                    dt.press('x')

                    time.sleep(0.7)  # 按下19秒
                    dt.press('x')
                    time.sleep(0.5)  # 按下两秒
                    dt.press('x')
                    self.FuBen_INFO66(7)
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键
                    self.FuBen_INFO1()
                    # time.sleep(1)
                    # self.forxunhuan(aa1, bb1, cc1, dd1)
                    dt.press('9')
                    time.sleep(0.5)
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.8 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键


                elif j == 8:  # 奇数 反之偶数

                    time.sleep(1)
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(2.5 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('q')
                    time.sleep(0.65)  # 按下19秒
                    dt.press('y')
                    time.sleep(0.75)  # 按下19秒
                    dt.press('d')
                    time.sleep(0.89)  # 按下19秒
                    dt.press('f')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('g')
                    time.sleep(0.89)  # 按下19秒
                    dt.press('alt')
                    time.sleep(0.89)  # 按下19秒
                    dt.press('alt')
                    time.sleep(1)  # 按下两秒
                    dt.press('x')
                    time.sleep(0.6)  # 按下两秒
                    dt.press('x')
                    time.sleep(0.6)  # 按下两秒
                    dt.press('x')
                    time.sleep(0.6)  # 按下两秒
                    dt.press('x')
                    self.FuBen_INFO2()
                    dt.press('0')

                    time.sleep(1)  # 按下两秒
                    for ii in range(1, 4):
                        num_num = num // 2
                        if i == num_num or i == 10 or i == num or i == 1:
                            time.sleep(5)  # 按下两秒
                            # from python_findpicture import Caozuolei1
                            self.sellGoods_xy()  # 点击一键出售按钮
                            print('等待3')
                            time.sleep(1.5)  # 按下两秒
                            dt.press('a')
                            time.sleep(1)  # 按下两秒
                            dt.press('space')
                            time.sleep(1)
                            dt.press('left')
                            time.sleep(1)
                            dt.press('space')
                            break
                        else:
                            break
                    time.sleep(5)  # 按下两秒
                    dt.press('esc')

                    # print('0')
                    # print(i)

                    time.sleep(2)

                    if i == num:
                        dt.press('.')
                        dt.press('f12')
                        time.sleep(2)
                    else:
                        dt.press('f10')
                    time.sleep(3)
                    # print('方法5运行{}'.format(j))

                time.sleep(0.5)
                # print('大循环{}'.format(j))
            sss = i * 6
            if i == num and Restart_computer == 0:
                # os.system("shutdown -s -t 30")
                b = time.strftime("%y-%m-%d_%H$%M$%S", time.localtime())

                img = pyautogui.screenshot(region=[48, 84, 848, 684])
                img = Image.fromarray(np.uint8(img))
                # img.save('D:\webdriver_new\lw\Tpshot{}s.png'.format(b))
                time.sleep(1)
                print(b)
                break
            elif i == num and Restart_computer == 1:
                os.system("shutdown -s -t 30")  # 30秒关闭电脑
                break

            print('费了{}疲劳'.format(sss))
        '''t
        dt.keyDown('a')#：模拟按键按下
        time.sleep(2) #按下两秒0
        dt.keyUp('a') #：模拟按键松开time.sleep(2)
        '''

    def nvQiGong(self, num_parameter, move_seepx, move_seepy, Restart_computer_parameter, sss, aa1, bb1, cc1, dd1,
                 hh1=0.75, hh2=0.75):
        # 女气功 花花
        print(num_parameter, move_seepx, move_seepy, Restart_computer_parameter, sss, aa1, bb1, cc1, dd1, hh1, hh2)
        # 定位坐标[606,401,75,499, 502,481, 460,481, 360,454]
        # self.forxunhuan(460,481)

        # ----》》测试
        # t1 = Thread(target=c.forxunhuanB,
        #             args=(sss, aa1, bb1, cc1, dd1, hh1, hh2,))  # 定义线程t1，线程任务为调用task1函数，task1函数的参数是6
        # t2 = Thread(target=self.forxunhuanA,
        #             args=(sss, aa1, bb1, cc1, dd1, hh1, hh2,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
        # t1.start()  # 开始运行t1线程
        # t2.start()  # 开始运行t2线程0
        # time.sleep(10000)
        # -------------》

        global gg_values
        gg_values[0] = 10
        for i in range(1, 60):
            print(gg_values[0])
            print(gg_values, "gg_values", ' 小循环', i, '大循环')

            time.sleep(0.5)
            num = num_parameter  # num不能是奇数 运行几次
            # move_seep = -0.52  # 57.7   气功4.2  40.8
            # move_seep1= -0.26

            move_seep = -move_seepx - 0.05
            move_seep1 = -move_seepy
            # move_seep = -0.8  # 57.70
            # move_seep1 = -0.23
            m_button = 'h'
            print(move_seep)
            a_error = 0
            print('pvp=', pvp)
            Restart_computer = Restart_computer_parameter  # Restart_computer为0或者1，0关闭电脑，1不关闭电脑
            t001 = Thread(target=self.timedaojishi, args=(int(pvp),))  # 定义线程t2，线程任务为5.30s倒计时，无参数
            t001.start()  # 开始运行t1线程
            print('开始')

            for j in range(1, 9):

                print(gg_values, "gg_values", ' 小循环', i, '大循环', j)
                # print('right开始按下{}次'.format(j))
                # a_error = 0
                if j == 1:

                    global ret_values
                    dt.keyUp('right')
                    dt.keyUp('up')
                    dt.keyUp('down')
                    ret_values = []
                    time.sleep(0.25)  # 按下两秒
                    dt.press('a')
                    t13 = Thread(target=self.FuBen_INFO3,
                                 args=(1,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t13.start()  # 开始运行t1线程
                    time.sleep(0.65)  # 按下两秒
                    dt.press('y')
                    time.sleep(0.68)  # 按下两秒
                    dt.press('g')
                    time.sleep(0.65)  # 按下两秒
                    dt.press('y')
                    time.sleep(0.65)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    if move_seepx > 0.98:
                        # move_seepx1= round(move_seepx / 1.5, 2)  #除以2 保留两位小数
                        time.sleep(2.6 + move_seep - 0.2)
                    else:
                        time.sleep(2.6 + move_seep)

                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('left')  # ：模拟按键按下
                    time.sleep(0.15)  # 按下两秒
                    dt.press('y')
                    # time.sleep(3)  # 按下两秒
                    # if self.FuBen_INFO3(1) == 1:
                    print(ret_values)
                    print(xxyy)
                    t13.join()
                    if ret_values[0] == 1:
                        j = 8
                        i = num
                        print('测试', t13, 'break 返回')
                        break
                    # elif self.FuBen_INFO3() == 2:
                    elif ret_values[0] == 2:
                        j = 8
                        print('break 返回')
                        break
                    else:
                        print('测试', t13)
                        self.FuBen_INFO66()

                        # self.FuBen_INFO6()
                        # dt.keyDown('down')  # ：模拟按键按下 向下
                        # time.sleep(0.5 + move_seep1)
                        # dt.keyUp('down')  # ：模拟按键松开按键
                        print('走到这不走了')
                        dt.press('9')
                        if self.FuBen_INFO() == 1:

                            self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                            print('成功')

                            dt.keyDown('right')  # ：模拟按键按下 向下
                            time.sleep(1.5 + move_seep1)
                            dt.keyUp('right')  # ：模拟按键松开按键
                        elif self.FuBen_INFO() == 100:
                            j = 9
                            print(9)
                            self.youjian()
                            time.sleep(600)
                            print(10)
                            self.movingfigur_right(2.5)  # 向右移动， 移动8秒，

                            # c.movingfigur_up(0.5)  # 向上移动， 移动0.15秒，
                            self.KeyPress1(190)  # 案件‘.’建，功能是隐藏技能和血功能
                            time.sleep(0.01)  # 睡眠0.5秒
                            dt.press('up')
                            self.FuBen_INFO11()
                            time.sleep(1.5)  # 睡眠1.5秒
                            dt.press('space')  # 单击空格操作
                            time.sleep(0.45)  # 睡眠1.5秒gdf
                        else:
                            pass

                elif j == 2:
                    dt.press('y')
                    # gg_values[0]=9
                    time.sleep(0.5)  # 按下两秒

                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(1 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键
                    # self.Find_srt("熟练者", "#422", "功师", "#422")
                    dt.press('y')
                    time.sleep(0.65)
                    dt.press('g')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('right')
                    dt.keyDown('right')  # ：模拟按键按下 向下
                    if move_seepx > 0.98:
                        # move_seepx1= round(move_seepx / 1.5, 2)  #除以2 保留两位小数
                        time.sleep(1.25 + move_seep + 0.1)

                    else:
                        time.sleep(1.25 + move_seep)

                    dt.keyUp('right')  # ：模拟按键松开按键

                    self.FuBen_INFO66()

                    # self.FuBen_INFO6()
                    dt.press('9')
                    time.sleep(0.5)
                    self.FuBen_INFO1()
                    # time.sleep(1)  # 按下两秒
                    print('二操作9成功')
                    # time.sleep(1.2)
                    # Caozuolei().FuBen_INFO1()

                    # time.sleep(1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.3 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键





                elif j == 3:
                    time.sleep(0.5)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.66 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    time.sleep(0.1)  # 按下两秒
                    dt.press('y')
                    time.sleep(0.5)  # 按下两
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('y')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('g')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('d')

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.5)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    self.FuBen_INFO66()
                    # self.FuBen_INFO6()
                    # time.sleep(0.5)
                    dt.press('9')
                    time.sleep(0.6)
                    self.FuBen_INFO1()
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    dt.press('left')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('left')  # ：模拟按键按下
                    # move_seep = -move_seepx - 0.05
                    print('move_seepx=={}'.format(move_seepx))
                    if move_seepx > 0.98:
                        time.sleep(3.45 + move_seep - move_seep1 - 0.3)
                    elif move_seepx > 0.65:
                        # move_seepx1= round(move_seepx / 1.5, 2)  #除以2 保留两位小数
                        time.sleep(3.45 + move_seep - move_seep1)
                    else:
                        time.sleep(3.45 + move_seep + 0.5)

                    dt.keyUp('left')  # ：模拟按键松开按键

                    time.sleep(0.5)
                    dt.press('y')
                    time.sleep(0.75)
                    dt.press('g')
                    time.sleep(0.75)
                    dt.press('f')
                    self.FuBen_INFO66()
                    #self.FuBen_INFO6(1)
                    dt.press('9')
                    time.sleep(0.7)
                    self.FuBen_INFO1()
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下

                    time.sleep(1.5 + move_seep)
                    xxyy[1] = 2
                    xxyy[3] = 4
                    # xxyy = [1, 2, 3, 4, 99, 6]
                    dt.keyUp('right')  # ：模拟按键松开按键

                    t3 = Thread(target=self.forxunhuanY,
                                args=(sss, aa1, bb1, cc1, dd1, hh1, hh2))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t3.start()  # 开始运行t1线程
                    time.sleep(0.56)
                    # self.forxunhuanY(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    print('<---')
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(3.3 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                elif j == 4:
                    time.sleep(0.5)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.75 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('y')

                    time.sleep(0.65)
                    dt.press('d')
                    time.sleep(0.65)  # 按下两秒
                    dt.press('f')
                    time.sleep(0.65)  # 按下两秒
                    dt.press('h')
                    time.sleep(0.65)  # 按下两秒
                    dt.press('e')
                    # self.FuBen_INFO66(1)

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    self.FuBen_INFO66(8)
                    # self.FuBen_INFO6(0, 6)

                    dt.press('9')
                    time.sleep(0.6)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    #self.FuBen_INFO66(1)

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.7 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 5:
                    time.sleep(0.05)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.75 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    if move_seepx > 0.98:
                        # move_seepx1= round(move_seepx / 1.5, 2)  #除以2 保留两位小数
                        time.sleep(1.205 + move_seep + 0.5)
                    else:
                        time.sleep(1.205 + move_seep)

                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.55 + move_seep1)
                    dt.keyUp('up')  # ：模拟按键松开按
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(1 + move_seep1)
                    dt.keyUp('left')  # ：模拟按键松开按键
                    dt.press('t')
                    time.sleep(0.6)
                    dt.press('c')

                    self.FuBen_INFO66()
                    ##self.FuBen_INFO6()

                    if self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2, 15) == 7:
                        print('结束')
                    elif self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2, 15) == 77:

                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(0.5)
                        dt.keyUp('left')  # ：模拟按键松开按键
                        time.sleep(0.56)
                    else:
                        dt.press('9')
                        time.sleep(0.55)
                        self.FuBen_INFO1()
                        # time.sleep(1)
                        dt.press('up')

                        time.sleep(0.56)
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.15 + move_seep1)
                        dt.keyUp('right')  # ：模拟按键松开按键


                elif j == 6:

                    time.sleep(0.75)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(0.77)
                    dt.keyDown('right')  # ：模拟按键按下 向下

                    time.sleep(0.55 + move_seep1)

                    dt.keyUp('up')  # ：模拟按键松开按键
                    time.sleep(0.75)
                    dt.keyUp('right')
                    dt.press('g')
                    time.sleep(0.7)
                    # self.Find_srt(aa1, bb1, cc1, dd1)
                    dt.keyDown('right')  # ：模拟按键按下 向下
                    time.sleep(0.65 + move_seep1)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    t31 = Thread(target=self.FuBen_INFO66,
                                 args=(0,))  # 定义线程t2，
                    t31.start()  # 开始运行t1线程
                    dt.press('down')
                    dt.press('down')

                    # self.FuBen_INFO6(0, 6)
                    dt.press('down')
                    time.sleep(0.7)
                    dt.press('9')
                    t31.join()
                    time.sleep(0.75)
                    self.FuBen_INFO1()
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2, -20)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 20)
                    if self.FuBen_INFO13() == 0:
                        dt.press('left')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(3.5 + move_seep)
                        dt.keyUp('left')  # ：模拟按键松开按键

                        time.sleep(0.5)
                        dt.press('y')
                        time.sleep(0.75)
                        dt.press('g')
                        time.sleep(0.75)
                        dt.press('f')
                        print('第6关开始')
                        ##self.FuBen_INFO66()
                        self.FuBen_INFO6(1)
                        dt.press('9')
                        time.sleep(0.85)
                        print('第6关开始1')
                        self.FuBen_INFO1()
                        print('第6关开始2')
                        self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                        # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 20)

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(3.05 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                    else:
                        # Caozuolei().FuBen_INFO1()

                        time.sleep(0.75)  # 按下两秒
                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按
                        time.sleep(1.6)  # 按下两秒

                    # c=1



                elif j == 7:  # 奇数 反之偶数
                    time.sleep(0.35)

                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.05 + move_seep1)
                    dt.keyUp('up')  # ：模拟按键松开按键
                    if move_seepx > 0.98:
                        # move_seepx1= round(move_seepx / 1.5, 2)  #除以2 保留两位小数
                        dt.press('right')
                        time.sleep(0.013)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.25 + move_seep - 0.15)  # 按下19秒
                        dt.keyUp('right')  # ：模拟按键松开按键
                        dt.press('alt')
                        dt.press('alt')
                    else:
                        dt.press('right')
                        time.sleep(0.013)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.25 + move_seep)  # 按下19秒
                        dt.keyUp('right')  # ：模拟按键松开按键
                        dt.press('alt')
                        dt.press('alt')
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.46 + move_seepy)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键
                    # time.sleep(1.3)  # 按下19秒
                    self.FuBen_INFO66(7)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2, -5)
                    # self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2, -5)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.FuBen_INFO1()
                    dt.press('up')
                    # time.sleep(1)
                    # self.forxunhuan(aa1, bb1, cc1, dd1)
                    dt.press('9')
                    time.sleep(0.5)
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.25 + move_seep)  # 按下19秒
                    # time.sleep(1.8 + move_seep)  # 按下19秒

                    dt.keyUp('right')  # ：模拟按键松开按键


                elif j == 8:  # 奇数 反之偶数
                    # gg_values[0]=7
                    global zjyz
                    zjyz = [9]  # zjyz意思为 “总经验值”
                    print(gg_values[0])
                    time.sleep(0.35)
                    if i == 1:

                        t38 = Thread(target=self.forxunhuanKF,
                                     args=(sss, aa1, bb1, cc1, dd1, hh1, hh2, 10))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                        t38.start()  # 开始运行t1线程
                    elif gg_values[0] == 0:
                        dt.press('up')
                    elif gg_values[0] == 1:
                        dt.press('down')
                    else:
                        pass
                    t360 = Thread(target=self.FuBen_INFO99)  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t360.start()  # 开始运行t1线程
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.3 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('q')
                    time.sleep(0.6)  # 按下19秒
                    dt.press('y')
                    time.sleep(0.65)  # 按下19秒
                    dt.press('f')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('g')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('e')
                    time.sleep(1.2)  # 按下19秒
                    dt.press('w')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('w')

                    for iiit in range(1, 100):
                        if zjyz[0] == 1:
                            break
                        else:
                            dt.press('q')
                            dt.press('q')
                            # dt.press('h')
                            dt.press('r')
                            dt.press('ctrl')
                            dt.press('f')
                            dt.press('ctrl')
                            dt.press('g')
                            continue

                    time.sleep(0.5)  # 按下19秒
                    if self.FuBen_INFO2() != 100 or zjyz[0] == 1:
                        dt.press('0')

                        time.sleep(0.5)  # 按下两秒
                        for ii in range(1, 4):
                            num_num = num // 2
                            if i == num_num or i == 10 or i == num or i == 1:
                                time.sleep(5)  # 按下两秒
                                # from python_findpicture import Caozuolei1
                                self.sellGoods_xy()  # 点击一键出售按钮
                                print('等待3')
                                time.sleep(1.5)  # 按下两秒
                                dt.press('a')
                                time.sleep(1)  # 按下两秒
                                dt.press('space')
                                time.sleep(1)
                                dt.press('left')
                                time.sleep(1)
                                dt.press('space')
                                break
                            else:
                                break
                        time.sleep(3)  # 按下两秒
                        dt.press('esc')

                        # print('0')
                        # print(i)

                        time.sleep(1)

                        if i == num:
                            dt.press('.')
                            dt.press('f12')
                            time.sleep(2)
                        else:
                            dt.press('f10')
                        time.sleep(2.8)
                        # print('方法5运行{}'.format(j))
                    else:
                        j = 9
                        print(9)
                        self.youjian()
                        time.sleep(600)
                        print(10)
                        self.movingfigur_right(2.5)  # 向右移动， 移动8秒，

                        # c.movingfigur_up(0.5)  # 向上移动， 移动0.15秒，
                        self.KeyPress1(190)  # 案件‘.’建，功能是隐藏技能和血功能
                        time.sleep(0.01)  # 睡眠0.5秒
                        dt.press('up')
                        self.FuBen_INFO11()
                        time.sleep(1.5)  # 睡眠1.5秒
                        dt.press('space')  # 单击空格操作
                        time.sleep(0.45)  # 睡眠1.5秒gdf
                time.sleep(0.5)
                # print('大循环{}'.format(j))
            sss = i * 6
            if i == num and Restart_computer == 0:
                # os.system("shutdown -s -t 30")
                b = time.strftime("%y-%m-%d_%H$%M$%S", time.localtime())

                img = pyautogui.screenshot(region=[48, 84, 848, 684])
                img = Image.fromarray(np.uint8(img))
                # img.save('D:\webdriver_new\lw\Tpshot{}s.png'.format(b))
                time.sleep(1)
                print(b)
                break
            elif i == num and Restart_computer == 1:
                os.system("shutdown -s -t 30")  # 30秒关闭电脑
                break

            print('费了{}疲劳'.format(sss))
        '''t
        dt.keyDown('a')#：模拟按键按下
        time.sleep(2) #按下两秒0
        dt.keyUp('a') #：模拟按键松开time.sleep(2)
        '''

    def kuangzhanshi(self, num_parameter, move_seepx, move_seepy, Restart_computer_parameter, sss, aa1, bb1, cc1,
                     dd1, hh1=0.75, hh2=0.75):

        for i in range(1, 60):
            time.sleep(2)
            num = num_parameter  # num不能是奇数 运行几次
            # move_seep = -0.52  # 57.7   气功4.2  40.8
            # move_seep1 = -0.26

            # move_seep = 0.189  # 57.70
            # move_seep1 = -0.23
            move_seep = -move_seepx
            move_seep1 = -move_seepy
            m_button = 'h'
            print(move_seep)
            Restart_computer = Restart_computer_parameter  # Restart_computer为0或者1，0关闭电脑，1不关闭电脑
            t001 = Thread(target=self.timedaojishi, args=(pvp,))  # 定义线程t2，线程任务为5.30s倒计时，无参数
            t001.start()  # 开始运行t1线程
            print('开始')

            Restart_computer = Restart_computer_parameter  # Restart_computer为0或者1，0关闭电脑，1不关闭电脑
            for j in range(1, 9):
                # print('right开始按下{}次'.format(j))
                if j == 1:
                    global ret_values
                    ret_values = []

                    time.sleep(0.3)  # 按下两秒
                    dt.press('h')
                    t13 = Thread(target=self.FuBen_INFO3, args=(1,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t13.start()  # 开始运行t1线程
                    time.sleep(0.54)  # 按下两秒
                    dt.press('d')
                    time.sleep(0.54)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(2.85 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    # time.sleep(3)  # 按下两秒
                    dt.press('left')
                    print(ret_values)
                    print(xxyy)
                    t13.join()
                    if ret_values[0] == 1:
                        j = 8
                        i = num
                        print('测试', t13)
                        break
                        # elif self.FuBen_INFO3() == 2:
                    elif ret_values[0] == 2:
                        j = 8
                        break
                    # if self.FuBen_INFO3() == 1:
                    #     j = 8
                    #     i = num
                    #     print('测试')
                    #     break
                    else:
                        self.FuBen_INFO66()
                        self.FuBen_INFO6()
                        dt.press('9')
                        # time.sleep(0.75)
                        # dt.keyDown('down')  # ：模拟按键按下 向下
                        # time.sleep(0.76 + move_seep1)
                        # dt.keyUp('down')  # ：模拟按键松开按键
                        if self.FuBen_INFO() == 1:
                            # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                            self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                            # time.sleep(1)  # 按下两秒
                            dt.keyDown('right')  # ：模拟按键按下 向下
                            time.sleep(1.5 + move_seep1)
                            dt.keyUp('right')  # ：模拟按键松开按键
                        else:
                            pass
                elif j == 2:

                    time.sleep(0.7)
                    dt.press('f')
                    time.sleep(0.3)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(1 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    dt.press('right')  # ：模拟按键按下
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.2 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('x')
                    time.sleep(0.3)
                    dt.press('x')
                    time.sleep(0.3)
                    dt.press('x')
                    time.sleep(0.3)
                    self.FuBen_INFO66()
                    self.FuBen_INFO6()
                    dt.press('9')
                    time.sleep(0.75)
                    self.FuBen_INFO1()

                    time.sleep(1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.3 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键




                elif j == 3:

                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.68 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('y')
                    time.sleep(0.75)  # 按下两秒

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.85 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    self.FuBen_INFO66()
                    dt.press('9')
                    time.sleep(0.8)
                    self.FuBen_INFO1()
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)

                    dt.press('left')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(4.0 + move_seep)
                    dt.keyUp('left')  # ：模拟按键松开按键

                    time.sleep(0.5)
                    dt.press('y')
                    time.sleep(0.75)
                    dt.press('f')
                    self.FuBen_INFO66()
                    self.FuBen_INFO6(1)
                    dt.press('9')
                    time.sleep(1)
                    self.FuBen_INFO1()
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)
                    t3 = Thread(target=self.forxunhuanY,
                                args=(sss, aa1, bb1, cc1, dd1, hh1, hh2,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t3.start()  # 开始运行t1线程

                    dt.keyUp('right')  # ：模拟按键松开按键
                    # self.forxunhuanY(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # print('<---')

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(3.75 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 4:
                    time.sleep(0.5)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.8 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('y')
                    time.sleep(0.75)
                    dt.press('f')
                    time.sleep(0.75)
                    dt.press('t')
                    time.sleep(0.7)
                    dt.press('a')

                    time.sleep(0.7)
                    dt.press('y')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('g')
                    time.sleep(0.8)  # 按下两秒

                    # self.FuBen_INFO66(1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(2.05 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    self.FuBen_INFO6(0, 6)
                    dt.press('9')
                    time.sleep(0.65)  # 按下两秒
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.FuBen_INFO1()
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.7 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 5:
                    time.sleep(0.03)
                    dt.keyDown('down')  # ：模拟按键按下 向下

                    time.sleep(0.75 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.2 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.9 + move_seep1)
                    dt.keyUp('up')  # ：模拟按键松开按键
                    dt.keyDown('left')  # ：模拟按键按下 向下
                    time.sleep(0.75 + move_seep1)
                    dt.keyUp('left')  # ：模拟按键松开按键
                    self.FuBen_INFO66()
                    self.FuBen_INFO6()

                    if self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2) == 7:
                        print('结束')

                    elif self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2) == 77:

                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(0.5)
                        dt.keyUp('left')  # ：模拟按键松开按键
                        time.sleep(0.6)

                    else:
                        dt.press('9')
                        time.sleep(0.7)
                        self.FuBen_INFO1()
                        dt.press('up')
                        time.sleep(0.5)
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1 + move_seep1)
                        dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 6:
                    time.sleep(0.75)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(0.6)
                    dt.keyDown('right')  # ：模拟按键按下 向下

                    time.sleep(1 + move_seep1)

                    dt.keyUp('up')  # ：模拟按键松开按键
                    time.sleep(0.7)
                    dt.keyUp('right')
                    dt.press('y')

                    time.sleep(0.65)  # 按下两秒
                    dt.press('s')

                    dt.press('right')
                    time.sleep(0.007)
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.7 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按
                    # self.FuBen_INFO66()
                    self.FuBen_INFO6(0, 6)

                    dt.press('9')
                    time.sleep(0.75)  # 按下两秒
                    self.FuBen_INFO1()
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 20)
                    if self.FuBen_INFO13() == 0:
                        dt.press('left')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(4.0 + move_seep)
                        dt.keyUp('left')  # ：模拟按键松开按键

                        time.sleep(0.5)
                        dt.press('y')
                        time.sleep(0.75)
                        dt.press('d')
                        time.sleep(0.75)
                        dt.press('f')
                        self.FuBen_INFO66()
                        self.FuBen_INFO6(1)
                        dt.press('9')
                        time.sleep(0.8)
                        self.FuBen_INFO1()
                        self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2, -5)
                        # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 20)

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(3.55 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键
                    else:
                        # Caozuolei().FuBen_INFO1()

                        # time.sleep(0.75)  # 按下两秒
                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按
                        time.sleep(0.75)  # 按下两秒
                    # c=1



                elif j == 7:  # 奇数 反之偶数

                    time.sleep(0.75)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.55 + move_seep1)
                    dt.keyUp('up')  # ：模拟按键松开按键

                    time.sleep(0.5)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.8 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('a')
                    time.sleep(0.5)  # 按下两秒
                    dt.press('y')
                    time.sleep(0.5)  # 按下两秒

                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.3 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键

                    self.FuBen_INFO66(7)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2 - 5)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.FuBen_INFO1()
                    dt.press('9')
                    time.sleep(0.7)

                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.34 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 8:  # 奇数 反之偶数
                    global zjyz
                    zjyz = [9]  # zjyz意思为 “总经验值”
                    time.sleep(0.65)
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.35 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键
                    t360 = Thread(target=self.FuBen_INFO99)  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t360.start()  # 开始运行t1线程
                    dt.press('ctrl')

                    time.sleep(0.7)  # 按下19秒
                    dt.press('d')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('f')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('y')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('t')
                    time.sleep(0.7)  # 按下19秒
                    for dazhao in range(1, 6):
                        dt.press('q')
                        time.sleep(0.65)  # 按下两秒
                    time.sleep(1.5)  # 按下两秒
                    dt.press('f')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('d')
                    time.sleep(0.6)  # 按下19秒
                    dt.press('y')
                    time.sleep(0.8)  # 按下19秒
                    for dazhao in range(1, 30):
                        if zjyz[0] == 1:
                            break
                        else:
                            dt.press('ctrl')
                            dt.press('left')
                            dt.press('f')
                            dt.press('g')
                            continue
                        # dt.press('a')
                    time.sleep(1)  # 按下两秒
                    if self.FuBen_INFO2() != 100 or zjyz[0] == 1:
                        dt.press('0')

                        time.sleep(0.5)  # 按下两秒
                        for ii in range(1, 4):
                            num_num = num // 2
                            if i == num_num or i == 10 or i == num or i == 1:
                                time.sleep(5)  # 按下两秒
                                # from python_findpicture import Caozuolei1
                                self.sellGoods_xy()  # 点击一键出售按钮
                                print('等待4')
                                time.sleep(1.5)  # 按下两秒
                                dt.press('a')
                                time.sleep(0.7)  # 按下两秒
                                dt.press('space')
                                time.sleep(1)
                                dt.press('left')
                                time.sleep(1)
                                dt.press('space')
                                break
                            else:
                                break
                        time.sleep(3)  # 按下两秒
                        dt.press('esc')

                        # print('0')
                        # print(i)

                        time.sleep(1)

                        if i == num:
                            dt.press('.')
                            dt.press('f12')
                            time.sleep(1.5)
                        else:
                            dt.press('f10')

                        time.sleep(2.8)
                        # print('方法5运行{}'.format(j))
                    else:
                        j = 9
                        print(9)
                        self.youjian()
                        time.sleep(600)
                        print(10)
                        self.movingfigur_right(2.5)  # 向右移动， 移动8秒，

                        # c.movingfigur_up(0.5)  # 向上移动， 移动0.15秒，
                        self.KeyPress1(190)  # 案件‘.’建，功能是隐藏技能和血功能
                        time.sleep(0.01)  # 睡眠0.5秒
                        dt.press('up')
                        self.FuBen_INFO11()
                        time.sleep(1.5)  # 睡眠1.5秒
                        dt.press('space')  # 单击空格操作
                        time.sleep(0.45)  # 睡眠1.5秒gdf
                time.sleep(0.5)
                # print('大循环{}'.format(j))
            sss = i * 6
            if i == num and Restart_computer == 0:
                # os.system("shutdown -s -t 30")
                b = time.strftime("%y-%m-%d_%H$%M$%S", time.localtime())

                img = pyautogui.screenshot(region=[48, 84, 848, 684])
                img = Image.fromarray(np.uint8(img))
                # img.save('D:\webdriver_new\lw\Tpshot{}s.png'.format(b))
                time.sleep(1)
                print(b)
                break
            elif i >= num and Restart_computer == 1:
                os.system("shutdown -s -t 30")  # 30秒关闭电脑
                break

            print('费了{}疲劳'.format(sss))
        '''t
        dt.keyDown('a')#：模拟按键按下
        time.sleep(2) #按下两秒0
        dt.keyUp('a') #：模拟按键松开time.sleep(2)
        '''

    def nanQiGong(self, num_parameter, Restart_computer_parameter, sss, aa1, bb1, cc1, dd1, hh1=0.75,
                  hh2=0.75):
        # 男气功

        for i in range(1, 60):
            time.sleep(2)
            num = num_parameter  # num不能是奇数 运行几次
            move_seep = -0.41  # 57.7
            m_button = 'h'
            print(move_seep)
            Restart_computer = Restart_computer_parameter  # Restart_computer为0或者1，0关闭电脑，1不关闭电脑
            t001 = Thread(target=self.timedaojishi, args=(pvp,))  # 定义线程t2，线程任务为5.30s倒计时，无参数
            t001.start()  # 开始运行t1线程
            print('开始')

            Restart_computer = Restart_computer_parameter  # Restart_computer为0或者1，0关闭电脑，1不关闭电脑
            for j in range(1, 9):
                # print('right开始按下{}次'.format(j))
                if j == 1:
                    global ret_values
                    ret_values = []
                    dt.press('t')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('h')
                    t13 = Thread(target=self.FuBen_INFO3,
                                 args=(1,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t13.start()  # 开始运行t1线程
                    time.sleep(0.75)  # 按下两秒
                    dt.press('s')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('w')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(2.85 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('left')
                    dt.press('y')
                    t13.join()
                    if ret_values[0] == 1:
                        j = 8
                        i = num
                        print('测试', t13)
                        break
                        # elif self.FuBen_INFO3() == 2:
                    elif ret_values[0] == 2:
                        j = 8
                        break
                    # if self.FuBen_INFO3() == 1:
                    #     j = 8
                    #     i = num
                    #     print('测试')
                    #     break
                    else:
                        self.FuBen_INFO66()
                        self.FuBen_INFO6()
                        dt.press('9')
                        time.sleep(0.75)
                        # dt.keyDown('down')  # ：模拟按键按下 向下
                        # time.sleep(0.45)
                        # dt.keyUp('down')  # ：模拟按键松开按键
                        if self.FuBen_INFO() == 1:
                            # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                            self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)

                            dt.keyDown('right')  # ：模拟按键按下 向下
                            time.sleep(1.5)
                            dt.keyUp('right')  # ：模拟按键松开按键
                        else:
                            pass


                elif j == 2:

                    time.sleep(0.1)
                    dt.press('y')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('s')
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    time.sleep(0.15)  # 按下两秒
                    dt.press('y')

                    time.sleep(0.75)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.015)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下 向下
                    time.sleep(0.6)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    self.FuBen_INFO66()
                    self.FuBen_INFO6()
                    dt.press('9')
                    time.sleep(0.75)
                    self.FuBen_INFO1()
                    print('二操作9成功')
                    # time.sleep(1.2)
                    # Caozuolei().FuBen_INFO1()

                    # time.sleep(1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.3 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键






                elif j == 3:
                    time.sleep(0.1)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.56)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    dt.press('y')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.25 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    time.sleep(0.75)  # 按下两秒
                    dt.press('e')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('e')
                    # time.sleep(0.75)  # 按下两秒
                    # dt.press('9')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.015)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.67)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    self.FuBen_INFO66()
                    self.FuBen_INFO6()

                    dt.press('9')
                    time.sleep(0.72)
                    self.FuBen_INFO1()
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)

                    dt.press('left')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(3.65 + move_seep)
                    dt.keyUp('left')  # ：模拟按键松开按键

                    dt.press('f')
                    time.sleep(0.75)
                    dt.press('s')
                    self.FuBen_INFO66()
                    self.FuBen_INFO6(1)
                    dt.press('9')
                    time.sleep(0.75)
                    self.FuBen_INFO1()
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    t3 = Thread(target=self.forxunhuanY,
                                args=(sss, aa1, bb1, cc1, dd1, hh1, hh2,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t3.start()  # 开始运行t1线程
                    # self.forxunhuanY(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # print('<---')

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(3.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键




                elif j == 4:
                    time.sleep(0.1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.75 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('y')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('e')
                    time.sleep(0.0075)  # 按下两秒
                    dt.press('e')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('f')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('s')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('alt')

                    time.sleep(0.7)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.65 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    self.FuBen_INFO66()
                    self.FuBen_INFO6()
                    # time.sleep(1)
                    dt.press('9')
                    time.sleep(0.7)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.FuBen_INFO1()
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 5:
                    time.sleep(0.1)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.75)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.05 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.6)
                    dt.keyUp('up')  # ：模拟按键松开按键
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(0.95 + move_seep)
                    dt.keyUp('left')  # ：模拟按键松开按键
                    dt.press('t')
                    time.sleep(0.8)
                    dt.press('c')

                    self.FuBen_INFO66()
                    self.FuBen_INFO1()
                    if self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2) == 7:
                        print('结束')

                    elif self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2) == 77:

                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(0.5)
                        dt.keyUp('left')  # ：模拟按键松开按键
                        time.sleep(0.56)
                    else:
                        # time.sleep(0.7)
                        dt.press('9')
                        time.sleep(0.55)
                        self.FuBen_INFO1()
                        # time.sleep(1)
                        dt.press('up')
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.05 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 6:
                    time.sleep(0.1)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(0.75)
                    dt.keyDown('right')
                    time.sleep(0.48)
                    dt.keyUp('up')  # ：模拟按键松开按键
                    time.sleep(0.45)
                    dt.keyUp('right')  # ：模拟按键松开按
                    time.sleep(0.85)  # 按下两秒
                    dt.press('e')
                    time.sleep(0.0075)  # 按下两秒
                    dt.press('e')
                    time.sleep(0.75)  # 按下两秒

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.7 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按
                    dt.press('down')
                    dt.press('9')
                    self.FuBen_INFO66()
                    self.FuBen_INFO1()
                    self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    if self.FuBen_INFO13() == 0:
                        dt.press('left')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(3.8 + move_seep)
                        dt.keyUp('left')  # ：模拟按键松开按键

                        time.sleep(0.5)
                        dt.press('y')
                        time.sleep(0.75)
                        dt.press('s')

                        self.FuBen_INFO66()
                        self.FuBen_INFO6(1)
                        dt.press('9')
                        time.sleep(0.75)
                        self.FuBen_INFO1()
                        self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, -15)

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(3.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                    else:
                        time.sleep(0.5)  # 按下两秒
                        dt.press('right')
                        time.sleep(0.0175)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.1 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按
                        time.sleep(1.4)  # 按下两秒
                    # c=1



                elif j == 7:  # 奇数 反之偶数

                    time.sleep(0.45)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.3 + move_seep)
                    dt.keyUp('up')  # ：模拟按键松开按键

                    time.sleep(0.01)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.2)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键

                    time.sleep(1)

                    time.sleep(0.75)  # 按下两秒
                    dt.press('y')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('d')

                    time.sleep(0.5)

                    self.FuBen_INFO66(7)
                    # self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, -25)

                    self.FuBen_INFO1()
                    dt.press('9')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('up')
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.25)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 8:  # 奇数 反之偶数
                    global zjyz
                    zjyz = [9]  # zjyz意思为 “总经验值”
                    time.sleep(0.4)
                    t38 = Thread(target=self.forxunhuanKF,
                                 args=(sss, aa1, bb1, cc1, dd1, hh1, hh2, 10))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t38.start()  # 开始运行t1线程
                    time.sleep(0.4)
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.05)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键
                    t360 = Thread(target=self.FuBen_INFO99)  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t360.start()  # 开始运行t1线程
                    # dt.press('e')
                    # time.sleep(0.65)  # 按下19秒
                    # dt.press('s')
                    # time.sleep(0.65)  # 按下19秒
                    # dt.press('f')
                    # time.sleep(0.65)  # 按下19秒
                    # dt.press('a')
                    # time.sleep(0.85)  # 按下两秒
                    dt.press('t')
                    time.sleep(0.6)
                    dt.press('ctrl')
                    time.sleep(0.6)
                    dt.press('q')
                    time.sleep(0.6)
                    dt.press('alt')
                    for dazhao in range(1, 100):
                        if zjyz[0] == 1:
                            break
                        else:
                            dt.press('t')
                            dt.press('ctrl')
                            dt.press('q')
                            dt.press('alt')
                            dt.press('t')
                            dt.press('g')
                            dt.press('e')
                            continue
                    if self.FuBen_INFO2() != 100 or zjyz[0] == 1:

                        dt.press('0')
                        time.sleep(0.7)  # 按下两秒
                        for ii in range(1, 4):
                            num_num = num // 2
                            if i == num_num or i == 10 or i == num or i == 1:
                                time.sleep(5)  # 按下两秒
                                # from python_findpicture import Caozuolei1
                                self.sellGoods_xy()  # 点击一键出售按钮
                                print('等待5')
                                time.sleep(1.5)  # 按下两秒
                                dt.press('a')
                                time.sleep(1)  # 按下两秒
                                dt.press('space')
                                time.sleep(1)
                                dt.press('left')
                                time.sleep(1)
                                dt.press('space')
                                break
                            else:
                                break
                        time.sleep(3)  # 按下两秒
                        dt.press('esc')

                        # print('0')
                        # print(i)

                        time.sleep(1)

                        if i == num:
                            dt.press('.')
                            dt.press('f12')
                            time.sleep(0.75)
                        else:
                            dt.press('f10')
                        time.sleep(3)
                        # print('方法5运行{}'.format(j))
                    else:
                        j = 9
                        print(9)
                        self.youjian()
                        time.sleep(600)
                        print(10)
                        self.movingfigur_right(2.5)  # 向右移动， 移动8秒，

                        # c.movingfigur_up(0.5)  # 向上移动， 移动0.15秒，
                        self.KeyPress1(190)  # 案件‘.’建，功能是隐藏技能和血功能
                        time.sleep(0.01)  # 睡眠0.5秒
                        dt.press('up')
                        self.FuBen_INFO11()
                        time.sleep(1.5)  # 睡眠1.5秒
                        dt.press('space')  # 单击空格操作
                        time.sleep(0.45)  # 睡眠1.5秒gdf
                    time.sleep(0.5)
                    # print('大循环{}'.format(j))
            sss = i * 6
            if i == num and Restart_computer == 0:
                # os.system("shutdown -s -t 30")
                b = time.strftime("%y-%m-%d_%H$%M$%S", time.localtime())

                img = pyautogui.screenshot(region=[48, 84, 848, 684])
                img = Image.fromarray(np.uint8(img))
                # img.save('D:\webdriver_new\lw\Tpshot{}s.png'.format(b))
                time.sleep(1)
                print(b)
                break
            elif i == num and Restart_computer == 1:
                os.system("shutdown -s -t 30")  # 30秒关闭电脑
                break

            print('费了{}疲劳'.format(sss))
        '''t
        dt.keyDown('a')#：模拟按键按下
        time.sleep(2) #按下两秒0
        dt.keyUp('a') #：模拟按键松开time.sleep(2)
        '''

    def zhaohuan(self, num_parameter, move_seepx, move_seepy, Restart_computer_parameter, sss, aa1, bb1, cc1, dd1,
                 hh1=0.75, hh2=0.75):
        # 召唤

        for i in range(1, 60):
            time.sleep(0.5)
            num = num_parameter  # num不能是奇数 运行几次
            # move_seep = -0.52  # 57.7   气功4.2  40.8
            # move_seep1= -0.26

            move_seep = -move_seepx
            move_seep1 = -move_seepy
            # move_seep = -0.8  # 57.70
            # move_seep1 = -0.23
            m_button = 'h'
            print(move_seep)
            Restart_computer = Restart_computer_parameter  # Restart_computer为0或者1，0关闭电脑，1不关闭电脑
            for j in range(1, 9):
                # print('right开始按下{}次'.format(j))
                if j == 1:
                    global ret_values
                    ret_values = []
                    time.sleep(0.35)  # 按下两秒
                    dt.press('t')
                    t13 = Thread(target=self.FuBen_INFO3,
                                 args=(1,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t13.start()  # 开始运行t1线程
                    time.sleep(0.75)  # 按下两秒
                    dt.press('s')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('t')
                    dt.press('g')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('f')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('y')
                    # time.sleep(1.75)  # 按下两秒
                    # dt.press('y')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(2.7 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    t13.join()
                    if ret_values[0] == 1:
                        j = 8
                        i = num
                        print('测试', t13)
                        break
                        # elif self.FuBen_INFO3() == 2:
                    elif ret_values[0] == 2:
                        j = 8
                        break
                    # if self.FuBen_INFO3() == 1:
                    #     j = 9
                    #     i = num
                    #     print('测试')
                    #     break
                    else:
                        self.FuBen_INFO66()
                        #self.FuBen_INFO6()
                        dt.keyDown('down')  # ：模拟按键按下 向下
                        time.sleep(0.5 + move_seep1)
                        dt.keyUp('down')  # ：模拟按键松开按键

                        dt.press('9')

                        if self.FuBen_INFO() == 1:

                            self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 15)
                            time.sleep(1)  # 按下两秒
                            dt.keyDown('right')  # ：模拟按键按下 向下
                            print('移动')
                            time.sleep(2 + move_seep1)
                            dt.keyUp('right')  # ：模拟按键松开按键
                        else:
                            pass

                elif j == 2:

                    time.sleep(0.5)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(1 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键
                    # dt.press('s')

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.4 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    self.FuBen_INFO66()
                    # self.FuBen_INFO6()
                    dt.press('9')
                    time.sleep(0.35)
                    self.FuBen_INFO1()

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.25 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键




                elif j == 3:
                    time.sleep(0.35)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.66 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    time.sleep(0.1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.55 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('d')
                    time.sleep(0.45)

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.45 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    self.FuBen_INFO66()
                    # Caozuolei().FuBen_INFO6(
                    dt.press('9')
                    time.sleep(0.3)
                    self.FuBen_INFO1()
                    self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 15)

                    dt.press('left')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(3.85 + move_seep)
                    dt.keyUp('left')  # ：模拟按键松开按键

                    self.FuBen_INFO66()
                    #self.FuBen_INFO6()

                    self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 15)
                    dt.press('9')
                    time.sleep(0.75)

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.45 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    t3 = Thread(target=self.forxunhuanY,
                                args=(sss, aa1, bb1, cc1, dd1, hh1, hh2))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t3.start()  # 开始运行t1线程
                    time.sleep(0.56)
                    # self.forxunhuanY(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    print('<---')
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下

                    time.sleep(3.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                elif j == 4:
                    time.sleep(0.55)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒

                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('s')
                    time.sleep(0.75)
                    dt.press('g')
                    time.sleep(0.75)
                    dt.press('w')
                    time.sleep(0.85)
                    dt.press('ctrl')
                    time.sleep(0.5)

                    # time.sleep(1.35)
                    dt.press('9')
                    time.sleep(0.7)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.7 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('down')
                    self.FuBen_INFO66(8)
                    self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 15)
                    self.FuBen_INFO1()
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.4 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 5:
                    time.sleep(0.15)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.8 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.68 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.keyDown('up')  # ：模拟按键按下 向下gygygygy
                    time.sleep(1.68 + move_seep1)
                    dt.keyUp('up')  # ：模拟按键松开按键
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(1.14 + move_seep)
                    dt.keyUp('left')  # ：模拟按键松开按键
                    dt.press('r')
                    self.FuBen_INFO66()
                    if self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2, 15) == 7:
                        print('结束')
                    elif self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2, 15) == 77:

                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(0.5)
                        dt.keyUp('left')  # ：模拟按键松开按键
                        time.sleep(0.56)
                    else:
                        dt.press('9')

                        self.FuBen_INFO1()

                        time.sleep(0.36)
                        dt.press('up')
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.05 + move_seep1)
                        dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 6:
                    time.sleep(0.35)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(0.6)
                    dt.keyDown('right')  # ：模拟按键按下 向下

                    time.sleep(0.8 + move_seep1)

                    dt.keyUp('up')  # ：模拟按键松开按键
                    time.sleep(0.9)
                    dt.keyUp('right')

                    t31 = Thread(target=self.FuBen_INFO66,
                                 args=(0,))  # 定义线程t2，
                    t31.start()  # 开始运行t1线程
                    dt.press('down')
                    dt.press('down')

                    # self.FuBen_INFO6(0, 6)
                    dt.press('down')
                    time.sleep(0.7)
                    dt.press('9')
                    t31.join()
                    time.sleep(0.75)
                    self.FuBen_INFO1()
                    # dt.press('9')
                    # time.sleep(0.7)
                    # self.FuBen_INFO1()

                    # dt.press('right')
                    # time.sleep(0.0075)  # 按下两秒
                    # dt.keyDown('right')  # ：模拟按键按下
                    # time.sleep(0.45 + move_seep)
                    # dt.keyUp('right')  # ：模拟按键松开按
                    self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 15)
                    if self.FuBen_INFO13() == 0:
                        dt.press('left')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(3.8 + move_seep)
                        dt.keyUp('left')  # ：模拟按键松开按键

                        time.sleep(0.5)
                        dt.press('y')
                        time.sleep(0.75)
                        dt.press('g')

                        self.FuBen_INFO66()
                        #self.FuBen_INFO6(1)
                        # dt.press('9')
                        # time.sleep(0.75)
                        self.FuBen_INFO1()
                        self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 15)

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5)
                        dt.keyUp('right')  # ：模拟按键松开按键

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(3.2 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                    else:
                        time.sleep(0.55)  # 按下两秒
                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按
                        time.sleep(1)
                        # c=1



                elif j == 7:  # 奇数 反之偶数

                    time.sleep(0.5)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(0.9 + move_seep1)
                    dt.keyUp('up')  # ：模拟按键松开按键
                    time.sleep(0.5)  # 按下两秒
                    dt.press('s')
                    time.sleep(0.7)
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.2 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键

                    time.sleep(0.5)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.7)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('g')

                    time.sleep(0.7)
                    dt.press('9')
                    time.sleep(0.7)
                    self.FuBen_INFO66(7)
                    self.FuBen_INFO1()
                    self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 15)
                    dt.press('up')
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.5 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 8:  # 奇数 反之偶数
                    global zjyz
                    zjyz = [9]  # zjyz意思为 “总经验值”
                    time.sleep(0.55)

                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.4 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键
                    time.sleep(0.1)  # 按下19秒
                    dt.press('q')
                    dt.press('q')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('g')
                    time.sleep(0.7)  # 按下19秒ff
                    # dt.press('ctrl')
                    t360 = Thread(target=self.FuBen_INFO99)  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t360.start()  # 开始运行t1线程
                    time.sleep(1.3)  # 按下19秒
                    dt.press('w')
                    if self.FuBen_INFO2(2) != 100:
                        for zz in range(1, 100):
                            if zjyz[0] == 1:
                                break
                            else:

                                time.sleep(1)  # 按下19秒

                        time.sleep(2)  # 按下两秒

                        dt.press('0')

                        time.sleep(3)  # 按下两秒
                        for ii in range(1, 4):
                            num_num = num // 2
                            if i == num_num or i == 10 or i == num or i == 1:
                                time.sleep(5)  # 按下两秒

                                self.sellGoods_xy()  # 点击一键出售按钮
                                print('等待6')
                                time.sleep(1.5)  # 按下两秒
                                dt.press('a')
                                time.sleep(1)  # 按下两秒
                                dt.press('space')
                                time.sleep(1)
                                dt.press('left')
                                time.sleep(1)
                                dt.press('space')
                                break
                            else:
                                break
                        time.sleep(3)  # 按下两秒
                        dt.press('esc')

                        # print('0')
                        # print(i)

                        time.sleep(2)

                        if i == num:
                            dt.press('.')
                            dt.press('f12')
                        else:
                            dt.press('f10')
                        time.sleep(3)
                        # print('方法5运行{}'.format(j))
                    else:
                        j = 9
                        print(9)
                        self.youjian(pvp)
                        time.sleep(600)
                        print(10)
                        self.movingfigur_right(2.5)  # 向右移动， 移动8秒，

                        # c.movingfigur_up(0.5)  # 向上移动， 移动0.15秒，
                        self.KeyPress1(190)  # 案件‘.’建，功能是隐藏技能和血功能
                        time.sleep(0.01)  # 睡眠0.5秒
                        dt.press('up')
                        self.FuBen_INFO11()
                        time.sleep(1.5)  # 睡眠1.5秒
                        dt.press('space')  # 单击空格操作
                        time.sleep(0.45)  # 睡眠1.5秒gdf
                time.sleep(0.5)
                # print('大循环{}'.format(j))
            sss = i * 6
            if i == num and Restart_computer == 0:
                # os.system("shutdown -s -t 30")
                b = time.strftime("%y-%m-%d_%H$%M$%S", time.localtime())

                img = pyautogui.screenshot(region=[48, 84, 848, 684])
                img = Image.fromarray(np.uint8(img))
                # img.save('D:\webdriver_new\lw\Tpshot{}s.png'.format(b))
                time.sleep(1)
                print(b)
                break
            elif i == num and Restart_computer == 1:
                os.system("shutdown -s -t 30")  # 30秒关闭电脑
                break

            print('费了{}疲劳'.format(sss))
        '''t
        dt.keyDown('a')#：模拟按键按下
        time.sleep(2) #按下两秒0
        dt.keyUp('a') #：模拟按键松开time.sleep(2)
        '''

    def excelboot01(self, nn):  # 参数是传几取列表（nn-1）行的数据
        zzzzz1 = []
        # data =openpyxl.load_workbook(wu)
        # data = openpyxl.load_workbook(r"C:\Users\light\webdriver_new\lw\game_name.xlsx")
        data = openpyxl.load_workbook(r"C:\Users\Administrator\webdriver_new\lw\game_name.xlsx")  # 只有C盘的用这个
        # data = openpyxl.load_workbook(r"D:\webdriver_new\lw\game_name.xlsx")
        # data = openpyxl.load_workbook(r"C:\Users\light\webdriver_new\lw\game_name.xlsx")
        # data = openpyxl.load_workbook(r"C:\Users\Administrator\webdriver_new\lw\game_name.xlsx") #只有C盘的用这个
        # data = openpyxl.load_workbook(r"D:\webdriver_new\lw\game_name.xlsx")
        # r"C:\Users\light\webdriver_new\lw\game_name.xlsx"
        # r"D:\webdriver_new\lw\game_name.xlsx"
        # 获取工作表 有三种方法
        zz1 = data.active  # 不知道表名称的 用这种
        # zz1=data['Sheet'] #知道表面的用第二种
        # zz1=data.get_sheet_by_name('Sheet1') #第三种不知道和第二种有什么区别
        # 获取表A1数据
        # z_max=zz1.max_row #获取最大行数
        # z_min=zz1.min_row #获取最小行数f'y
        # c_min=zz1.max_column #获取最大数列
        # c_min = zz1.min_column  # 获取最小数列
        # cell.coordinate 定位数据
        # 获取行数nn
        max_row_num = zz1.max_row
        row_list2 = []
        for i in range(nn + 1, max_row_num + 1):  # nn代表从哪行开始取  第一行有标题，要去掉标题党，所有nn+1才是列表的值

            for row in zz1[i]:
                row_list2.append(row.value)
        print(row_list2[0:8])
        return row_list2[2:8]  # 取数据直接截取前6条数据 并返回给调用方

    def FindStr(self, x1, y1, x2, y2, string, color_format, sim, isbackcolor, err='0'):

        try:
            # Caozuolei.mutex1.acquire()
            # time.sleep(0.05)
            ret = self.lw.FindStr(x1, y1, x2, y2, string, color_format, sim, isbackcolor)
            # Caozuolei.mutex1.release()
            if ret == 1:
                return self.lw.x(), self.lw.y()
            else:

                print('零')
                # dt.press('alt')
                return 0

        except OSError as de:
            print(de, 'FindStr', err)
            return '崩溃'
            traceback.print_exc()
        except Exception as e:
            print(e, 'FindStr', err)
            return '崩溃'

    def Find_srt(self, usr_string1, usr_color_format1, usr_string2, usr_color_format2, usr_HH1=0.75,
                 usr_HH2=0.75):  # 人物坐标
        # self.Use_Dict(0)
        # print(usr_HH1, type(usr_HH1))
        b = 0
        while True:
            b = b + 1
            # sleep(0.15)
            z = self.FindStr(
                x1=0,
                y1=0,
                x2=800,
                y2=600,
                string=usr_string1,  # "先驱者",  # "挑战者"
                color_format=usr_color_format1,  # "#422",  ##380
                sim=usr_HH1,
                isbackcolor=0,
                err='Find_srt')
            print(z, "人物坐标Find_srt<>")
            if '崩溃' in str(z):
                print('崩溃？？')
                continue
            elif z != 0 and z is not None:
                x = int(z[0]) + 68
                y = int(z[1]) + 100
                # xxyy[0:2] = x, y
                xxyy[0] = x
                xxyy[1] = y
                # xxyy[5] = 999
                # print('人物坐标{},{},门坐标{},{}'.format(xxyy[0], xxyy[1],xxyy[2],xxyy[3]))
                print('人物坐标Find_srt先驱者', z[0], '+', '68', '=', xxyy[0])
                print('人物坐标Find_srt先驱者Y轴', z[1], '+', '100', '=', xxyy[1])
                return
            elif z == 0:

                zy = self.FindStr(
                    x1=0,
                    y1=0,
                    x2=800,
                    y2=600,
                    string=usr_string2,  # "不足",
                    color_format=usr_color_format2,  # "#422",
                    sim=usr_HH2,
                    isbackcolor=0)
                if '崩溃' in str(zy):
                    print('崩溃？？')
                    continue
                elif zy != 0 and zy is not None:
                    x = zy[0]
                    y = zy[1] + 114
                    # xxyy[5] = 999
                    xxyy[0] = x
                    xxyy[1] = y
                    # print('人物坐标{},{},门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]), ',?')
                    print('人物坐标Find_srt不足', zy[0], '+', 0, '=', xxyy[0])
                    return
            #
            elif b == 60:
                # self.youjian()
                print('!!!')
                return
            elif 32 > b > 30:
                dt.press('right')
                dt.press('alt')
                continue
            # print(0, xxyy)
            #
            # print("从新找")
            # continue
            else:
                try:
                    print('不知道什么情况')
                    continue
                except OSError as de:
                    print(de, 'forxunhuan', )

                    traceback.print_exc()
                    return self
                except Exception as e:
                    print(e, 'forxunhuan')
                    traceback.print_exc()
                    return self

    def menzuobiao(self, renwuzuobiao):  # 门坐标
        # self.Use_Dict(0)
        while True:

            zuer = self.FindStr(
                x1=392,
                y1=63,
                x2=853,
                y2=574,
                string="开洞",  # "先驱者",  # "挑战者"
                color_format="#380",  # "#422",  ##380
                sim=0.8,
                isbackcolor=0,
                err='menzuobiao')
            print(zuer, 'zuer')
            if '崩溃' in str(zuer):
                print('崩溃？？')
                continue
            elif zuer != 0 and zuer is not None:
                xxyy[2] = int(zuer[0]) - 140
                xxyy[3] = int(zuer[1]) + 50 + renwuzuobiao  # 65
                xxyy[4] = 88  # 找到数据传88
                # xxyy[2:2] = x, y
                print('开洞', zuer[0], '-', 140, '=', xxyy[2])
                print('开洞yy', zuer[0], '+50+renwuzuobiao', 140, renwuzuobiao, '=', xxyy[2], '(开洞Y坐标)')
                print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                return


            elif xxyy[3] == 4 and xxyy[0] > 500:
                print('测试', xxyy[0])
                return
            elif xxyy[3] == 4 and xxyy[0] < 90:
                dt.press('right')
                time.sleep(0.0075)  # 按下两秒
                dt.keyDown('right')  # ：模拟按键按下
                time.sleep(1)
                dt.keyUp('right')  # ：模拟按键松开按键
                print("9988")
                return
            else:

                try:
                    # return
                    continue
                except OSError as de:
                    print(de, 'menzuobiao')

                    traceback.print_exc()
                    continue
                except Exception as e:
                    print(e, ' menzuobiao')

                    traceback.print_exc()
                    continue

    def menzuobiao1(self, renwuzuobiao):  # 门坐标
        # self.Use_Dict(0)
        while True:

            z = self.FindStr(
                x1=0,
                y1=158,
                x2=121,
                y2=438,
                string="开洞",  # "先驱者",  # "挑战者"
                color_format="#380",  # "#422",  ##380
                sim=0.8,
                isbackcolor=0,
                err='menzuobiao1')
            print(z, 'z')
            if '崩溃' in str(z):
                print('崩溃？？')
                continue
            elif z != 0:
                xxyy[2] = int(z[0]) - 140
                xxyy[3] = int(z[1]) + 50 + renwuzuobiao  # 65
                xxyy[4] = 88  # 找到数据传88
                # xxyy[2:2] = x, y
                print('开洞', z[0], '-', 140, '=', xxyy[2])
                print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                return
            else:
                print('测试', xxyy[0])
                if xxyy[3] == 4 and xxyy[0] > 500:

                    return
                elif xxyy[3] == 4 and xxyy[0] < 90:
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    print("9988")
                    return
                else:
                    return

    # 人物移动到地图某一个x，y坐标点
    def forxunhuanY(self, sss, aa, bb, cc, dd, hh1=0.75, hh2=0.75, renwuzuobiao=0):  # ,aa,bb,cc,dd
        # self.Use_Dict(0)
        print('--->')
        a = [1]
        num_ss = [0]
        self.Find_srt(aa, bb, cc, dd, hh1, hh2)
        # self.Find_srt("先驱者","#422", "不足", "#422")
        # # self.menzuobiao()
        # x, y, x1, y1, z, h = xxyy
        # # print('x=',x,'y=',y,x1,y1,z,h)
        # if -11 < (x-x1) < 11 and -11 < (y- y1)  < 11 and z==88 :
        #
        #     print((x-x1), 'y','外层循环')
        #     return
        # else:
        for j in range(1, 3):
            print(num_ss[0])
            self.menzuobiao1(renwuzuobiao)
            # self.Find_srt(aa1, bb1, cc1, dd1)
            # time.sleep(0.5)

            x, y, x1, y1, z, h = xxyy
            print(xxyy)
            print((x - x1, '=x - x1'), (y - y1, '=y - y1'), '----')
            if h == 1000:
                print(1000)
                break
            elif (y == 0 and y1 == 4) or (y == 2 and y1 == 4):
                print('y=', y, ' y1=', y1, '测试1')
                continue
            elif y >= 423:  # 449
                dt.press('up')
                print('y=', y, ' y1=', y1, '测试2')
                return
            elif y < 380:
                dt.press('down')
                print('y=', y, ' y1=', y1, '测试3')
                return
            # elif int(y - y1) > 60 and z == 88:  # 6
            #     # print((y- y1) , 'y2')
            #     dt.press('up')
            #     # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
            #     print('y=', y, ' y1=', (y - y1), "up,#6  int(y - y1) > 16 and z == 88 ")
            #     num_ss[0] = 0
            #     return
            #
            # elif (int(y - y1) > 22 and z == 88) and (y1!=4 or y!=2):  # 61
            #     # print((y- y1) , 'y2')
            #     # dt.press('up')
            #     # dt.press('up')
            #     dt.keyDown('up')  # ：模拟按键按下
            #     time.sleep(0.1)
            #     dt.keyUp('up')  # ：模拟按键松开按键
            #     # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
            #     print("up,#6  int(y - y1) > 16 and z == 88 ",'y分别为',y,'y1分别为',y1)
            #     print('y=', y, ' y1=', y1, '2023')
            #     num_ss[0] = 0
            #     return
            # elif (int(y - y1) < -110 and z == 88) and (y1!=4 or y!=2) :  # 5
            #     # print((y- y1), 'y1')
            #     dt.press('down')
            #     print("down,#5 int(y - y1) < -16 and z == 88 ",'y分别为',y,'y1分别为',y1)
            #     num_ss[0] = 0
            #     print('y=', y, ' y1=', y1, '测试')
            #     return
            # elif (33 > int(y - y1) > -60 and z == 88) and (y1!=4 or y!=2):  # 6
            #     # print((y- y1) , 'y2')
            #     # dt.press('up')
            #     # dt.press('down')
            #     print('不动，数据不名称，', (y - y1), '> -150 ','y分别为',y,'y1分别为',y1)
            #     return

            # elif 11 < int(y- y1) > -11 and z==88:
            #     print((y- y1) , 'y3')
            #    # continue
            # elif 11 > int(x - x1) > -11 and z==88:
            #     print((x - x1), 'x,3')
            #    # continu
            #     print((y - y1)>80)
            # elif -24 <= (x - x1) <= 24 and -16 <= (y - y1) <= 16 and z == 88:  # 7
            #     # print((x-x1), 'y？？？？？？')
            #     # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
            #     print("up,#7")
            #
            #     a.append("真")
            #     return
            else:

                print('????')
                return

    # 狂战士专用，对照forxunhuanY
    def forxunhuanY_kuang(self, sss, aa, bb, cc, dd, hh1=0.75, hh2=0.75, renwuzuobiao=0):  # ,aa,bb,cc,dd
        # self.Use_Dict(0)
        print('--->')
        a = [1]
        num_ss = [0]
        self.Find_srt(aa, bb, cc, dd, hh1, hh2)
        # self.Find_srt("先驱者","#422", "不足", "#422")
        # # self.menzuobiao()
        # x, y, x1, y1, z, h = xxyy
        # # print('x=',x,'y=',y,x1,y1,z,h)
        # if -11 < (x-x1) < 11 and -11 < (y- y1)  < 11 and z==88 :
        #
        #     print((x-x1), 'y','外层循环')
        #     return
        # else:
        for j in range(1, 3):
            # print(num_ss[0])
            self.menzuobiao1(renwuzuobiao)
            # self.Find_srt(aa1, bb1, cc1, dd1)
            # time.sleep(0.5)

            x, y, x1, y1, z, h = xxyy
            # print(xxyy)
            # print((x - x1, '=x - x1'), (y - y1, '=y - y1'), '----')
            if h == 1000:
                # print(1000)
                break
            elif (y == 0 and y1 == 4) or (y == 2 and y1 == 4):
                # print('y=', y, ' y1=', y1, '测试1')
                continue
            elif y >= 349:  # 449 #423
                dt.press('up')
                print('y=', y, ' y1=', y1, '测试2')
                return
            elif y < 380:
                dt.press('down')
                print('y=', y, ' y1=', y1, '测试3')
                return

            else:

                print('????')
                return

    # 人物移动到地图某一个x，y坐标点
    def forxunhuanYnama(self, sss, aa, bb, cc, dd, hh1=0.75, hh2=0.75, renwuzuobiao=0):  # ,aa,bb,cc,dd
        # self.Use_Dict(0)
        print('--->')
        a = [1]
        num_ss = [0]
        self.Find_srt(aa, bb, cc, dd, hh1, hh2)
        # self.Find_srt("先驱者","#422", "不足", "#422")
        # # self.menzuobiao()
        # x, y, x1, y1, z, h = xxyy
        # # print('x=',x,'y=',y,x1,y1,z,h)
        # if -11 < (x-x1) < 11 and -11 < (y- y1)  < 11 and z==88 :
        #
        #     print((x-x1), 'y','外层循环')
        #     return
        # else:
        for j in range(1, 3):
            print(num_ss[0])
            self.menzuobiao1(renwuzuobiao)
            # self.Find_srt(aa1, bb1, cc1, dd1)
            # time.sleep(0.5)

            x, y, x1, y1, z, h = xxyy
            print(xxyy)
            print((x - x1, '=x - x1'), (y - y1, '=y - y1'), '----')
            if h == 1000:
                print(1000)
                break
            elif (y == 0 and y1 == 4) or (y == 2 and y1 == 4):
                print('y=', y, ' y1=', y1, '测试1')
                continue
            elif y > 349:  # 449
                dt.press('up')
                print('y=', y, ' y1=', y1, '测试2')
                return
            elif y < 480:
                dt.press('down')
                print('y=', y, ' y1=', y1, '测试3')
                return

            else:

                print('????')
                return

    def menzuobiao2(self, renwuzuobiao):  # 开府门 颜色 #334 相似度 0.8
        # self.Use_Dict(0)                # 开府口 颜色 #388 相似度 0.75
        while True:

            z = self.FindStr(
                x1=0,
                y1=0,
                x2=250,
                y2=600,
                string="开府口",  # "先驱者",  # "挑战者"
                color_format="#388",  # "#422",  ##380
                sim=0.75,
                isbackcolor=0)
            print(z, 'z')
            if '崩溃' in str(z):
                print('崩溃？？')
                continue
            elif z != 0:
                xxyy[2] = z[0] - 140
                xxyy[3] = z[1] + 50 + renwuzuobiao  # 65
                xxyy[4] = 88  # 找到数据传88
                # xxyy[2:2] = x, y
                print('开洞X', z[0], '-', 140, '=', xxyy[2])
                print('开洞Y', z[1], '+', 50, '+', renwuzuobiao, '=', xxyy[3])
                print('人物坐标{}x,{}y,开府坐标{}x,{}y'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                return
            else:
                print('测试', xxyy[0])
                if xxyy[3] == 4 and xxyy[0] > 500:

                    return
                elif xxyy[3] == 4 and xxyy[0] < 90:
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    print("9988")
                    return
                else:
                    return

    def forxunhuanKF(self, sss, aa, bb, cc, dd, hh1, hh2, renwuzuobiao=0):  # ,aa,bb,cc,dd
        # self.Use_Dict(0)
        print('forxunhuanKF--->')
        # 内部调用方法吗，方法内部用 nonlocal

        # nonlocal gg_values

        a = 2
        num_ss = [0]
        self.Find_srt(aa, bb, cc, dd, hh1, hh2)
        # self.Find_srt("先驱者","#422", "不足", "#422")
        # # self.menzuobiao()
        # x, y, x1, y1, z, h = xxyy
        # # print('x=',x,'y=',y,x1,y1,z,h)
        # if -11 < (x-x1) < 11 and -11 < (y- y1)  < 11 and z==88 :
        #
        #     print((x-x1), 'y','外层循环')
        #     return
        # else:
        for j in range(1, 2):

            print(num_ss[0], renwuzuobiao, '等于renwuzuobiao')
            self.menzuobiao2(renwuzuobiao)
            # self.Find_srt(aa1, bb1, cc1, dd1)
            # time.sleep(0.5)

            x, y, x1, y1, z, h = xxyy
            print(xxyy)
            print((x - x1, '=x - x1'), (y - y1, '=y - y1'), '----forxunhuanKF000000')
            if h == 1000:
                print(1000)
                break
            elif int(y - y1) > 15 and z == 88:  # 6
                print(y, y1, (y - y1), 'y2', 'forxunhuanKF111')
                dt.press('up')
                # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                # print("up,#6  int(y - y1) > 16 and z == 88 ")
                num_ss[0] = 0
                print()

                gg_values[0] = 0

                return gg_values[0]

            elif int(y - y1) > 125 and z == 88:  # 6
                print(y, y1, (y - y1), 'y2', 'forxunhuanKF222')
                # dt.press('up')
                # dt.press('up')
                dt.keyDown('up')  # ：模拟按键按下
                time.sleep(0.25)
                dt.keyUp('up')  # ：模拟按键松开按键
                # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                print("up,#6  int(y - y1) > 16 and z == 88 ")
                # gg_values = 0
                gg_values[0] = 0
                return gg_values[0]
            elif int(y - y1) < -150 and z == 88:  # 6
                # print((y- y1) , 'y2')
                # dt.press('up')
                # dt.press('up')
                print('不动，数据不名称，', (y - y1), '> -150 ')
                # gg_values = 2
                gg_values[0] = 2
                return gg_values[0]

            elif int(y - y1) < -10 and z == 88:  # 5
                # print((y- y1), 'y1')
                dt.press('down')
                print("down,#5 int(y - y1) < -16 and z == 88 ")
                # gg_values = 1
                gg_values[0] = 1
                return gg_values[0]


            # elif 11 < int(y- y1) > -11 and z==88:
            #     print((y- y1) , 'y3')
            #    # continue
            # elif 11 > int(x - x1) > -11 and z==88:
            #     print((x - x1), 'x,3')
            #    # continu
            #     print((y - y1)>80)
            # elif -24 <= (x - x1) <= 24 and -16 <= (y - y1) <= 16 and z == 88:  # 7
            #     # print((x-x1), 'y？？？？？？')
            #     # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
            #     print("up,#7")
            #
            #     a.append("真")
            #     return
            else:
                gg_values[0] = 2
                print('????')
                return

    # 人物移动到地图某一个x，y坐标点
    def forxunhuan(self, sss, aa, bb, cc, dd, hh1=0.75, hh2=0.75, renwuzuobiao=0):  # ,aa,bb,cc,dd
        # self.Use_Dict(0)
        a = [1]
        num_ss = [0]
        while '真' not in a:

            print(' else #10,forxunhuan', (a))
            self.Find_srt(aa, bb, cc, dd, hh1, hh2)
            # self.Find_srt("先驱者","#422", "不足", "#422")
            # # self.menzuobiao()
            # x, y, x1, y1, z, h = xxyy
            # # print('x=',x,'y=',y,x1,y1,z,h)
            # if -11 < (x-x1) < 11 and -11 < (y- y1)  < 11 and z==88 :
            #
            #     print((x-x1), 'y','外层循环')
            #     return
            # else:
            for j in range(1, 20):
                print(num_ss[0])
                self.menzuobiao(renwuzuobiao)
                # self.Find_srt(aa1, bb1, cc1, dd1)
                # time.sleep(0.5)

                x, y, x1, y1, z, h = xxyy
                print(xxyy)
                print((x - x1), (y - y1), '----')
                if h == 1000:
                    break
                elif sss == 6 and self.FuBen_INFO() == 7:

                    return 7
                elif sss == 6 and self.FuBen_INFO() == 77:

                    return 77
                elif -24 <= (x - x1) <= 34 and num_ss[0] == 0 and z == 88:  # 00
                    num_ss[0] = 1
                    print("# 00", sss)
                    continue

                elif (x - x1) < int(-80) and z == 88 and num_ss[0] == 0:  # 0
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.1)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键
                    print('#0,人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                    # self.Find_srt("熟练者", "#422", "功师", "#422")
                    num_ss[0] = 1
                    break
                elif (int(x - x1) < -24 and z == 88 and num_ss[0] == 0) and (
                        int(y - y1) < -24 and z == 88 and num_ss[0] == 0):  # 1
                    # print((x - x1), 'x,1')
                    dt.press('right')
                    dt.press('down')
                    # time.sleep(0.1)  # 按下19秒
                    # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1],xxyy[2],xxyy[3]))
                    print("right,#11")
                    # self.Find_srt("熟练者", "#422", "功师", "#422")
                    num_ss[0] = 1
                    break

                elif int(x - x1) < -24 and z == 88 and num_ss[0] == 0:  # 1
                    # print((x - x1), 'x,1')
                    dt.press('right')
                    # time.sleep(0.1)  # 按下19秒
                    # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1],xxyy[2],xxyy[3]))
                    print("right,#1")
                    # self.Find_srt("熟练者", "#422", "功师", "#422")
                    num_ss[0] = 1
                    break
                elif int(x - x1) > 34 and z == 88 and num_ss[0] == 0:  # 2
                    # print((x - x1), 'x,2')
                    dt.press('left')
                    # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1],xxyy[2],xxyy[3]))
                    print("left,#2")
                    # self.Find_srt("熟练者", "#422", "功师", "#422")
                    num_ss[0] = 1
                    break

                elif int(y - y1) < -90 and z == 88 and num_ss[0] == 1:
                    dt.keyDown('down')  # ：模拟按键按下
                    time.sleep(0.4)  # 按下19秒
                    dt.keyUp('down')  # ：模拟按键松开按键
                    # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                    print("down,#3")
                    num_ss[0] = 0
                    break
                elif int(y - y1) > 80 and z == 88 and num_ss[0] == 1:  # 4
                    dt.keyDown('up')  # ：模拟按键按下
                    time.sleep(0.4)  # 按下19秒
                    dt.keyUp('up')  # ：模拟按键松开按键
                    print("down,#4")
                    num_ss[0] = 0
                    break
                elif int(y - y1) < -24 and z == 88 and num_ss[0] == 1:  # 5
                    # print((y- y1), 'y1')
                    dt.press('down')
                    print("down,#5")
                    num_ss[0] = 0
                    break

                elif int(y - y1) > 24 and z == 88 and num_ss[0] == 1:  # 6
                    # print((y- y1) , 'y2')
                    dt.press('up')
                    # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                    print("up,#6")
                    num_ss[0] = 0
                    break
                # elif 11 < int(y- y1) > -11 and z==88:
                #     print((y- y1) , 'y3')
                #    # continue
                # elif 11 > int(x - x1) > -11 and z==88:
                #     print((x - x1), 'x,3')
                #    # continu
                #     print((y - y1)>80)
                elif (-24 <= (x - x1) <= 34) and (-24 <= (y - y1) <= 24) and z == 88:  # 7
                    # print((x-x1), 'y？？？？？？')
                    # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                    print("up,#7")

                    a.append("真")
                    return

                elif num_ss[0] == 1:
                    num_ss[0] = 0
                    print(';;')
                    break
                elif num_ss[0] == 0:
                    num_ss[0] = 1
                    print(';')
                    break
                elif x > 600 and z == 88:  # 8
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(1)  # 按下19秒
                    dt.keyUp('left')  # ：模拟按键松开按键
                    # print((x-x1) , 'x', (y- y1), 'y','怎么回事')
                    # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                    print("up,#8")
                else:
                    try:
                        print(xxyy)
                        # self.youjian()
                        dt.press('right')

                        print("up,#100")
                        break
                    except OSError as de:
                        print(de, 'forxunhuan')

                        traceback.print_exc()
                        break
                    except Exception as e:
                        print(e, 'forxunhuan')
                        traceback.print_exc()
                        break

    def forxunhuanB(self, sss, aa, bb, cc, dd, hh1=0.75, hh2=0.75, renwuzuobiao=0):  # ,aa,bb,cc,dd
        # self.Use_Dict(0)
        a = [1]
        num_ss = [0]
        while True:

            print(' else #10,forxunhuanB', (a))
            self.Find_srt(aa, bb, cc, dd, hh1, hh2)
            # self.Find_srt("先驱者","#422", "不足", "#422")
            # # self.menzuobiao()
            # x, y, x1, y1, z, h = xxyy
            # # print('x=',x,'y=',y,x1,y1,z,h)
            # if -11 < (x-x1) < 11 and -11 < (y- y1)  < 11 and z==88 :
            #
            #     print((x-x1), 'y','外层循环')
            #     return
            # else:

            print(num_ss[0])
            self.menzuobiao(renwuzuobiao)
            # self.Find_srt(aa1, bb1, cc1, dd1)
            # time.sleep(0.5)

            x, y, x1, y1, z, h = xxyy
            print(xxyy)
            print((x - x1), (y - y1), '----')
            if h == 1000:
                print(1000)
                break
            elif sss == 6 and self.FuBen_INFO() == 7:
                print(7)
                return 7
            elif sss == 6 and self.FuBen_INFO() == 77:
                print(77)
                return 77

            elif int(y - y1) < -85 and z == 88:
                dt.keyDown('down')  # ：模拟按键按下
                time.sleep(0.4)  # 按下19秒
                dt.keyUp('down')  # ：模拟按键松开按键
                # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                print("down,#3")
                continue
            elif int(y - y1) < -58 and z == 88:
                dt.keyDown('down')  # ：模拟按键按下
                time.sleep(0.2)  # 按下19秒
                dt.keyUp('down')  # ：模拟按键松开按键
                # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                print("down,#3")
                continue
            elif int(y - y1) > 85 and z == 88:  # 4
                dt.keyDown('up')  # ：模拟按键按下
                time.sleep(0.4)  # 按下19秒
                dt.keyUp('up')  # ：模拟按键松开按键
                print("down,#4")

                continue
            elif int(y - y1) > 58 and z == 88:  # 4
                dt.keyDown('up')  # ：模拟按键按下
                time.sleep(0.2)  # 按下19秒
                dt.keyUp('up')  # ：模拟按键松开按键
                print("down,#4")

                continue
            elif int(y - y1) < -24 and z == 88:  # 5
                # print((y- y1), 'y1')
                dt.press('down')
                print("down,#5")

                continue

            elif int(y - y1) > 24 and z == 88:  # 6
                # print((y- y1) , 'y2')
                dt.press('up')
                # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                print("up,#6")

                continue
            # elif 11 < int(y- y1) > -11 and z==88:
            #     print((y- y1) , 'y3')
            #    # continue
            # elif 11 > int(x - x1) > -11 and z==88:
            #     print((x - x1), 'x,3')
            #    # continu
            #     print((y - y1)>80)
            elif (-24 <= (y - y1) <= 24) and z == 88:  # 7
                # print((x-x1), 'y？？？？？？')
                # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                print("up,#7,forxunhuan B")

                a.append("真")

                return 1
            if x > 600 and z == 88:  # 8
                # dt.keyDown('left')  # ：模拟按键按下
                # time.sleep(1)  # 按下19秒
                # dt.keyUp('left')  # ：模拟按键松开按键
                # print((x-x1) , 'x', (y- y1), 'y','怎么回事')
                # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                print("up,#8")
                continue

            else:

                try:
                    print(xxyy)
                    # self.youjian()
                    # dt.press('right')
                    # dt.press('alt')
                    print("up,#100")

                    continue
                except OSError as de:
                    print(de, 'forxunhuanb-try')
                    continue
                    traceback.print_exc()
                except Exception as e:
                    print(e, 'forxunhuanb-try')
                    continue

    def forxunhuanA(self, sss, aa, bb, cc, dd, hh1=0.75, hh2=0.75, renwuzuobiao=0):  # ,aa,bb,cc,dd
        # self.Use_Dict(0)
        a = [1]
        num_ss = [0]
        while True:

            print(' else #10,forxunhuanA', (a))
            self.Find_srt(aa, bb, cc, dd, hh1, hh2)
            # self.Find_srt("先驱者","#422", "不足", "#422")
            # # self.menzuobiao()
            # x, y, x1, y1, z, h = xxyy
            # # print('x=',x,'y=',y,x1,y1,z,h)
            # if -11 < (x-x1) < 11 and -11 < (y- y1)  < 11 and z==88 :
            #
            #     print((x-x1), 'y','外层循环')
            #     return
            # else:

            print(num_ss[0])
            self.menzuobiao(renwuzuobiao)
            # self.Find_srt(aa1, bb1, cc1, dd1)
            # time.sleep(0.5)

            x, y, x1, y1, z, h = xxyy
            # print(xxyy)
            print((x - x1), (y - y1), '----')
            if h == 1000:
                print('h', h)
                break
            elif sss == 6 and self.FuBen_INFO() == 7:

                return 7
            elif sss == 6 and self.FuBen_INFO() == 77:

                return 77
            # elif -24 <= (x - x1) <= 24 and z == 88:  # 00
            #     num_ss[0] = 1
            #     print("# 00", sss)
            #     continue

            elif (x - x1) < int(-90) and z == 88:  # 0
                dt.press('right')
                time.sleep(0.013)  # 按下两秒
                dt.keyDown('right')  # ：模拟按键按下
                time.sleep(0.075)  # 按下19秒
                dt.keyUp('right')  # ：模拟按键松开按键
                # print('#0,人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                # self.Find_srt("熟练者", "#422", "功师", "#422")

                continue
            # elif int(x - x1) > 90 and z == 88:  # 2
            #     # print((x - x1), 'x,2')
            #     dt.press('left')
            #     # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1],xxyy[2],xxyy[3]))
            #     print("left,#2")
            #     dt.press('left')
            #     time.sleep(0.013)  # 按下两秒
            #     dt.keyDown('left')  # ：模拟按键按下
            #     time.sleep(0.3)  # 按下19秒
            #     dt.keyUp('left')  # ：模拟按键松开按键
            #     continue
            elif (x - x1) < int(-60) and z == 88:  # 0

                dt.keyDown('right')  # ：模拟按键按下
                time.sleep(0.3)  # 按下两秒
                dt.keyUp('right')  # ：模拟按键松开按键
                print('#0,人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                # self.Find_srt("熟练者", "#422", "功师", "#422")

                continue


            elif int(x - x1) > 66 and z == 88:  # 2
                # print((x - x1), 'x,2')
                dt.press('left')
                # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1],xxyy[2],xxyy[3]))

                dt.press('left')
                time.sleep(0.013)  # 按下两秒
                dt.keyDown('left')  # ：模拟按键按下

                dt.keyUp('left')  # ：模拟按键松开按键
                # self.Find_srt("熟练者", "#422", "功师", "#422")
                continue
            elif int(x - x1) < -26 and z == 88:
                # print((x - x1), 'x,1')
                dt.press('right')
                # time.sleep(0.1)  # 按下19秒
                # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1],xxyy[2],xxyy[3]))
                print("right,#1")
                # self.Find_srt("熟练者", "#422", "功师", "#422")

                continue
            elif int(x - x1) > 36 and z == 88:  # 2
                # print((x - x1), 'x,2')
                dt.press('left')
                # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1],xxyy[2],xxyy[3]))
                print("left,#2")
                # self.Find_srt("熟练者", "#422", "功师", "#422")

                continue

            # elif 11 < int(y- y1) > -11 and z==88:
            #     print((y- y1) , 'y3')
            #    # continue
            # elif 11 > int(x - x1) > -11 and z==88:
            #     print((x - x1), 'x,3')
            #    # continu
            #     print((y - y1)>80)
            elif -26 <= (x - x1) <= 36 and z == 88:  # 7
                # print((x-x1), 'y？？？？？？')
                # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                print("up,#7，forxunhuanA")

                a.append("真")

                return 1
            elif x > 600 and z == 88:  # 8
                # dt.keyDown('left')  # ：模拟按键按下
                # time.sleep(1)  # 按下19秒
                # dt.keyUp('left')  # ：模拟按键松开按键
                # print((x-x1) , 'x', (y- y1), 'y','怎么回事')
                # print('人物坐标{},{},dong门坐标{},{}'.format(xxyy[0], xxyy[1], xxyy[2], xxyy[3]))
                print("up,#8")
                continue
            else:
                try:
                    print(xxyy)
                    # self.youjian()
                    # dt.press('right')
                    # dt.press('alt')
                    print("up,#100")
                    continue
                except OSError as de:
                    print(de, 'forxunhuanA-try')
                    continue
                    traceback.print_exc()
                except Exception as e:
                    print(e, 'forxunhuanA-try')
                    continue

    def kuangzhanshi2(self, num_parameter, move_seepx, moveseepx_jia, move_seepy, Restart_computer_parameter, sss, aa1,
                      bb1, cc1,
                      dd1, hh1=0.75, hh2=0.75):

        for i in range(1, 60):
            time.sleep(1)
            num = num_parameter  # num不能是奇数 运行几次
            # move_seep = -0.52  # 57.7   气功4.2  40.8
            # move_seep1 = -0.26
            moveseepx_jia1 = -moveseepx_jia
            # move_seep = 0.189  # 57.70
            # move_seep1 = -0.23
            move_seep = -move_seepx
            move_seep1 = -move_seepy
            m_button = 'h'
            print(move_seep)
            Restart_computer = Restart_computer_parameter  # Restart_computer为0或者1，0关闭电脑，1不关闭电脑
            t001 = Thread(target=self.timedaojishi, args=(pvp,))  # 定义线程t2，线程任务为5.30s倒计时，无参数
            t001.start()  # 开始运行t1线程
            print('开始')

            Restart_computer = Restart_computer_parameter  # Restart_computer为0或者1，0关闭电脑，1不关闭电脑
            for j in range(1, 9):
                # print('right开始按下{}次'.format(j))
                if j == 1:
                    global ret_values
                    ret_values = []

                    time.sleep(0.3)  # 按下两秒
                    dt.press('h')
                    t13 = Thread(target=self.FuBen_INFO3, args=(1,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t13.start()  # 开始运行t1线程
                    time.sleep(0.54)  # 按下两秒
                    dt.press('w')
                    dt.press('d')
                    time.sleep(0.54)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(2.4 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    # time.sleep(3)  # 按下两秒
                    dt.press('left')
                    print(ret_values)
                    print(xxyy)
                    t13.join()
                    if ret_values[0] == 1:
                        j = 8
                        i = num
                        print('测试', t13)
                        break
                        # elif self.FuBen_INFO3() == 2:
                    elif ret_values[0] == 2:
                        j = 8
                        break
                    # if self.FuBen_INFO3() == 1:
                    #     j = 8
                    #     i = num
                    #     print('测试')
                    #     break
                    else:
                        self.FuBen_INFO66()
                        # self.FuBen_INFO6()
                        dt.press('9')
                        # time.sleep(0.75)
                        # dt.keyDown('down')  # ：模拟按键按下 向下
                        # time.sleep(0.76 + move_seep1)
                        # dt.keyUp('down')  # ：模拟按键松开按键
                        if self.FuBen_INFO() == 1:
                            # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                            self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                            # time.sleep(1)  # 按下两秒
                            dt.keyDown('right')  # ：模拟按键按下 向下
                            time.sleep(1.5 + move_seep1)
                            dt.keyUp('right')  # ：模拟按键松开按键
                        else:
                            pass
                elif j == 2:

                    time.sleep(0.4)
                    dt.press('w')
                    time.sleep(0.3)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(1 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    dt.press('right')  # ：模拟按键按下
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.05 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('e')
                    time.sleep(0.7)
                    dt.press('f')
                    time.sleep(0.7)
                    dt.press('g')
                    time.sleep(0.7)
                    self.FuBen_INFO66()
                    # self.FuBen_INFO6()
                    dt.press('9')
                    time.sleep(0.3)
                    self.FuBen_INFO1()

                    # time.sleep(1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.2 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键




                elif j == 3:
                    time.sleep(0.3)
                    dt.keyDown('down')  # ：模拟按键按下 向下
                    time.sleep(0.68 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.85 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('s')
                    time.sleep(0.75)  # 按下两秒
                    dt.press('e')
                    time.sleep(0.75)  # 按下两秒

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.15 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    self.FuBen_INFO66()
                    dt.press('9')
                    time.sleep(0.3)
                    self.FuBen_INFO1()
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)

                    dt.press('left')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('left')  # ：模拟按键按下
                    time.sleep(3.65 + move_seep)
                    dt.keyUp('left')  # ：模拟按键松开按键

                    time.sleep(0.2)
                    dt.press('y')
                    time.sleep(0.75)
                    dt.press('f')
                    time.sleep(0.75)
                    dt.press('w')
                    self.FuBen_INFO66()
                    # self.FuBen_INFO6(1)
                    dt.press('9')
                    time.sleep(0.3)
                    self.FuBen_INFO1()
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.3 + move_seep)
                    t3 = Thread(target=self.forxunhuanY_kuang,
                                args=(sss, aa1, bb1, cc1, dd1, hh1, hh2,))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t3.start()  # 开始运行t1线程
                    dt.keyUp('right')  # ：模拟按键松开按键
                    time.sleep(0.7)
                    # self.forxunhuanY(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # print('<---')

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(3.0 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 4:
                    time.sleep(0.3)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.55 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.press('y')
                    time.sleep(0.75)
                    dt.press('f')
                    time.sleep(0.75)
                    dt.press('t')
                    time.sleep(0.7)
                    dt.press('w')

                    time.sleep(0.7)
                    dt.press('e')
                    time.sleep(0.7)  # 按下两秒
                    dt.press('g')
                    time.sleep(0.7)  # 按下两秒

                    # self.FuBen_INFO66(1)
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.15 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键
                    dt.press('down')
                    self.FuBen_INFO66(8)
                    # self.FuBen_INFO6(0, 6)
                    dt.press('9')
                    time.sleep(0.65)  # 按下两秒
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.FuBen_INFO1()
                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.3 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 5:
                    time.sleep(0.01)
                    dt.press('w')
                    time.sleep(0.7)  # 按下两秒
                    dt.keyDown('down')  # ：模拟按键按下 向下

                    time.sleep(0.75 + move_seep1)
                    dt.keyUp('down')  # ：模拟按键松开按键

                    dt.press('right')
                    time.sleep(0.0075)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.85 + moveseepx_jia1)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.7 + moveseepx_jia1)
                    dt.keyUp('up')  # ：模拟按键松开按键
                    dt.keyDown('left')  # ：模拟按键按下 向下
                    time.sleep(0.75 + move_seep1)
                    dt.keyUp('left')  # ：模拟按键松开按键
                    dt.press('g')
                    time.sleep(0.7)  # 按下两秒

                    dt.press('q')
                    time.sleep(0.3)  # 按下两秒
                    # self.FuBen_INFO66()
                    self.FuBen_INFO6()

                    if self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2) == 7:
                        print('结束')

                    elif self.forxunhuan(6, aa1, bb1, cc1, dd1, hh1, hh2) == 77:

                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(0.5)
                        dt.keyUp('left')  # ：模拟按键松开按键
                        time.sleep(0.6)

                    else:
                        dt.press('9')
                        time.sleep(0.4)
                        self.FuBen_INFO1()
                        dt.press('up')
                        dt.press('up')
                        time.sleep(0.2)
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.3 + move_seep1)
                        dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 6:
                    time.sleep(0.75)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(0.5)
                    dt.keyDown('right')  # ：模拟按键按下 向下

                    time.sleep(0.8 + move_seep1)

                    dt.keyUp('up')  # ：模拟按键松开按键
                    time.sleep(0.4)
                    dt.keyUp('right')
                    dt.press('d')

                    time.sleep(0.65)  # 按下两秒
                    dt.press('e')
                    time.sleep(0.65)  # 按下两秒

                    dt.press('right')
                    time.sleep(0.007)
                    dt.keyDown('gright')  # ：模拟按键按下
                    time.sleep(0.57 + move_seep)

                    dt.keyUp('right')  # ：模拟按键松开按
                    t31 = Thread(target=self.FuBen_INFO66,
                                 args=(0,))  # 定义线程t2，
                    t31.start()  # 开始运行t1线程
                    # self.FuBen_INFO66()

                    dt.press('down')
                    dt.press('down')
                    dt.press('down')
                    dt.press('9')
                    t31.join()
                    time.sleep(0.35)  # 按下两秒
                    self.FuBen_INFO1()
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2, 15)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 20)
                    if self.FuBen_INFO13() == 0:
                        dt.press('left')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('left')  # ：模拟按键按下
                        time.sleep(0.7)
                        dt.press('up')
                        time.sleep(2.65 + move_seep)
                        dt.keyUp('left')  # ：模拟按键松开按键

                        time.sleep(0.5)
                        dt.press('y')
                        time.sleep(0.75)
                        dt.press('d')
                        time.sleep(0.75)
                        dt.press('f')
                        # self.FuBen_INFO66()
                        t31 = Thread(target=self.FuBen_INFO66,
                                     args=(1,))  # 定义线程t2，
                        t31.start()  # 开始运行t1线程
                        dt.press('9')
                        t31.join()
                        time.sleep(0.75)
                        self.FuBen_INFO1()
                        self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2, -5)
                        # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2, 20)

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.3 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键

                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(2.9 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按键
                    else:
                        # Caozuolei().FuBen_INFO1()

                        # time.sleep(0.75)  # 按下两秒
                        dt.press('right')
                        time.sleep(0.0075)  # 按下两秒
                        dt.keyDown('right')  # ：模拟按键按下
                        time.sleep(1.5 + move_seep)
                        dt.keyUp('right')  # ：模拟按键松开按
                        time.sleep(0.75)  # 按下两秒
                    # c=1



                elif j == 7:  # 奇数 反之偶数

                    time.sleep(0.65)
                    dt.keyDown('up')  # ：模拟按键按下 向下
                    time.sleep(1.35 + move_seep1)
                    dt.keyUp('up')  # ：模拟按键松开按键

                    time.sleep(0.1)  # 按下两秒
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.6 + move_seep)  # 按下19秒
                    dt.press('a')
                    time.sleep(0.7 + move_seep)
                    dt.keyUp('right')  # ：模拟按键松开按键

                    time.sleep(0.5)  # 按下两秒
                    dt.press('y')
                    time.sleep(0.5)  # 按下两秒

                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(0.8 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键

                    # self.FuBen_INFO66
                    self.FuBen_INFO66(7)
                    self.forxunhuanC(sss, aa1, bb1, cc1, dd1, hh1, hh2 - 15)
                    # self.forxunhuan(sss, aa1, bb1, cc1, dd1, hh1, hh2)
                    self.FuBen_INFO1()
                    dt.press('9')
                    time.sleep(0.7)

                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.2 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键

                elif j == 8:  # 奇数 反之偶数
                    global zjyz
                    zjyz = [9]  # zjyz意思为 “总经验值”
                    time.sleep(0.65)
                    dt.press('right')
                    time.sleep(0.013)  # 按下两秒
                    dt.keyDown('right')  # ：模拟按键按下
                    time.sleep(1.35 + move_seep)  # 按下19秒
                    dt.keyUp('right')  # ：模拟按键松开按键
                    t360 = Thread(target=self.FuBen_INFO99)  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
                    t360.start()  # 开始运行t1线程

                    time.sleep(0.7)  # 按下19秒
                    dt.press('d')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('f')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('y')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('t')
                    time.sleep(0.7)  # 按下19秒
                    for dazhao in range(1, 6):
                        dt.press('e')
                        dt.press('ctrl')
                        time.sleep(0.65)  # 按下两秒
                    time.sleep(1.5)  # 按下两秒
                    dt.press('f')
                    time.sleep(0.7)  # 按下19秒
                    dt.press('d')
                    time.sleep(0.6)  # 按下19秒
                    dt.press('y')
                    time.sleep(0.8)  # 按下19秒
                    for dazhao in range(1, 30):
                        if zjyz[0] == 1:
                            break
                        else:
                            dt.press('ctrl')
                            dt.press('left')
                            dt.press('f')
                            dt.press('g')
                            continue
                        # dt.press('a')
                    time.sleep(1)  # 按下两秒
                    if self.FuBen_INFO2() != 100 or zjyz[0] == 1:
                        dt.press('0')

                        time.sleep(0.5)  # 按下两秒
                        for ii in range(1, 4):
                            num_num = num // 2
                            if i == num_num or i == 10 or i == num or i == 1:
                                time.sleep(5)  # 按下两秒
                                # from python_findpicture import Caozuolei1
                                self.sellGoods_xy()  # 点击一键出售按钮
                                print('等待4')
                                time.sleep(1.5)  # 按下两秒
                                dt.press('a')
                                time.sleep(0.7)  # 按下两秒
                                dt.press('space')
                                time.sleep(1)
                                dt.press('left')
                                time.sleep(1)
                                dt.press('space')
                                break
                            else:
                                break
                        time.sleep(3)  # 按下两秒
                        dt.press('esc')

                        # print('0')
                        # print(i)

                        time.sleep(1)

                        if i == num:
                            dt.press('.')
                            dt.press('f12')
                            time.sleep(1.5)
                        else:
                            dt.press('f10')

                        time.sleep(2.8)
                        # print('方法5运行{}'.format(j))
                    else:
                        j = 9
                        print(9)
                        self.youjian()
                        time.sleep(600)
                        print(10)
                        self.movingfigur_right(2.5)  # 向右移动， 移动8秒，

                        # c.movingfigur_up(0.5)  # 向上移动， 移动0.15秒，
                        self.KeyPress1(190)  # 案件‘.’建，功能是隐藏技能和血功能
                        time.sleep(0.01)  # 睡眠0.5秒
                        dt.press('up')
                        self.FuBen_INFO11()
                        time.sleep(1.5)  # 睡眠1.5秒
                        dt.press('space')  # 单击空格操作
                        time.sleep(0.45)  # 睡眠1.5秒gdf
                time.sleep(0.5)
                # print('大循环{}'.format(j))
            sss = i * 6
            if i == num and Restart_computer == 0:
                # os.system("shutdown -s -t 30")
                b = time.strftime("%y-%m-%d_%H$%M$%S", time.localtime())

                img = pyautogui.screenshot(region=[48, 84, 848, 684])
                img = Image.fromarray(np.uint8(img))
                # img.save('D:\webdriver_new\lw\Tpshot{}s.png'.format(b))
                time.sleep(1)
                print(b)
                break
            elif i >= num and Restart_computer == 1:
                os.system("shutdown -s -t 30")  # 30秒关闭电脑
                break

            print('费了{}疲劳'.format(sss))

    def forxunhuanC(self, sss, aa1, bb1, cc1, dd1, hh1=0.75, hh2=0.75, renwuzuobiao=0):
        print('forxunhuanC内')
        t1 = Thread(target=self.forxunhuanB,
                    args=(sss, aa1, bb1, cc1, dd1, hh1, hh2, -10))  # 定义线程t1，线程任务为调用task1函数，task1函数的参数是6
        t2 = Thread(target=self.forxunhuanA,
                    args=(sss, aa1, bb1, cc1, dd1, hh1, hh2, -10))  # 定义线程t2，线程任务为调用task2函数，task2函数无参数
        t1.start()  # 开始运行t1线程
        t2.start()  #

        t1.join()
        t2.join()
        print(t1, t2, 't1,t2')
        if t1.is_alive() or t2.is_alive():
            print("t1,t2运行结束")
            return 1
        # if t1 == 1 and t2 == 1:
        #     return 1
        else:
            try:
                print('没有找到没有找到没有找到')
                return self
            except OSError as de:
                print(de, 'forxunhuanC')
                return self
                traceback.print_exc()
            except Exception as e:
                print(e, 'forxunhuanC')
                return self
                traceback.print_exc()

    def timedaojishi(self, pvp=0):
        # Caozuolei.mutex2.acquire()
        # for i in range(330,0,-1):
        for i in range(330, 0, -1):
            aabb = self.Find_Ocr(
                x1=0,
                y1=0,
                x2=800,
                y2=300,
                color_format="#360",
                sim=0.8,
                linesign=" ",
                isbackcolor=0)

            # Caozuolei.mutex2.release()
            time.sleep(0.5)
            # print('pvp=', pvp)
            if str(aabb) is None:
                continue
            elif i == 1:
                #   print(i)
                self.youjian(pvp)
                return
            elif i > 170 and ("最后" in str(aabb) or "再次挑战" in str(aabb)):
                #  print(i, aabb, 'timedaojishi')
                # self.youjian(pvp)
                #  print(i)
                return
            elif i > 170:
                time.sleep(0.5)
                # print(i)
                continue
            elif i <= 170:
                if "最后" in str(aabb) or "再次挑战" in str(aabb):
                    # print(i, aabb, 'timedaojishi')
                    # self.youjian(pvp)
                    # print(i)
                    return
                else:
                    # print(i, aabb, "没有找到timedaojishi")
                    # print(i, aabb, 'timedaojishi')
                    # self.youjian()
                    continue
            else:
                try:
                    #  print(i)
                    self.youjian(pvp)
                #   print("查询timedaojishi%.f" % i)
                # except OSError:
                #     ...
                except OSError as de:
                    #  print(de,'timedaojishi')
                    #   print(i)
                    traceback.print_exc()
                except Exception as e:
                    #   print(e,'timedaojishi')
                    #   print(i)
                    traceback.print_exc()

                # else:
                #     ...


if __name__ == '__main__':
    global xxyy
    global gg_values

    gg_values = [10]
    # global ret_values
    # ret_values = [9]
    xxyy = [1, 2, 3, 4, 99, 6]

    # pmp=['阳1v4号','qqlight-1v1号','撒旦1v2号','心动依恋2v2号','阿斯顿1v3号','','','','','','']
    # 定位坐标[606,401,75,499,502,481,460,481,360,454]
    # c.forxunhuan(606, 401)
    # move_seep = -0.54  # 57.7   气功4.2  40.8
    # move_seep1 = -0.26
    # 睡眠不足十62.5%

    time.sleep(1.5)
    c = Caozuolei()  # 注册乐玩


    # dt.press('alt')
    print(1)
    # time.sleep(1000)

    c.lw.MoveWindow(c.hwnd, 1, 1, 0, 0)  # 移动窗口，抢两个零是xy，后两个是窗口高度和宽度，默认为0不生效
    # c.Set_Dict(1, '测试2.txt')
    # c.Set_Dict(0, 'test3.txt')
    #
    #
    # while True:
    #     c.FuBen_INFO()
    #     time.sleep(1)
    # else:
    #     pass
    c.Set_Dict(0, 'test3.txt')
    # c.Set_Dict(1, '测试2.txt')

    # c.forxunhuan("熟练者", "#422", "功师", "#422")
    # c.forxunhuan("熟练者", "#410", "师很", "#140")
    print(1)

    # c.Find_srt("熟练者", "#422", "师很", "#422")
    # time.sleep(1000)

    # c.FuBen_INFOxx(3)
    # time.sleep(1)
    # c.forxunhuan(627,462)
    # c.Find_srt("熟练者", "#422", "功师", "#422")
    time.sleep(5)

    # canshu = c.excelboot01(22)
    # n = 24
    # c.nvQiGgdfyong(n, 0.64, 0.4, 0, 0, *canshu)  # 3p 气功运气

    # time.sleep(1000)y
    # pvp = c.FuBen_INFO12()
    pvp = 3  # 1是1P 阳 |2是2p   |3p 是11011011 撒旦
    # pvp = 2 # 1是1P 阳 |2是2p   |3p 是11011011 撒旦

    print(pvp)

    # x = [[150, 290, 1], [270, 280,g'd'f'ygg'd'f 2], [380, 215, 3], [490, 215], 4, [719, 285, 5], [80, 501, 6]]
    x = [[134, 231, 1], [274, 258, 2], [412, 244, 3], [556, 247, 4], [691, 256, 5], [67, 464, 6], [204, 466, 7],
         [350, 487, 8], [450, 487, 9], [550, 487, 10]]
    for aa in range(0, 5):  # qr打图设置ddd
        # c.Set_Dict(1, '1测试q2.txt')
        # c.Set_Dict(0, 'test3n.t1xt')f
        if aa == 11 and pvp == 1:  # aa
            # if aa == 3 and pvp. == 1:# aa
            aa = 4
        # elif aa == 6 and pvp == 1:
        #     aa = 7gdfyhe
        elif aa == 13 and pvp == 2:
            aa = 14
        time.sleep(1.85)  # 选一个任务
        c.LeftClick(x[aa][0], x[aa][1])  # 441 , 310
        time.sleep(0.015)
        c.LeftClick(x[aa][0], x[aa][1])  # 单机两下鼠标左键0
        time.sleep(1.8)
        # c.KeyPress1(190)  # 案件‘.’建，功能是隐藏技能和血功能
        if pvp == 1:
            sss1 = 12  # 气功师很烂丶
            print('pvp=', pvp, 'excel=>', sss1)
        elif pvp == 2:
            sss1 = 1  # 睡眠不足十
            print('pvp=', pvp, 'excel=>', sss1)
        elif pvp == 3:
            sss1 = 22  # ll0110ll
            print('pvp=', pvp, 'excel=>', sss1)
        elif pvp == 4:
            sss1 = 30  # 快递员 气功师很水
            print('pvp=', pvp, 'excel=>', sss1)
        elif pvp == 5:
            sss1 = 40  # 能量不足了 19129832 QQ ?
            print('pvp=', pvp, 'excel=>', sss1)
        elif pvp == 6:
            sss1 = 49  # 造就师很狂 ?
            print('pvp=', pvp, 'excel=>', sss1)
        elif pvp == 7:
            sss1 = 58  # 风化雪月  色彩不足啊
            print('pvp=', pvp, 'excel=>', sss1)
        elif pvp == 8:
            sss1 = 67  # 柳杨  力气师很大
            print('pvp=', pvp, 'excel=>', sss1)
        elif pvp == 9:
            sss1 = 77  # 超越起跑线  无敌师很猛
            print('pvp=', pvp, 'excel=>', sss1)
        elif pvp == 10:
            sss1 = 96  # 智力不足吗
            print('pvp=', pvp, 'excel=>', sss1)
        else:

            print('什么都不是')
            break
            # d
        # 女气功的操作流程，从选人物到进入图，在到刷图
        time.sleep(5)  # pvp==2 是睡眠不足
        if (x[aa][2] == 5 and pvp == 3) or (x[aa][2] == 7 and pvp == 1) or (x[aa][2] == 9 and pvp == 1) \
                or (pvp == 3 and x[aa][2] == 3) \
                or (pvp == 2 and x[aa][2] == 4) \
                or (pvp == 4 and x[aa][2] == 3) \
                or (pvp == 1 and x[aa][2] == 6) \
                or (pvp == 4 and x[aa][2] == 1) \
                or (pvp == 10 and x[aa][2] == 2) \
                or (pvp == 5 and x[aa][2] == 5):
            # or (pvp == 9 and x[aa][2] == 1):

            c.movingfigur_Down(0.65)  # 向下移动，移动一秒
        else:

            c.movingfigur_Down(0.95)  # 向下移动，移动一秒

        time.sleep(0.5)  # 睡眠0.5秒
        c.movingfigur_right(7.5)  # 向右移动， 移动8秒，

        # c.movingfigur_up(0.5)  # 向上移动， 移动0.15秒，
        c.KeyPress1(190)  # 案件‘.’建，功能是隐藏技能和血功能
        time.sleep(0.01)  # 睡眠0.5秒
        dt.press('up')
        c.FuBen_INFO11()
        time.sleep(0.5)
        # time.sleep(1.5)  # 睡眠1.5秒
        # for i in range(10):
        #     time.sleep(0.25)  # 睡眠1.5秒
        #     dt.press('left')  # 向上移动， 移动0.15秒，

        time.sleep(0.5)  # 睡眠1.5秒
        dt.press('right')  # 向右移动， 移动1秒，
        dt.press('right')  # 向右移动， 移动1秒，
        dt.press('right')  # 向右移动， 移动1秒，
        dt.press('right')  # 向右移动， 移动1秒，
        dt.press('right')  # 向右移动， 移动1秒，
        dt.press('right')  # 向右移动， 移动1秒，
        dt.press('right')  # 向右移动， 移动1秒
        dt.press('right')  # 向右移动， 移动1秒，
        # if (x[aa][2] == 4 and pvp == 2) or (pvp == 3 and x[aa][2] == 11) or (pvp == 2 and x[aa][2] == 8):
        if (pvp == 3 and x[aa][2] == 12) or (pvp == 2 and x[aa][2] == 12):

            dt.press('left')  # 向右移动， 移动1秒，
            pass
        elif x[aa][2] == 12 and pvp == 2:
            dt.press('right')  # 向右移动， 移动1秒，
            # dt.press('right')  # 向右移动， 移动1秒，
            # dt.press('right')  # 向右移动， 移动1秒，
        # elif x[aa][2] == 5 and pvp == 2 or (x[aa][2] == 7 and pvp == 2):  # or  (x[aa][2] == 3 and pvp ==2)
        # elif (x[aa][2] == 2 and pvp == 7) or (x[aa][2] == 3 and pvp == 7) :  # or  (x[aa][2] == 3 and pvp ==2)
        #
        #     dt.press('left')  # 向右移动， 移动1秒，
        #     dt.press('left')  # 向右移动， 移动1秒，
        #     dt.press('left')  # 向右移动， 移动1秒，
        # elif (x[aa][2] == 2 and pvp == 8) or (x[aa][2] == 3 and pvp == 8) :  # or  (x[aa][2] == 3 and pvp ==2)
        #     dt.press('left')  # 向右移动， 移动1秒，
        #     dt.press('left')  # 向右移动， 移动1秒，
        #     dt.press('left')  # 向右移动， 移动1秒，
        else:
            dt.press('right')  # 向右移动， 移动1秒，
        # dt.press('right')  # 向右移动， 移动1秒，
        # dt.press('right')  # 向右移动， 移动1秒，

        time.sleep(1.7)  # 睡眠1.5秒

        dt.press('space')  # 单击空格操作
        time.sleep(0.5)  # 睡眠1.5秒gdf
        # aa1, bb1, cc1, dd1 = "挑战者", "#360", "师很", "#140"
        canshu = c.excelboot01(aa + sss1)  # 调用excel表数据取值# 2p是加1 1p是加7  3p是 14
        n = 60
        print(aa, sss1, '    sss')
        if aa == 0:
            if pvp == 1:
                c.nvQiGong(n, 0.62, 0.26, 0, 0, *canshu)  # 1p 气功师很烂丶

            elif pvp == 2:

                c.nvQiGong(n, 0.65, 0.26, 0, 0, *canshu)  # 2p 睡眠不足十
                # aa1, bb1, cc1, dd1 = "先驱者", "#422", "师很", "#140"
                # aa1, bb1, cc1, dd1="熟练者", "#410", "师很", "#140"
                # c.nvQiGong(n, 0.54, 0.26, 0)  # 气功师该加强了

            elif pvp == 3:
                c.kuangzhanshi2(n, 0.53, 0.25, 0.26, 0, 0, *canshu)  # 1p  阿修罗，睡眠不足♂
                # c.nvQiGong(n, 0.58, 0.26, 0, 0, *canshu)  # 3p ll0110ll
            elif pvp == 4:
                c.nvQiGong(n, 0.64, 0.26, 0, 0, *canshu)  # 4p 气功师很水 快递员
            elif pvp == 5:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 5p
            elif pvp == 6:
                c.nvQiGong(n, 0.48, 0.26, 0, 0, *canshu)  # 6p 造就师很狂
                # c.nvQiGong(n, 0.59, 0.26, 0, 0, *canshu)  # 6p 造就师很狂

            elif pvp == 7:
                c.nvQiGong(n, 0.68, 0.26, 0, 0, *canshu)  # 7p 色彩不足S
            elif pvp == 8:
                c.nvQiGong(n, 0.64, 0.26, 0, 0, *canshu)  # 8p 力气师很大
            elif pvp == 9:
                c.nvQiGong(n, 0.64, 0.26, 0, 0, *canshu)  # 8p 无敌师很猛 0
            elif pvp == 10:
                c.nvQiGong(n, 0.64, 0.26, 0, 0, *canshu)  # 10p  智力不足

            else:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  #
            time.sleep(2)
            # x11, y11 = Caozuolei().left + 378,Caozuolei().top + 452,
            c.KeyPress1(27)  # 案件esc建
            # 425, 532 选人位置 424, 533
            time.sleep(3)
            # c.LeftClick(392,444)  # 单机鼠标左键
            # c.LeftClick(399, 524)  # 单机鼠标左键
            c.LeftClick(389, 514)  # 单机鼠标左键
            continue
        elif aa == 1:
            if pvp == 1:
                #
                # c.nvQiGong(n, 0.48, 0.2, 0)  # 1p 气功师很懒
                c.nvQiGong(n, 0.61, 0.26, 0, 0, *canshu)  # 1p 气功师狠烂
            elif pvp == 2:
                c.zhaohuan(n, 0.98, 0.26, 0, 0, *canshu)  # 2p 睡眠不足十
                # break
                # c.nvQiGong(n, 0.66, 0.26, 0, 0, *canshu)  # 2p ll1ll数据i
            elif pvp == 3:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 3p 气功师废了
            # elif pvp == 4:
            #     break
            #     #c.nvQiGong(n, 0.48, 0.26, 0, 0, *canshu)  # 4p 快递员
            elif pvp == 4:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 4p 气功师很丶
            elif pvp == 5:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 5p 史上最菜气功
                # break
            elif pvp == 6:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 6p 狗头师很差

            elif pvp == 7:
                c.nvQiGong(n, 0.58, 0.26, 0, 0, *canshu)  # 7p 炼狱不足钢
            elif pvp == 8:
                c.nvQiGong(n, 0.64, 0.26, 0, 0, *canshu)  # 8p 冒险师很帅
            elif pvp == 9:
                c.nvQiGong(n, 0.64, 0.26, 0, 0, *canshu)  # 8p 乐乐师很猛
            elif pvp == 10:
                c.nvQiGong(n, 0.64, 0.26, 0, 0, *canshu)  # 10p  智力不足

            else:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  #
                # c.nvQiGong(n, 0.48, 0.26, 0)  # 气功师该加强了
            time.sleep(3)
            # x11, y11 = Caozuolei().left + 378,Caozuolei().top + 452,
            c.KeyPress1(27)  # 案件esc建
            # 425, 532 选人位置 424, 533
            time.sleep(3)
            # c.LeftClick(390,500)  # 单机鼠标左键
            # c.LeftClick(399, 524)  # 单机鼠标左键
            c.LeftClick(389, 514)  # 单机鼠标左键
            continue
        elif aa == 2:
            if pvp == 1:
                c.nvQiGong(n, 0.61, 0.26, 0, 0, *canshu)  # 1p 气功师很懒
                # c.nvQiGong(n, 0.62, 0.26, 0)  # 2p 睡眠不足丶
                #
            elif pvp == 2:
                c.nvQiGong(n, 0.56, 0.3, 0, 0, *canshu)  # 2p 睡眠不足丶
            elif pvp == 3:
                c.nvQiGong(n, 0.66, 0.26, 0, 0, *canshu)  # 3p 气功运气 2023 11 16
            elif pvp == 4:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 4p 快递员
                # break
            elif pvp == 5:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 5p 能量不足？
            elif pvp == 6:
                c.nvQiGong(n, 0.59, 0.26, 0, 0, *canshu)  # 6p 建筑师很牛

            elif pvp == 7:
                c.nvQiGong(n, 0.6, 0.26, 0, 0, *canshu)  # 7p 炼铁不足金
            elif pvp == 8:
                c.nvQiGong(n, 0.72, 0.26, 0, 0, *canshu)  # 8p 冒险师很坏

            elif pvp == 10:
                c.kuangzhanshi2(n, 0.45, 0.1, 0.26, 0, 0, *canshu)  # 10p  睡眠不足♂ 丫

            else:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  #
            time.sleep(3)
            # x11, y11 = Caozuolei().left + 378,Caozuolei().top + 452,
            c.KeyPress1(27)  # 案件esc建
            # 425, 532 选人位置 424, 533
            time.sleep(3)
            # c.LeftClick(392,444)  # 单机鼠标左键
            # c.LeftClick(399, 524)  # 单机鼠标左键
            c.LeftClick(389, 514)  # 单机鼠标左键
            continue
        elif aa == 3:
            if pvp == 1:

                c.kuangzhanshi(n, 0.4, 0.23, 0, 0, *canshu)  # 1p 狂战士
            elif pvp == 2:
                c.nvQiGong(n, 0.61, 0.26, 0, 0, *canshu)  # 2p 十睡眠不足
            elif pvp == 3:
                c.nvQiGong(n, 0.77, 0.26, 0, 0, *canshu)  # 3p 气功该加强了
            elif pvp == 4:
                # c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 4p l11l
                c.kuangzhanshi2(n, 0.4, 0.4, 0.26, 0, 0, *canshu)  # 1p  阿修罗
                # c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 4p l11l
            elif pvp == 6:
                c.nvQiGong(n, 0.59, 0.26, 0, 0, *canshu)  # 6p 拉梅师很开
            elif pvp == 5:
                c.kuangzhanshi2(n, 0.2, 0.4, 0.23, 0, 0, *canshu)  # 1p 狂战士
                # c.naiMa(n, 0.56, 0.26, 0, 0, *canshu) #5P 爷的圣光最叼
            elif pvp == 7:
                c.nvQiGong(n, 0.6, 0.26, 0, 0, *canshu)  # 7p 炼铁不足了
            elif pvp == 8:
                c.nvQiGong(n, 0.6, 0.26, 0, 0, *canshu)  # 8p 铸造师很广

            elif pvp == 9:
                c.nvQiGong(n, 0.6, 0.26, 0, 0, *canshu)  # 9P 雪域师很水


            else:
                c.nvQiGong(n, 0.66, 0.26, 0, 0, *canshu)  #
                # c.nvQiGong(n, 0.54, 0.32, 0)  # 气功师运气

                # c.zhaohuan(n, 0.05, 0.21, 0)  # 2p s睡眠不足s
            time.sleep(2)
            # x11, y11 = Caozuolei().left + 378,Caozuolei().top + 452,
            c.KeyPress1(27)  # 案件esc建
            # 425, 532 选人位置 424, 533
            time.sleep(3)
            # c.LeftClick(392,444)  # 单机鼠标左键
            # c.LeftClick(399, 524)  # 单机鼠标左键
            c.LeftClick(389, 514)  # 单机鼠标左键
            continue
        elif aa == 4:
            if pvp == 1:

                c.nanQiGong(n, 0, 0, *canshu)  # 1p  男气功
            elif pvp == 2:
                c.zhaohuan(n, 0.94, 0.26, 0, 0, *canshu)  # 2p 睡眠不足十
                # c.nvQiGong(n, 0.66, 0.26, 0, 0, *canshu)  # 2p SS睡眠不足SS
            elif pvp == 3:
                c.kuangzhanshi2(n, 0.45, 0.47, 0.26, 0, 0, *canshu)
                # c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 3p 气功师很送
                # break
            elif pvp == 4:
                c.nvQiGong(n, 0.5, 0.26, 0, 0, *canshu)  # 4p 气功师很水i
            elif pvp == 6:
                c.kuangzhanshi2(n, 0.4, 0.45, 0.26, 0, 0, *canshu)  # 1p 狂战士
                # c.nvQiGong(n, 0.59, 0.26, 0, 0, *canshu)  # 6p 梦想师很阔
            elif pvp == 5:
                c.naiMa(n, 0.5, 0.26, 0, 0, *canshu)  # 5P 爷的太阳最混
            elif pvp == 8:
                c.nvQiGong(n, 0.6, 0.26, 0, 0, *canshu)  # 8p 铸造师很气

            elif pvp == 7:
                c.kuangzhanshi2(n, 0.47, 0.4, 0.26, 0, 0, *canshu)  # 士
            elif pvp == 9:
                c.kuangzhanshi2(n, 0.48, 0.25, 0.26, 0, 0, *canshu)  # 1p  阿修罗，无敌师很猛丨
            else:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  #
                break
                # c.nvQiGong(n, 0.54, 0.32, 0)  # 气功师运气
                # c.nvQiGong(n, 0.54, 0.2, 0, *canshu)  # 2p 十睡眠不足

            time.sleep(2)
            # x11, y11 = Caozuolei().left + 378,Caozuolei().top + 452,
            c.KeyPress1(27)  # 案件esc建
            # 425, 532 选人位置 424, 533
            time.sleep(3)
            # c.LeftClick(392,444)  # 单机鼠标左键
            # c.LeftClick(399, 524)  # 单机鼠标左键
            c.LeftClick(389, 514)  # 单机鼠标左键
            continue
        elif aa == 5:
            if pvp == 1:
                c.nvQiGong(n, 0.66, 0.26, 0, 0, *canshu)  # 气功师很烂
            elif pvp == 2:
                c.naiMa(n, 0.45, 0.26, 0, 0, *canshu)  # 2p 圣骑士很烂 光明骑士
                # c.nvQiGong(n, 0.54, 0.26, 0, 0, *canshu)  # 2p 睡眠不足啊
            elif pvp == 3:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 3p 气功师很水丨
            elif pvp == 4:
                c.nvQiGong(n, 0.5, 0.26, 0, 0, *canshu)  # 4p 气功师运气
            elif pvp == 6:
                c.nvQiGong(n, 0.59, 0.26, 0, 0, *canshu)  # 6p 高台师很恶

            elif pvp == 5:
                c.naiMa(n, 0.5, 0.26, 0, 0, *canshu)  # 5P 爷的曙光最混
                # c.nvQiGong(n, 0.54, 0.32, 0)  # 气功师运气
                # c.nvQiGong(n, 0.69, 0.26, 0, 0, *canshu)  # 2p SS睡眠不足SS
                # c.zhaohuan(n, 0.05, 0.21, 0)  # 2p
            elif pvp == 7:
                c.nvQiGong(n, 0.58, 0.26, 0, 0, *canshu)  # 7p 宽带不足了
                break
            elif pvp == 8:
                c.nvQiGong(n, 0.6, 0.26, 0, 0, *canshu)  # 8p 铸造师很气

            elif pvp == 9:
                c.nvQiGong(n, 0.64, 0.26, 0, 0, *canshu)  # 9P 格斗师很猛

            else:
                break
            time.sleep(2)
            # x11, y11 = Caozuolei().left + 378,Caozuolei().top + 452,
            c.KeyPress1(27)  # 案件esc建
            # 425, 532 选人位置 424, 533
            time.sleep(3)
            # c.LeftClick(392,444)  # 单机鼠标左键
            # c.LeftClick(399, 524)  # 单机鼠标左键
            c.LeftClick(389, 514)  # 单机鼠标左键
            continue
        elif aa == 6:
            if pvp == 1:

                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 1p 气功师很猛
            elif pvp == 2:
                c.kuangzhanshi2(n, 0.45, 0.4, 0.26, 0, 0, *canshu)
                # c.zhaohuan(n, 0.18, 0.24, 0, 0, *canshu)  # 2p s睡眠不足s
                # c.yuren(n, 0.5, 0.26, 0, 0, *canshu)  # 2p 睡眠不足罗
                # c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 2p 睡眠不足风
            elif pvp == 3:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 3p 睡眠不足巛
            elif pvp == 4:
                c.nvQiGong(n, 0.5, 0.26, 0, 0, *canshu)  # 4p 气功师很行
                break
            elif pvp == 6:
                c.nvQiGong(n, 0.59, 0.26, 0, 0, *canshu)  # 6p 储藏师很狂
            elif pvp == 8:
                c.nvQiGong(n, 0.6, 0.26, 0, 0, *canshu)  # 8p 铸造师很气
            elif pvp == 5:
                c.naiMa(n, 0.56, 0.26, 0, 0, *canshu)  # 5P 爷的圣光最香
            elif pvp == 9:
                c.nvQiGong(n, 0.64, 0.26, 0, 0, *canshu)  # 9P 零零师很辣

            else:
                break
            time.sleep(2)
            # x11, y11 = Caozuolei().left + 378,Caozuolei().top + 452,
            c.KeyPress1(27)  # 案件esc建
            # 425, 532 选人位置 424, 533
            time.sleep(3)
            # c.LeftClick(392,444)  # 单机鼠标左键
            # c.LeftClick(399, 524)  # 单机鼠标左键
            c.LeftClick(389, 514)  # 单机鼠标左键
            continue
        elif aa == 7:
            if pvp == 1:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 1p 气功师很屌

            elif pvp == 2:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 2p 睡眠不足啊啊
            elif pvp == 3:
                # c.naiMa(n, 0.56, 0.26, 0, 0, *canshu)  # 5P 爷的奶无敌了
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 3p 睡眠不足丨
                break
            elif pvp == 6:
                c.nvQiGong(n, 0.59, 0.26, 0, 0, *canshu)  # 6p 朱雀师很美
            elif pvp == 8:
                c.nvQiGong(n, 0.6, 0.26, 0, 0, *canshu)  # 8p 铸造师很气
            elif pvp == 4:
                c.nvQiGong(n, 0.5, 0.26, 0, 0, *canshu)  # 4p 气功师很刚
                break
            elif pvp == 5:
                c.naiMa(n, 0.56, 0.26, 0, 0, *canshu)  # 5P 爷的曙光最叼
            elif pvp == 9:
                c.nvQiGong(n, 0.64, 0.26, 0, 0, *canshu)  # 9P 可可师很说

            else:
                break
            time.sleep(2)
            # x11, y11 = Caozuolei().left + 378,Caozuolei().top + 452,
            c.KeyPress1(27)  # 案件esc建
            # 425, 532 选人位置 424, 533
            time.sleep(3)
            # c.LeftClick(392,444)  # 单机鼠标左键
            c.LeftClick(399, 524)  # 单机鼠标左键
            continue
        elif aa == 8:
            if pvp == 1:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 1p 气功师很懒
                # c.yuren(n, 0.52, 0.26, 0, 0, *canshu)  # 1p 悟空师很烂 缪斯
            elif pvp == 2:
                c.nvQiGong(n, 0.56, 0.26, 0, 0, *canshu)  # 2p 睡眠不足啊啊

            elif pvp == 6:
                c.nvQiGong(n, 0.59, 0.26, 0, 0, *canshu)  # 6p 武功师很利
                break
            elif pvp == 5:
                c.naiMa(n, 0.56, 0.26, 0, 0, *canshu)  # 5P 爷的太阳最叼
                break
            elif pvp == 8:
                c.nvQiGong(n, 0.6, 0.26, 0, 0, *canshu)  # 8p 铸造师很气
                # c.naiMa(n, 0.49, 0.26, 1, 0, *canshu)  # 2p 圣骑士很烂 光明骑士
            elif pvp == 9:
                c.nvQiGong(n, 0.64, 0.26, 0, 0, *canshu)  # 9P 帆帆师很水
                break
            else:
                break
            time.sleep(2)
            # x11, y11 = Caozuolei().left + 378,Caozuolei().top + 452,
            c.KeyPress1(27)  # 案件esc建
            # 425, 532 选人位置 424, 533
            time.sleep(3)
            # c.LeftClick(392,444)  # 单机鼠标左键
            # c.LeftClick(399, 524)  # 单机鼠标左键
            c.LeftClick(389, 514)  # 单机鼠标左键
            break
        elif aa == 9:
            if pvp == 1:
                c.naiMa(n, 0.5, 0.26, 1, 0, *canshu)  # 1p 奇特木偶 光明骑士
            elif pvp == 2:
                c.zhaohuan(n, 0.18, 0.24, 0, 0, *canshu)  # 2p s睡眠不足s
                # c.naiMa(n, 0.49, 0.26, 1, 0, *canshu)  # 2p 圣骑士很烂 光明骑士
                break
            elif pvp == 8:
                c.nvQiGong(n, 0.6, 0.26, 0, 0, *canshu)  # 8p 铸造师很气
                break
            else:
                break
            time.sleep(2)
            # x11, y11 = Caozuolei().left + 378,Caozuolei().top + 452,
            c.KeyPress1(27)  # 案件esc建
            # 425, 532 选人位置 424, 533
            time.sleep(3)
            # c.LeftClick(392,444)  # 单机鼠标左键
            # c.LeftClick(399, 524)  # 单机鼠标左键
            c.LeftClick(389, 514)  # 单机鼠标左键
            break
    '''
    time.sleep(1.5)  # 选一个任务
    c.LeftClick(x[0][0], x[0][1])  # 441, 310
    time.sleep(0.015)
    c.LeftClick(150, 290)  # 单机两下鼠标左键
    time.sleep(1.8)
    c.KeyPress1(190)  # 案件‘.’建，功能是隐藏技能和血功能

    # ----
    # 女气功的操作流程，从选人物到进入图，在到刷图
    time.sleep(5)
    c.movingfigur_Down(0.95)  # 向下移动，移动一秒

    time.sleep(0.5)  # 睡眠0.5秒
    c.movingfigur_right(9)  # 向右移动， 移动8秒，

    # c.movingfigur_up(0.5)  # 向上移动， 移动0.15秒，
    dt.press('up')
    time.sleep(1.5)  # 睡眠1.5秒
    for i in range(10):
        time.sleep(0.45)  # 睡眠1.5秒
        dt.press('left')  # 向上移动， 移动0.15秒，

    time.sleep(0.5)  # 睡眠1.5秒

    dt.press('right')  # 向右移动， 移动1秒，
    dt.press('right')  # 向右移动， 移动1秒，

    time.sleep(3)  # 睡眠1.5秒

    dt.press('space')  # 单击空格操作
    time.sleep(0.01)  # 睡眠1.5秒
    c.zhaohuan(28, 0.05, 0.26, 0)
    '''
    # c.nanQiGong(20,0)
    # c.nvQiGong(28,0.54,0.26,0)#睡眠不足数据0
    # c.zhaohuan(28, 0.05, 0.26, 1)#召唤数据i
    # c.nvQiGong(28, 0.54, 0.26, 0)  # 睡眠不足十数据
    # c.nvQiGong(18, 0.56, 0.26, 0)  # ll1ll数据i
    # c.nvQiGong(26, 0.54, 0.2, 0)  # 气功师很烂丶
    # c.nvQiGong(28, 0.48, 0.2, 0)  # 气功师狠烂
    # c.nvQiGong(8, 0.54, 0.26, 0g)  # 气功师该加强了
    # c.nvQiGong(8, 0.54, 0.32, 0)  # 气功师运气
    # c.kuangzhanshi(24,0.21,0.23,0)
    # c.nvQiGong(26, 0.54, 0.26, 0)  # 气功师不足丶

    # for i in range(1,10000):
    #     sleep(1)
    #     z=c.FindStr(x1=0,
    #               y1=0,
    #               x2=1000,
    #               y2=1000,
    #               string="不足",
    #               color_format="#422",
    #               sim=0.8,
    #               isbackcolor=0)
    #     if z!=0:
    #         x=z[0]
    #         y=z[1]+131
    #         print('x={},y={}'.format(x, y))
    #     else:
    #         print(0)

    # c.lw.MoveWindow(c.hwnd, 0, 0, 0, 0)  # 移动窗口，抢两个零是xy，后两个是窗口高度和宽度，默认为0不生效
    # for i in range(1, 10000):
    #     sleep(1)
    #     z = c.FindStr(x1=0,
    #                   y1=0,
    #                   x2=1000,
    #                   y2=1000,
    #                   string="挑战者",
    #                   color_format="#422",
    #                   sim=0.8,
    #                   isbackcolor=0)
    #     if z != 0:
    #         x = z[0] + 68
    #         y = z[1] + 135
    #
    #         print('x={},y={}'.format(x, y))
    #     else:
    #         for j in range(0, 1):
    #             sleep(1)
    #             z = c.FindStr(x1=0,
    #                           y1=0,
    #                           x2=1000,
    #                           y2=1000,
    #                           string="不足",
    #                           color_format="#422",
    #                           sim=0.8,
    #                           isbackcolor=0)
    #             if z != 0:
    #                 x = z[0]
    #                 y = z[1] + 131
    #                 print('x={},y={}'.format(x, y))
    #             else:
    #                 print(0)
    #                 break

    # c.lw.MoveTO(z[0],z[1])

    # c.Set_Dict(0, "测试2.txt")
    # Caozuolei().FuBen_INFO1()
    # c.Set_Dict(0, "测试2.txt")
    # c.kuangzhanshi(28,0)
    # c.nvQiGong(28,0.46,0.245,0)

    # if Caozuolei1().Find_Ocr(
    #     x1 = 0,
    #     y1 = 0,
    #     x2 = 1200,
    #     y2 = 1200,
    #     color_format= "#360",
    #     sim = 0.95,
    #     linesign = "五陵",
    #     isbackcolor = 0) == 1:
    #     print("识别成功！")
    # else:
    #     print("识别失败")

    # '''

    sleep(random.randint(0, 3))  # 随机睡眠一个小会儿
    c.UnBind()  # 解除绑定
